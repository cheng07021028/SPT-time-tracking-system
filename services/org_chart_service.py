from __future__ import annotations

import hashlib
import html
import json
import re
from datetime import date, datetime
from typing import Any, Iterable

import pandas as pd

from .config import PERSISTENT_DIR
from .persistent_store import load_json, log_audit, save_json


DIM_DEPARTMENT = "部門"
DIM_COURSE = "課別"
DIM_TITLE = "職稱"
DIM_MACHINE = "機型"
DIM_STAGE = "工段"
DIM_SOURCE = "人力來源"
PERSON_NAME = "姓名"
PERSON_KEY = "人員Key"
ROSTER_SOURCE = "名單來源"
MANPOWER_CLASS = "人力分類"

FALLBACK_DEPARTMENT = "製造部"
FALLBACK_COURSE = "未設定課別"
FALLBACK_MACHINE = "未設定機型"
FALLBACK_STAGE = "未設定工段"

LEADER_TITLE_KEYWORDS = ("總經理", "經理", "課長", "主任", "組長", "顧問", "特助")
TOP_LEADER_KEYWORDS = ("總經理", "經理", "顧問", "特助")
COURSE_LEADER_KEYWORDS = ("課長", "主任")
STAGE_LEADER_KEYWORDS = ("組長",)


ORG_LAYOUT_EXPORT_VERSION = "spt_org_layout_v1"
ORG_LAYOUTS_PATH = PERSISTENT_DIR / "org_chart_layouts.json"
ORG_STRUCTURE_PATH = PERSISTENT_DIR / "org_chart_structure_settings.json"

STRUCTURE_COLUMNS = [
    "刪除",
    "啟用",
    "層級",
    "架構類型",
    "顯示名稱",
    "上層",
    "課別",
    "工段清單",
    "管理職稱清單",
    "成員職稱清單",
    "排序",
]



def _layout_record_id(year: object, name: object) -> str:
    raw = f"{_safe_text(year, 'all')}|{_safe_text(name, '未命名版面')}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:14]


def _layout_payload() -> dict[str, Any]:
    payload = load_json(ORG_LAYOUTS_PATH, default={}) or {}
    if not isinstance(payload, dict):
        return {"records": []}
    records = payload.get("records", [])
    if not isinstance(records, list):
        records = []
    payload["records"] = [r for r in records if isinstance(r, dict)]
    return payload


def list_org_layout_records(year: object | None = None) -> list[dict[str, Any]]:
    """List saved organization chart presentation layouts.

    These records are display-layout snapshots only. They do not replace 01/02
    authority people data. The file is saved under data/persistent so it can be
    synced by the existing GitHub persistence helper when configured.
    """
    records = _layout_payload().get("records", [])
    if year is not None:
        year_text = _safe_text(year, "")
        records = [r for r in records if _safe_text(r.get("year"), "") == year_text]
    return sorted(records, key=lambda r: _safe_text(r.get("saved_at"), ""), reverse=True)


def get_org_layout_record(record_id: str) -> dict[str, Any] | None:
    rid = _safe_text(record_id, "")
    if not rid:
        return None
    for record in _layout_payload().get("records", []):
        if _safe_text(record.get("id"), "") == rid:
            return record
    return None


def _sanitize_layout_html(layout_html: object) -> str:
    text = _safe_text(layout_html, "")
    if not text:
        raise ValueError("永久記錄內容是空白，請先在組織圖內按『複製永久記錄碼』。")
    if len(text) > 2_000_000:
        raise ValueError("永久記錄內容過大，請先重置不需要的臨時卡片或線條後再匯出。")
    lower = text.lower()
    blocked = ["<script", "</script", "<iframe", "</iframe", "javascript:", " onerror=", " onload=", "<form", "</form"]
    if any(token in lower for token in blocked):
        raise ValueError("永久記錄內容包含不允許的程式碼，已拒絕儲存。")
    if "xl-org-canvas" not in text and "data-drag-type" not in text:
        raise ValueError("這不是製造部階層式組織圖的永久記錄碼。")
    return text


def parse_org_layout_export(export_text: object) -> str:
    """Parse the copy/paste export from the browser toolbar and return safe HTML."""
    raw = _safe_text(export_text, "")
    if not raw:
        raise ValueError("請貼上永久記錄碼。")
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        return _sanitize_layout_html(raw)
    if not isinstance(payload, dict):
        raise ValueError("永久記錄碼格式不正確。")
    version = _safe_text(payload.get("version"), "")
    if version and version != ORG_LAYOUT_EXPORT_VERSION:
        raise ValueError(f"永久記錄碼版本不相容：{version}")
    return _sanitize_layout_html(payload.get("html", ""))


def save_org_layout_record(name: object, year: object, export_text: object, filters: dict[str, Any] | None = None, user: str = "streamlit") -> dict[str, Any]:
    name_text = _safe_text(name, "")
    if not name_text:
        raise ValueError("請輸入永久記錄名稱。")
    year_text = _safe_text(year, "全部")
    layout_html = parse_org_layout_export(export_text)
    record_id = _layout_record_id(year_text, name_text)
    payload = _layout_payload()
    records = [r for r in payload.get("records", []) if _safe_text(r.get("id"), "") != record_id]
    record = {
        "id": record_id,
        "name": name_text,
        "year": year_text,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "version": ORG_LAYOUT_EXPORT_VERSION,
        "filters": filters or {},
        "html": layout_html,
    }
    records.append(record)
    payload["records"] = records
    payload["saved_at"] = datetime.now().isoformat(timespec="seconds")
    save_json(ORG_LAYOUTS_PATH, payload)
    log_audit("save_org_chart_layout", {"id": record_id, "name": name_text, "year": year_text}, user=user)
    return record


def delete_org_layout_record(record_id: str, user: str = "streamlit") -> bool:
    rid = _safe_text(record_id, "")
    if not rid:
        return False
    payload = _layout_payload()
    before = len(payload.get("records", []))
    payload["records"] = [r for r in payload.get("records", []) if _safe_text(r.get("id"), "") != rid]
    if len(payload["records"]) == before:
        return False
    payload["saved_at"] = datetime.now().isoformat(timespec="seconds")
    save_json(ORG_LAYOUTS_PATH, payload)
    log_audit("delete_org_chart_layout", {"id": rid}, user=user)
    return True




def _structure_record_id(year: object) -> str:
    return _safe_text(year, "全部")


def _split_structure_values(value: object) -> list[str]:
    text = _safe_text(value, "")
    if not text:
        return []
    parts = re.split(r"[,，、/|;；\n]+", text)
    return [p.strip() for p in parts if p and p.strip()]


def _contains_any_text(value: object, keywords: Iterable[str]) -> bool:
    text = _safe_text(value, "")
    words = [w for w in keywords if _safe_text(w, "")]
    if not words:
        return True
    return any(w in text for w in words)


def _structure_payload() -> dict[str, Any]:
    payload = load_json(ORG_STRUCTURE_PATH, default={}) or {}
    if not isinstance(payload, dict):
        payload = {}
    records = payload.get("records", {})
    if not isinstance(records, dict):
        records = {}
    payload["records"] = records
    return payload


def _normalize_structure_frame(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=STRUCTURE_COLUMNS)
    out = df.copy()
    for col in STRUCTURE_COLUMNS:
        if col not in out.columns:
            if col == "啟用":
                out[col] = True
            elif col == "刪除":
                out[col] = False
            elif col in {"層級", "排序"}:
                out[col] = 0
            else:
                out[col] = ""
    out = out[STRUCTURE_COLUMNS].copy()
    out["刪除"] = out["刪除"].map(lambda v: str(v).strip().lower() in {"true", "1", "是", "刪除", "yes", "y"})
    out["啟用"] = out["啟用"].map(lambda v: str(v).strip().lower() not in {"false", "0", "否", "停用", "none", "nan", ""})
    out["層級"] = pd.to_numeric(out["層級"], errors="coerce").fillna(0).astype(int)
    out["排序"] = pd.to_numeric(out["排序"], errors="coerce").fillna(999).astype(int)
    for col in [c for c in STRUCTURE_COLUMNS if c not in {"刪除", "啟用", "層級", "排序"}]:
        out[col] = out[col].map(lambda v: "" if _safe_text(v, "") == "" else _safe_text(v, ""))
    out = out[~out["刪除"]].copy()
    out = out[out["顯示名稱"].astype(str).str.strip().ne("")].copy()
    return out.sort_values(["層級", "排序", "顯示名稱"], kind="stable").reset_index(drop=True)


def default_org_structure_frame(people: pd.DataFrame | None = None) -> pd.DataFrame:
    """Build a practical default structure: 製造部 → 課別 → 組別/工段."""
    normalized = _normalize_people(people) if people is not None and not people.empty else pd.DataFrame()
    rows: list[dict[str, Any]] = [
        {
            "啟用": True,
            "層級": 1,
            "架構類型": "部門",
            "顯示名稱": "製造部",
            "上層": "",
            "課別": "",
            "工段清單": "",
            "管理職稱清單": "經理,總經理,顧問,特助",
            "成員職稱清單": "",
            "排序": 10,
        }
    ]
    if normalized.empty:
        courses = ["製一課", "製二課"]
        course_stage_map = {"製一課": ["模組", "配電", "配管", "前置鈑金", "收包機"], "製二課": ["機電", "製程", "領料", "Packing"]}
    else:
        # Build architecture from real course/stage workers. Top managers stay in the
        # 製造部 root card, and 課長/主任 stay in the course supervisor layer.
        # They must not create a fake 製造部/未設定工段 group during rebuild.
        top_mask = normalized[DIM_TITLE].map(lambda v: _contains_any(v, TOP_LEADER_KEYWORDS))
        course_source = normalized[~top_mask].copy()
        courses = [c for c in _sort_unique(course_source[DIM_COURSE].tolist(), _dept_sort_key) if c not in {FALLBACK_DEPARTMENT, "製造部"}]
        course_stage_map = {}
        for course in courses:
            cdf = course_source[course_source[DIM_COURSE].astype(str).eq(course)].copy()
            course_stage_source = cdf[~cdf[DIM_TITLE].map(lambda v: _contains_any(v, COURSE_LEADER_KEYWORDS))].copy()
            course_stage_map[course] = [s for s in _sort_unique(course_stage_source[DIM_STAGE].tolist(), _stage_sort_key) if s and s != FALLBACK_STAGE]
    order = 20
    for course in courses:
        rows.append(
            {
                "啟用": True,
                "層級": 2,
                "架構類型": "課別",
                "顯示名稱": course,
                "上層": "製造部",
                "課別": course,
                "工段清單": "",
                "管理職稱清單": "課長,主任",
                "成員職稱清單": "",
                "排序": order,
            }
        )
        order += 10
        stage_order = order
        for stage in course_stage_map.get(course, []):
            if _safe_text(stage, "") == "":
                continue
            rows.append(
                {
                    "啟用": True,
                    "層級": 3,
                    "架構類型": "組別",
                    "顯示名稱": stage,
                    "上層": course,
                    "課別": course,
                    "工段清單": stage,
                    "管理職稱清單": "組長",
                    "成員職稱清單": "技術員,助理工程師,工程師,高級工程師,資深工程師,製程工程師",
                    "排序": stage_order,
                }
            )
            stage_order += 5
        order = stage_order + 5
    return _normalize_structure_frame(pd.DataFrame(rows))



def _structure_row_key(row: pd.Series | dict[str, Any]) -> tuple[str, str, str, str]:
    """Stable key used to preserve manual structure settings during rebuild."""
    getter = row.get if hasattr(row, "get") else (lambda key, default="": default)
    return (
        _safe_text(getter("架構類型", ""), ""),
        _safe_text(getter("顯示名稱", ""), ""),
        _safe_text(getter("上層", ""), ""),
        _safe_text(getter("課別", ""), ""),
    )


def _merge_rebuilt_structure_with_existing(rebuilt: pd.DataFrame, existing: pd.DataFrame | None) -> pd.DataFrame:
    """Rebuild hierarchy from current roster while preserving saved enable/delete choices.

    Rebuild should not turn every card back on. Matching rows keep the user's saved
    啟用 state, sort, title/member rules, and manual group mapping where possible.
    New rows from current roster are enabled by default; rows the user deleted remain
    deleted because they are not present in the saved enabled structure.
    """
    rebuilt_norm = _normalize_structure_frame(rebuilt)
    existing_norm = _normalize_structure_frame(existing)
    if rebuilt_norm.empty or existing_norm.empty:
        return rebuilt_norm

    existing_by_key = {_structure_row_key(row): row for _, row in existing_norm.iterrows()}
    merged_rows: list[dict[str, Any]] = []
    for _, row in rebuilt_norm.iterrows():
        data = row.to_dict()
        old = existing_by_key.get(_structure_row_key(row))
        if old is not None:
            data["啟用"] = bool(old.get("啟用", True))
            for col in ["管理職稱清單", "成員職稱清單", "排序"]:
                if col in old and _safe_text(old.get(col, ""), "") != "":
                    data[col] = old.get(col)
        merged_rows.append(data)
    return _normalize_structure_frame(pd.DataFrame(merged_rows))

def get_org_structure_frame(year: object | None = None, people: pd.DataFrame | None = None) -> pd.DataFrame:
    payload = _structure_payload()
    year_key = _structure_record_id(year)
    records = payload.get("records", {})
    record = records.get(year_key)
    if not isinstance(record, dict):
        return default_org_structure_frame(people)
    rows = record.get("rows", [])
    if not isinstance(rows, list) or not rows:
        return default_org_structure_frame(people)
    return _normalize_structure_frame(pd.DataFrame(rows))


def save_org_structure_frame(year: object | None, structure_df: pd.DataFrame, user: str = "streamlit") -> pd.DataFrame:
    normalized = _normalize_structure_frame(structure_df)
    if normalized.empty:
        raise ValueError("架構設定不可空白，請至少保留製造部、課別或組別卡片。")
    payload = _structure_payload()
    year_key = _structure_record_id(year)
    payload["records"][year_key] = {
        "year": year_key,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "rows": normalized.to_dict(orient="records"),
    }
    payload["saved_at"] = datetime.now().isoformat(timespec="seconds")
    save_json(ORG_STRUCTURE_PATH, payload)
    log_audit("save_org_chart_structure", {"year": year_key, "rows": len(normalized)}, user=user)
    return normalized


def delete_org_structure_record(year: object | None, user: str = "streamlit") -> bool:
    payload = _structure_payload()
    year_key = _structure_record_id(year)
    if year_key not in payload.get("records", {}):
        return False
    payload["records"].pop(year_key, None)
    payload["saved_at"] = datetime.now().isoformat(timespec="seconds")
    save_json(ORG_STRUCTURE_PATH, payload)
    log_audit("delete_org_chart_structure", {"year": year_key}, user=user)
    return True


def rebuild_org_structure_frame(year: object | None, people: pd.DataFrame | None, user: str = "streamlit") -> pd.DataFrame:
    existing = get_org_structure_frame(year, people=None)
    frame = default_org_structure_frame(people)
    frame = _merge_rebuilt_structure_with_existing(frame, existing)
    return save_org_structure_frame(year, frame, user=user)

def _safe_text(value: object, default: str = "未設定") -> str:
    if value is None:
        return default
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "nat", "<na>"}:
        return default
    return text


def _slug(value: object, prefix: str = "node") -> str:
    text = _safe_text(value, prefix)
    digest = hashlib.md5(text.encode("utf-8")).hexdigest()[:10]
    return f"{prefix}_{digest}"




def _person_key_from_row(row: pd.Series) -> str:
    employee_id = _safe_text(row.get("員工編號", ""), "") if isinstance(row, pd.Series) else ""
    source = _safe_text(row.get(DIM_SOURCE, ""), "") if isinstance(row, pd.Series) else ""
    name = _safe_text(row.get(PERSON_NAME, ""), "") if isinstance(row, pd.Series) else ""
    title = _safe_text(row.get(DIM_TITLE, ""), "") if isinstance(row, pd.Series) else ""
    course = _safe_text(row.get(DIM_COURSE, ""), "") if isinstance(row, pd.Series) else ""
    stage = _safe_text(row.get(DIM_STAGE, ""), "") if isinstance(row, pd.Series) else ""
    machine = _safe_text(row.get(DIM_MACHINE, ""), "") if isinstance(row, pd.Series) else ""
    raw = employee_id or "|".join([source, name, title, course, stage, machine])
    return _slug(f"{source}|{raw}", "person")


def org_people_signature(people: pd.DataFrame) -> str:
    """Return a stable signature of current 01/02 organization data for browser layout cache busting."""
    normalized = _normalize_people(people)
    if normalized.empty:
        return "empty"
    cols = [PERSON_KEY, DIM_COURSE, DIM_STAGE, DIM_TITLE, DIM_SOURCE, DIM_MACHINE, PERSON_NAME, "是否直接人力", "可用比例"]
    payload = normalized.reindex(columns=cols, fill_value="").astype(str).sort_values(cols, kind="stable").to_csv(index=False)
    return hashlib.md5(payload.encode("utf-8")).hexdigest()[:12]

def _dept_label(value: object) -> str:
    text = _safe_text(value, FALLBACK_COURSE)
    aliases = {
        "製造一課": "製一課",
        "製造二課": "製二課",
        "制造一課": "製一課",
        "制造二課": "製二課",
    }
    return aliases.get(text, text)


def _dept_sort_key(value: object) -> tuple[int, str]:
    text = _dept_label(value)
    if "一" in text or "1" in text:
        return (1, text)
    if "二" in text or "2" in text:
        return (2, text)
    if text == FALLBACK_COURSE:
        return (99, text)
    return (50, text)


def _title_sort_key(value: object) -> tuple[int, str]:
    text = _safe_text(value, "未設定職稱")
    order = ["總經理", "經理", "副理", "顧問", "特助", "課長", "主任", "組長", "副組長", "資深", "高級", "工程師", "助理", "技術員", "派遣"]
    for idx, key in enumerate(order, start=1):
        if key in text:
            return (idx, text)
    return (80, text)


def _stage_sort_key(value: object) -> tuple[int, str]:
    text = _safe_text(value, FALLBACK_STAGE)
    order = ["NTB", "FCLP", "配電", "配線", "模組", "水平", "S.T", "ST", "前製", "鈑金", "收包機", "點料", "packing", "GPTC", "BWBS", "加工"]
    for idx, key in enumerate(order, start=1):
        if key.lower() in text.lower():
            return (idx, text)
    if text == FALLBACK_STAGE:
        return (99, text)
    return (50, text)




def _machine_sort_key(value: object) -> tuple[int, str]:
    """Stable machine/category order for the organization chart.

    The requested hierarchy is now: 製造部 → 課別 → 機型群 → 組別/工段 → 主管 → 組員.
    Keeping the machine order stable prevents cards from jumping after rerun.
    """
    text = _safe_text(value, FALLBACK_MACHINE)
    order = ["Sorter", "EFEM", "BWBS", "NTB", "GPTC", "FCLP", "Packing", "配電", "配管", "鈑金"]
    for idx, key in enumerate(order, start=1):
        if key.lower() in text.lower():
            return (idx, text)
    if text == FALLBACK_MACHINE:
        return (99, text)
    return (50, text)

def _column_first_value(df: pd.DataFrame, candidates: Iterable[str], default: str) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype="object")
    result = pd.Series([default] * len(df), index=df.index, dtype="object")
    for col in candidates:
        if col not in df.columns:
            continue
        raw = df[col].map(lambda v: _safe_text(v, ""))
        mask = raw.astype(str).str.strip().ne("") & result.astype(str).str.strip().eq(default)
        result.loc[mask] = raw.loc[mask]
        if result.astype(str).str.strip().ne(default).all():
            break
    return result.map(lambda v: _safe_text(v, default))


def _sort_unique(values: Iterable[object], key=None) -> list[str]:
    items = [_safe_text(v, "") for v in values]
    items = [x for x in items if x]
    unique = list(dict.fromkeys(items))
    return sorted(unique, key=key or (lambda x: x))


def _contains_any(value: object, keywords: Iterable[str]) -> bool:
    text = _safe_text(value, "")
    return any(k in text for k in keywords)


def build_people_frame(employees: pd.DataFrame, dispatch: pd.DataFrame) -> pd.DataFrame:
    """Combine employee and dispatch records for organization visualization.

    The formal data source remains 01/02 authority tables. This function only normalizes
    column names so the organization chart can consistently use:
    部門、課別、職稱、機型、工段、人力來源、姓名。
    """
    frames: list[pd.DataFrame] = []
    for source_name, df in [("超慧正職", employees), ("派遣/外包", dispatch)]:
        if df is None or df.empty:
            continue
        temp = df.copy()
        temp[ROSTER_SOURCE] = "01. 超慧員工名單" if source_name == "超慧正職" else "02. 派遣名單"
        temp = temp.rename(
            columns={
                "職 稱": "職稱",
                "職務": "職稱",
                "部門別": "部門",
                "部 門": "部門",
                "單位": "部門",
                "課 別": "課別",
                "姓名 ": "姓名",
                "工 段": "工段",
                "站別": "工段",
                "製程": "工段",
                "組別": "工段",
                "機種": "機型",
                "Type": "機型",
                "TYPE": "機型",
                "負責機型": "機型",
                "工作機型": "機型",
                "人員來源": "人力來源",
            }
        )
        if DIM_SOURCE not in temp.columns:
            temp[DIM_SOURCE] = source_name
        else:
            temp[DIM_SOURCE] = temp[DIM_SOURCE].map(lambda v: source_name if _safe_text(v, "") == "" else _safe_text(v, source_name))
        temp[MANPOWER_CLASS] = temp.apply(
            lambda row: _classify_manpower_source(
                row.get(ROSTER_SOURCE, ""),
                row.get(DIM_SOURCE, ""),
                row.get("人力類別", ""),
                row.get("備註", ""),
            ),
            axis=1,
        )
        frames.append(temp)

    base_columns = [DIM_DEPARTMENT, DIM_COURSE, DIM_TITLE, DIM_MACHINE, DIM_STAGE, DIM_SOURCE, ROSTER_SOURCE, MANPOWER_CLASS, PERSON_NAME, "員工編號", "到職日", "離職日", "是否直接人力", "可用比例", "啟用", PERSON_KEY]
    if not frames:
        return pd.DataFrame(columns=base_columns)

    people = pd.concat(frames, ignore_index=True)
    people[DIM_DEPARTMENT] = _column_first_value(people, [DIM_DEPARTMENT, "部門", "部門別", "部 門", "單位"], FALLBACK_DEPARTMENT)
    people[DIM_COURSE] = _column_first_value(people, [DIM_COURSE, "課別", "課 別"], FALLBACK_COURSE).map(_dept_label)
    people[DIM_TITLE] = _column_first_value(people, [DIM_TITLE, "職稱", "職 稱", "職務"], "未設定職稱")
    people[DIM_MACHINE] = _column_first_value(people, [DIM_MACHINE, "機種", "Type", "TYPE", "負責機型", "工作機型", "產品別", "Category"], FALLBACK_MACHINE)
    people[DIM_STAGE] = _column_first_value(people, [DIM_STAGE, "工段", "工 段", "製程", "站別", "組別"], FALLBACK_STAGE)

    for col, default in [(DIM_SOURCE, "未設定來源"), (ROSTER_SOURCE, "01. 超慧員工名單"), (MANPOWER_CLASS, "正職"), (PERSON_NAME, "未命名"), ("是否直接人力", "否"), ("啟用", "是")]:
        if col not in people.columns:
            people[col] = default
        people[col] = people[col].map(lambda v, d=default: _safe_text(v, d))
    if MANPOWER_CLASS in people.columns:
        people[MANPOWER_CLASS] = people.apply(
            lambda row: _classify_manpower_source(
                row.get(ROSTER_SOURCE, ""),
                row.get(DIM_SOURCE, ""),
                row.get(MANPOWER_CLASS, ""),
                row.get("備註", ""),
            ),
            axis=1,
        )
    if "員工編號" not in people.columns:
        people["員工編號"] = ""
    people["員工編號"] = people["員工編號"].map(lambda v: _safe_text(v, ""))
    if "可用比例" not in people.columns:
        people["可用比例"] = 1.0
    people["可用比例"] = pd.to_numeric(people["可用比例"], errors="coerce").fillna(0)
    active_mask = ~people["啟用"].astype(str).str.strip().isin(["否", "停用", "離職", "離場", "0", "False", "false"])
    people = people.loc[active_mask].copy()
    if not people.empty:
        people = people.loc[_employment_active_mask(people)].copy()
    people[PERSON_KEY] = people.apply(_person_key_from_row, axis=1)

    sort_cols = [DIM_DEPARTMENT, DIM_COURSE, DIM_STAGE, DIM_TITLE, DIM_SOURCE, DIM_MACHINE, PERSON_NAME]
    people = people.sort_values(sort_cols, kind="stable").reset_index(drop=True)
    return people


def _normalize_people(people: pd.DataFrame) -> pd.DataFrame:
    if people is None or people.empty:
        return pd.DataFrame(columns=[DIM_DEPARTMENT, DIM_COURSE, DIM_TITLE, DIM_MACHINE, DIM_STAGE, DIM_SOURCE, ROSTER_SOURCE, MANPOWER_CLASS, PERSON_NAME, "員工編號", "到職日", "離職日", "是否直接人力", "可用比例", "啟用", PERSON_KEY])
    people = people.copy()
    people[DIM_DEPARTMENT] = _column_first_value(people, [DIM_DEPARTMENT, "部門", "部門別", "部 門", "單位"], FALLBACK_DEPARTMENT)
    people[DIM_COURSE] = _column_first_value(people, [DIM_COURSE, "課別", "課 別"], FALLBACK_COURSE).map(_dept_label)
    people[DIM_TITLE] = _column_first_value(people, [DIM_TITLE, "職稱", "職 稱", "職務"], "未設定職稱")
    people[DIM_MACHINE] = _column_first_value(people, [DIM_MACHINE, "機種", "Type", "TYPE", "負責機型", "工作機型", "產品別", "Category"], FALLBACK_MACHINE)
    people[DIM_STAGE] = _column_first_value(people, [DIM_STAGE, "工段", "工 段", "製程", "站別", "組別"], FALLBACK_STAGE)
    for col, default in [(DIM_SOURCE, "未設定來源"), (ROSTER_SOURCE, "01. 超慧員工名單"), (MANPOWER_CLASS, "正職"), (PERSON_NAME, "未命名"), ("是否直接人力", "否"), ("啟用", "是")]:
        if col not in people.columns:
            people[col] = default
        people[col] = people[col].map(lambda v, d=default: _safe_text(v, d))
    if MANPOWER_CLASS in people.columns:
        people[MANPOWER_CLASS] = people.apply(
            lambda row: _classify_manpower_source(
                row.get(ROSTER_SOURCE, ""),
                row.get(DIM_SOURCE, ""),
                row.get(MANPOWER_CLASS, ""),
                row.get("備註", ""),
            ),
            axis=1,
        )
    if "員工編號" not in people.columns:
        people["員工編號"] = ""
    people["員工編號"] = people["員工編號"].map(lambda v: _safe_text(v, ""))
    if "可用比例" not in people.columns:
        people["可用比例"] = 0
    people["可用比例"] = pd.to_numeric(people["可用比例"], errors="coerce").fillna(0)
    people[PERSON_KEY] = people.apply(_person_key_from_row, axis=1)
    return people




def _parse_org_date_series(series: pd.Series, index: pd.Index | None = None) -> pd.Series:
    if series is None:
        return pd.Series(pd.NaT, index=index, dtype="datetime64[ns]")
    parsed = pd.to_datetime(series, errors="coerce")
    if isinstance(parsed, pd.Series):
        return parsed
    return pd.Series(parsed, index=index, dtype="datetime64[ns]")


def _employment_active_mask(df: pd.DataFrame, as_of: date | pd.Timestamp | None = None) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype=bool)
    ref = pd.Timestamp(as_of or date.today()).normalize()
    hire_col = next((c for c in ["到職日", "到職日期", "入職日", "入職日期", "到任日", "進場日", "報到日"] if c in df.columns), None)
    leave_col = next((c for c in ["離職日", "離職日期", "退職日", "退場日", "離場日"] if c in df.columns), None)
    hire = _parse_org_date_series(df[hire_col], df.index) if hire_col else pd.Series(pd.NaT, index=df.index, dtype="datetime64[ns]")
    leave = _parse_org_date_series(df[leave_col], df.index) if leave_col else pd.Series(pd.NaT, index=df.index, dtype="datetime64[ns]")
    return ((hire.isna() | hire.le(ref)) & (leave.isna() | leave.ge(ref))).fillna(False).astype(bool)


def _is_truthy_direct(value: object) -> bool:
    text = _safe_text(value, "").strip().lower()
    return text in {"是", "yes", "y", "true", "1", "直接", "dl"}


def _classify_manpower_source(roster_source: object, source: object = "", category: object = "", note: object = "") -> str:
    """Classify people for the top manpower summary.

    01 employee roster counts as regular DL/IDL unless explicitly marked as dispatch,
    outsource or field. 02 dispatch roster counts as dispatch by default because vendor
    names such as 德興、晟銘、東方 do not contain the word 派遣.
    """
    roster_text = _safe_text(roster_source, "")
    detail_text = " ".join(_safe_text(x, "") for x in [source, category, note])
    if "外場" in detail_text:
        return "外場"
    if "外包" in detail_text:
        return "外包"
    if "02." in roster_text or "派遣名單" in roster_text:
        return "派遣"
    if "派遣" in detail_text:
        return "派遣"
    return "正職"


def _regular_people(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    if MANPOWER_CLASS not in df.columns:
        return df.copy()
    return df[df[MANPOWER_CLASS].astype(str).eq("正職")].copy()


def _direct_people(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "是否直接人力" not in df.columns:
        return pd.DataFrame()
    return df[df["是否直接人力"].map(_is_truthy_direct)].copy()


def _indirect_people(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "是否直接人力" not in df.columns:
        return pd.DataFrame()
    return df[~df["是否直接人力"].map(_is_truthy_direct)].copy()


def _is_management_title(value: object) -> bool:
    return _contains_any(value, ("總經理", "經理", "顧問", "特助", "課長", "主任", "組長"))


def _summary_dl_people(df: pd.DataFrame) -> pd.DataFrame:
    direct = _direct_people(df)
    if direct.empty or DIM_TITLE not in direct.columns:
        return direct
    return direct[~direct[DIM_TITLE].map(_is_management_title)].copy()


def _summary_idl_people(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    dl_keys = set(_summary_dl_people(df).get(PERSON_KEY, pd.Series(dtype="object")).astype(str).tolist())
    if PERSON_KEY in df.columns:
        return df[~df[PERSON_KEY].astype(str).isin(dl_keys)].copy()
    return df.drop(_summary_dl_people(df).index, errors="ignore").copy()


def _count_title_bucket(df: pd.DataFrame, bucket: str) -> int:
    if df is None or df.empty or DIM_TITLE not in df.columns:
        return 0
    titles = df[DIM_TITLE].map(lambda v: _safe_text(v, ""))
    if bucket == "senior_engineer":
        mask = titles.str.contains("資深", na=False)
    elif bucket == "advanced_engineer":
        mask = titles.str.contains("高級", na=False)
    elif bucket == "assembly_engineer":
        mask = titles.str.contains("工程師", na=False) & ~titles.str.contains("資深|高級|助理|製程|設備", regex=True, na=False)
    elif bucket == "assistant_equipment_engineer":
        mask = titles.str.contains("助理", na=False) & titles.str.contains("設備|工程師", regex=True, na=False)
    elif bucket == "technician":
        mask = titles.str.contains("技術", na=False)
    elif bucket == "assistant_engineer":
        mask = titles.str.contains("助理工程師", na=False) & ~titles.str.contains("設備", na=False)
    elif bucket == "engineer_assistant":
        mask = titles.str.contains("工程師助理", na=False)
    elif bucket == "leader":
        mask = titles.str.contains("組長", na=False)
    elif bucket == "supervisor":
        mask = titles.str.contains("主任", na=False)
    elif bucket == "section_manager":
        mask = titles.str.contains("課長", na=False)
    elif bucket == "manager":
        mask = titles.str.contains("經理", na=False)
    else:
        mask = pd.Series([False] * len(titles), index=titles.index)
    return int(mask.sum())

def _direct_count(df: pd.DataFrame) -> int:
    return int(len(_direct_people(df)))


def _count_people(df: pd.DataFrame, keyword: str | None = None) -> int:
    if df is None or df.empty:
        return 0
    if not keyword:
        return int(len(df))
    if DIM_TITLE not in df.columns:
        return 0
    return int(df[DIM_TITLE].astype(str).str.contains(keyword, na=False).sum())


def _source_count(df: pd.DataFrame, keyword: str) -> int:
    if df is None or df.empty:
        return 0
    normalized = _normalize_people(df)
    if MANPOWER_CLASS not in normalized.columns:
        return 0
    return int(normalized[MANPOWER_CLASS].astype(str).eq(keyword).sum())


def _role_class(title: object, source: object = "") -> str:
    t = _safe_text(title, "")
    s = _safe_text(source, "")
    if "派遣" in t or "外包" in t or "派遣" in s or "外包" in s:
        return "role-dispatch"
    if _contains_any(t, ("經理", "課長", "主任", "組長", "顧問", "特助")):
        return "role-leader"
    if "資深" in t or "高級" in t:
        return "role-senior"
    if "助理" in t:
        return "role-assistant"
    if "技術" in t:
        return "role-tech"
    if "工程師" in t:
        return "role-engineer"
    return "role-general"


def _short_machine_values(df: pd.DataFrame, max_items: int = 3) -> str:
    if df is None or df.empty or DIM_MACHINE not in df.columns:
        return FALLBACK_MACHINE
    values = [x for x in _sort_unique(df[DIM_MACHINE].tolist()) if x != FALLBACK_MACHINE]
    if not values:
        return FALLBACK_MACHINE
    if len(values) <= max_items:
        return " / ".join(values)
    return " / ".join(values[:max_items]) + f" +{len(values) - max_items}"


def _delete_button() -> str:
    return '<button class="dim-edit" type="button" title="修改卡片內容">✎</button><button class="dim-delete" type="button" title="刪除此卡片">×</button>'


def _summary_table(people: pd.DataFrame, year: object | None = None) -> str:
    people = _normalize_people(people)
    year_text = html.escape(_safe_text(year, "2026"))

    regular_df = _regular_people(people)
    dl_df = _summary_dl_people(regular_df)
    idl_df = _summary_idl_people(regular_df)

    dl_items = [
        ("資深<br>工程師", _count_title_bucket(dl_df, "senior_engineer")),
        ("組裝<br>高級工程師", _count_title_bucket(dl_df, "advanced_engineer")),
        ("組裝<br>工程師", _count_title_bucket(dl_df, "assembly_engineer")),
        ("助理<br>設備<br>工程師", _count_title_bucket(dl_df, "assistant_equipment_engineer")),
        ("技術員", _count_title_bucket(dl_df, "technician")),
    ]
    idl_items = [
        ("助理<br>工程師", _count_title_bucket(idl_df, "assistant_engineer")),
        ("工程師<br>助理", _count_title_bucket(idl_df, "engineer_assistant")),
        ("組長", _count_title_bucket(idl_df, "leader")),
        ("主任", _count_title_bucket(idl_df, "supervisor")),
        ("課長", _count_title_bucket(idl_df, "section_manager")),
        ("經理", _count_title_bucket(idl_df, "manager")),
    ]

    dl_role_cells = "".join(f"<th>{label}</th>" for label, _ in dl_items)
    idl_role_cells = "".join(f"<th>{label}</th>" for label, _ in idl_items)
    dl_counts = "".join(f"<td>{count}</td>" for _, count in dl_items)
    idl_counts = "".join(f"<td>{count}</td>" for _, count in idl_items)
    dispatch_count = _source_count(people, "派遣")
    outsource_count = _source_count(people, "外包")
    field_count = _source_count(people, "外場")
    total_count = int(len(regular_df) + dispatch_count + outsource_count + field_count)

    return f"""
    <div class="xl-summary" draggable="true" data-drag-type="summary" data-card-id="summary_top">
      <table>
        <thead>
          <tr><th colspan="17" class="year">{year_text}年｜依目前 01/02 名單自動計算</th></tr>
          <tr><th rowspan="2">單位</th><th colspan="6">DL</th><th colspan="6">IDL</th><th rowspan="2" title="派遣：02. 派遣名單中未標示為外包或外場的啟用人員。">派遣</th><th rowspan="2" title="外包：01/02 人力來源、人力類別或備註含外包的人員。">外包</th><th rowspan="2" title="外場：01/02 人力來源、人力類別或備註含外場的人員。">外場</th><th>製造部</th></tr>
          <tr>{dl_role_cells}<th>總人數</th>{idl_role_cells}<th>Total</th></tr>
        </thead>
        <tbody>
          <tr><td>製造課</td>{dl_counts}<td class="red" title="DL總人數：01. 超慧員工名單中啟用、是否直接人力=是，且非組長/主任/課長/經理等管理職的人員；不含02派遣名單。">{len(dl_df)}</td>{idl_counts}<td>{dispatch_count}</td><td>{outsource_count}</td><td>{field_count}</td><td class="red">{total_count}</td></tr>
        </tbody>
      </table>
    </div>
    """

def _top_person_box(row: pd.Series, label: str | None = None, idx: object = "top") -> str:
    name = html.escape(_safe_text(row.get(PERSON_NAME, ""), "未設定")) if isinstance(row, pd.Series) else "未設定"
    title = html.escape(label or _safe_text(row.get(DIM_TITLE, ""), "主管")) if isinstance(row, pd.Series) else html.escape(label or "主管")
    return f'<div class="xl-top-box role-leader" draggable="true" data-drag-type="person" data-card-id="top_{idx}"><span class="xl-grip">☰</span><b>{title}</b><em>{name}</em>{_delete_button()}</div>'


def _department_manager_label(people: pd.DataFrame) -> str:
    """Return the manufacturing manager line shown inside the root department card.

    The root of the organization chart should be 製造部 with the manager attached
    to that root, not a floating yellow card above the chart. This matches the
    requested hierarchy sketch: 製造部 → 經理 → 製一課/製二課.
    """
    people = _normalize_people(people)
    manager_df = people[people[DIM_TITLE].astype(str).str.contains("經理", na=False)].copy()
    if manager_df.empty:
        return "經理｜未設定"
    names = _sort_unique(manager_df[PERSON_NAME].tolist())
    return "經理｜" + "、".join(names[:3]) + (f" +{len(names) - 3}" if len(names) > 3 else "")


def _top_management(people: pd.DataFrame) -> str:
    """Show only actual upper-level support roles above the root.

    Do not create placeholder 總經理/經理 cards. The 經理 is now rendered inside
    the 製造部 root card to avoid floating cards and preserve the fixed hierarchy.
    """
    people = _normalize_people(people)
    top_df = people[people[DIM_TITLE].map(lambda v: _contains_any(v, ("總經理", "顧問", "特助")))].copy()
    boxes: list[str] = []
    for label in ["總經理", "顧問", "特助"]:
        hit = top_df[top_df[DIM_TITLE].astype(str).str.contains(label, na=False)]
        for i, row in hit.iterrows():
            boxes.append(_top_person_box(row, label, f"{label}_{i}"))
    return "".join(boxes)


def _leader_cards(df: pd.DataFrame, keywords: Iterable[str], prefix: str, parent_card_id: str | None = None) -> str:
    if df is None or df.empty:
        return ""
    leader_df = df[df[DIM_TITLE].map(lambda v: _contains_any(v, keywords))].copy()
    if leader_df.empty:
        return ""
    cards: list[str] = []
    for i, row in leader_df.sort_values(DIM_TITLE, key=lambda s: s.map(_title_sort_key), kind="stable").iterrows():
        name = html.escape(_safe_text(row.get(PERSON_NAME, ""), "未命名"))
        title = html.escape(_safe_text(row.get(DIM_TITLE, ""), "未設定職稱"))
        parent_attr = f' data-parent-card-id="{html.escape(parent_card_id)}"' if parent_card_id else ""
        cards.append(
            f'<div class="xl-leader-card role-leader" draggable="true" data-drag-type="person" data-card-id="{prefix}_{i}"{parent_attr}>'
            f'<span class="xl-grip">☰</span><b>{title}(1員)</b><em>{name}</em>{_delete_button()}</div>'
        )
    return "".join(cards)


def _exclude_title_keywords(df: pd.DataFrame, keywords: Iterable[str]) -> pd.DataFrame:
    if df is None or df.empty or DIM_TITLE not in df.columns:
        return pd.DataFrame(columns=df.columns if df is not None else [])
    words = tuple(dict.fromkeys([_safe_text(w, "") for w in keywords if _safe_text(w, "")]))
    if not words:
        return df.copy()
    return df[~df[DIM_TITLE].map(lambda v: _contains_any(v, words))].copy()


def _dispatch_count(df: pd.DataFrame) -> int:
    return _source_count(df, "派遣")


def _person_rows(df: pd.DataFrame, parent_card_id: str | None = None) -> str:
    if df is None or df.empty:
        return '<div class="xl-empty">未配置人員</div>'
    rows: list[str] = []
    for i, row in df.sort_values(PERSON_NAME, kind="stable").iterrows():
        name = html.escape(_safe_text(row.get(PERSON_NAME, ""), "未命名"))
        title = html.escape(_safe_text(row.get(DIM_TITLE, ""), "未設定職稱"))
        raw_source = _safe_text(row.get(DIM_SOURCE, ""), "未設定來源")
        source = html.escape(raw_source)
        raw_machine = _safe_text(row.get(DIM_MACHINE, ""), FALLBACK_MACHINE)
        machine = html.escape(raw_machine)
        roster_source = _safe_text(row.get(ROSTER_SOURCE, ""), "")
        cls = _role_class(title, source)
        person_key = html.escape(_safe_text(row.get(PERSON_KEY, ""), f"person_{i}"))
        if roster_source.startswith("02."):
            detail_text = f"{raw_machine}｜{raw_source}" if raw_machine != FALLBACK_MACHINE and raw_source not in {"", "未設定來源"} else (raw_source or raw_machine or "02. 派遣名單")
        elif raw_machine != FALLBACK_MACHINE and raw_source not in {"", "未設定來源", "超慧正職"}:
            detail_text = f"{raw_machine}｜{raw_source}"
        else:
            detail_text = raw_machine if raw_machine != FALLBACK_MACHINE else raw_source
        detail = html.escape(detail_text)
        parent_attr = f' data-parent-card-id="{html.escape(parent_card_id)}"' if parent_card_id else ""
        rows.append(
            f'<div class="xl-person-row {cls}" draggable="true" data-drag-type="person" data-card-id="{person_key}" data-person-key="{person_key}"{parent_attr}>'
            f'<span class="xl-grip">⋮</span><b>{name}</b><small>{detail}</small>{_delete_button()}</div>'
        )
    return "".join(rows)


def _role_blocks(stage_df: pd.DataFrame, member_title_keywords: Iterable[str] | None = None, parent_card_id: str | None = None) -> str:
    if stage_df is None or stage_df.empty:
        return '<div class="xl-empty">此工段目前沒有資料</div>'
    blocks: list[str] = []
    non_leaders = stage_df[~stage_df[DIM_TITLE].map(lambda v: _contains_any(v, LEADER_TITLE_KEYWORDS))].copy()
    member_words = [w for w in (member_title_keywords or []) if _safe_text(w, "")]
    if member_words and not non_leaders.empty:
        matched = non_leaders[non_leaders[DIM_TITLE].map(lambda v: _contains_any_text(v, member_words))].copy()
        if not matched.empty:
            non_leaders = matched
    if non_leaders.empty:
        return '<div class="xl-empty">未配置組員</div>'
    for title in _sort_unique(non_leaders[DIM_TITLE].tolist(), _title_sort_key):
        title_df = non_leaders[non_leaders[DIM_TITLE].astype(str).eq(title)]
        for source in _sort_unique(title_df[DIM_SOURCE].tolist()):
            source_df = title_df[title_df[DIM_SOURCE].astype(str).eq(source)]
            cls = _role_class(title, source)
            block_id = _slug(f"{parent_card_id or 'root'}|{title}|{source}|{_short_machine_values(source_df)}", "role")
            source_label = source if source != "超慧正職" else ""
            parent_attr = f' data-parent-card-id="{html.escape(parent_card_id)}"' if parent_card_id else ""
            blocks.append(
                f'<article class="xl-role-block {cls}" draggable="true" data-drag-type="role" data-card-id="{block_id}"{parent_attr}>'
                f'<div class="xl-role-title"><span class="xl-grip">☰</span><b>{html.escape(title)}({len(source_df)}員)</b><em>{html.escape(source_label)}</em><button class="xl-add" type="button" data-add-type="person">＋人員</button>{_delete_button()}</div>'
                f'<div class="xl-person-zone xl-drop-zone" data-accept="person">{_person_rows(source_df, block_id)}</div>'
                f'</article>'
            )
    return "".join(blocks)


def _stage_card(stage: str, stage_df: pd.DataFrame, parent_card_id: str | None = None) -> str:
    stage_id = _slug(f"{parent_card_id or 'root'}|{stage}", "stage")
    member_df = _exclude_title_keywords(stage_df, STAGE_LEADER_KEYWORDS)
    machine_label = html.escape(_short_machine_values(stage_df))
    subtitle_parts = [f"{len(stage_df)}員", f"直接 {_direct_count(stage_df)}"]
    dispatch_count = _dispatch_count(stage_df)
    if machine_label:
        subtitle_parts.append(machine_label)
    if dispatch_count:
        subtitle_parts.append(f"派遣 {dispatch_count}")
    return f"""
    <section class="xl-stage-card" draggable="true" data-drag-type="stage" data-card-id="{stage_id}" data-parent-card-id="{html.escape(_safe_text(parent_card_id, "")) if "parent_card_id" in locals() else ""}">
      <div class="xl-stage-head"><span class="xl-grip">☰</span><b>{html.escape(stage)}</b><em>{html.escape('｜'.join(subtitle_parts))}</em><button class="xl-add" type="button" data-add-type="role">＋職稱</button>{_delete_button()}</div>
      <div class="xl-role-zone xl-drop-zone" data-accept="role">{_role_blocks(member_df, parent_card_id=stage_id)}</div>
    </section>
    """




def _configured_group_card(group_row: pd.Series, group_df: pd.DataFrame, parent_card_id: str | None = None) -> str:
    group_name = _safe_text(group_row.get("顯示名稱"), FALLBACK_STAGE)
    group_id = _slug(
        f"structure_group|{parent_card_id or 'root'}|{group_row.get('課別', '')}|{group_name}|{group_row.get('工段清單', '')}",
        "group",
    )
    manager_words = _split_structure_values(group_row.get("管理職稱清單", "")) or list(STAGE_LEADER_KEYWORDS)
    member_words = _split_structure_values(group_row.get("成員職稱清單", ""))
    member_df = _exclude_title_keywords(group_df, manager_words)
    stages = _split_structure_values(group_row.get("工段清單", ""))
    stage_text = " / ".join(stages[:4]) + (f" +{len(stages) - 4}" if len(stages) > 4 else "")
    machine_label = html.escape(_short_machine_values(group_df))
    subtitle_parts = [f"{len(group_df)}員", f"直接 {_direct_count(group_df)}"]
    dispatch_count = _dispatch_count(group_df)
    if stage_text:
        subtitle_parts.append(stage_text)
    elif machine_label:
        subtitle_parts.append(machine_label)
    if dispatch_count:
        subtitle_parts.append(f"派遣 {dispatch_count}")
    return f"""
    <section class="xl-stage-card xl-structure-group" draggable="true" data-drag-type="stage" data-card-id="{group_id}" data-parent-card-id="{html.escape(_safe_text(parent_card_id, ""))}" data-structure-group="{html.escape(group_name)}">
      <div class="xl-stage-head"><span class="xl-grip">☰</span><b>{html.escape(group_name)}</b><em>{html.escape('｜'.join(subtitle_parts))}</em><button class="xl-add" type="button" data-add-type="role">＋職稱</button>{_delete_button()}</div>
      <div class="xl-role-zone xl-drop-zone" data-accept="role">{_role_blocks(member_df, member_words, group_id)}</div>
    </section>
    """


def _machine_group_cards(machine_df: pd.DataFrame, machine_id: str, group_rows: pd.DataFrame | None = None) -> str:
    """Render the group/stage cards under one machine/category card.

    This enforces the corrected hierarchy: machine first, then groups/stages. A stage
    card is never placed directly under the course when a machine can be determined.
    """
    if machine_df is None or machine_df.empty:
        return '<div class="xl-empty">此機型目前沒有組別資料</div>'
    stages_html: list[str] = []
    assigned_stages: set[str] = set()
    active_groups = _normalize_structure_frame(group_rows) if group_rows is not None else pd.DataFrame(columns=STRUCTURE_COLUMNS)
    if not active_groups.empty:
        active_groups = active_groups[active_groups["啟用"]]
    for _, group_row in active_groups.sort_values(["排序", "顯示名稱"], kind="stable").iterrows():
        stages = _split_structure_values(group_row.get("工段清單", ""))
        if not stages:
            continue
        group_df = machine_df[machine_df[DIM_STAGE].astype(str).isin(stages)].copy()
        if group_df.empty:
            continue
        assigned_stages.update(stages)
        stages_html.append(_configured_group_card(group_row, group_df, machine_id))

    unassigned = [s for s in _sort_unique(machine_df[DIM_STAGE].tolist(), _stage_sort_key) if s not in assigned_stages]
    for stage in unassigned:
        stage_df = machine_df[machine_df[DIM_STAGE].astype(str).eq(stage)].copy()
        stages_html.append(_stage_card(stage, stage_df, machine_id))
    return "".join(stages_html) if stages_html else '<div class="xl-empty">未設定組別 / 工段</div>'


def _machine_card(machine: str, machine_df: pd.DataFrame, parent_card_id: str, group_rows: pd.DataFrame | None = None) -> str:
    machine_name = _safe_text(machine, FALLBACK_MACHINE)
    machine_id = _slug(f"{parent_card_id}|machine|{machine_name}", "machine")
    group_count = len([s for s in _sort_unique(machine_df[DIM_STAGE].tolist(), _stage_sort_key) if s and s != FALLBACK_STAGE])
    stage_cols = max(1, min(group_count or 1, 4))
    leader_html = _leader_cards(machine_df, STAGE_LEADER_KEYWORDS, f"machine_leader_{machine_id}", machine_id)
    leader_lane = f'<div class="xl-machine-leader-lane xl-drop-zone" data-accept="person">{leader_html}</div>' if leader_html else ''
    member_df = _exclude_title_keywords(machine_df, STAGE_LEADER_KEYWORDS)
    subtitle_parts = [f"{len(machine_df)}員", f"直接 {_direct_count(machine_df)}", f"{group_count}工段"]
    dispatch_count = _dispatch_count(machine_df)
    if dispatch_count:
        subtitle_parts.append(f"派遣 {dispatch_count}")
    return f"""
    <section class="xl-machine-card" draggable="true" data-drag-type="machine" data-card-id="{machine_id}" data-parent-card-id="{html.escape(parent_card_id)}">
      <div class="xl-machine-head"><span class="xl-grip">☰</span><b>{html.escape(machine_name)}</b><em>{html.escape('｜'.join(subtitle_parts))}</em><button class="xl-add" type="button" data-add-type="stage">＋組別</button>{_delete_button()}</div>
      {leader_lane}
      <div class="xl-machine-group-lane xl-drop-zone" data-accept="stage" style="--machine-stage-cols:{stage_cols};">{_machine_group_cards(member_df, machine_id, group_rows)}</div>
    </section>
    """


def _machine_lane_html(work_df: pd.DataFrame, course_id: str, group_rows: pd.DataFrame | None = None) -> str:
    if work_df is None or work_df.empty:
        return '<div class="xl-empty xl-empty-wide">未設定機型 / 組別</div>'
    machine_values = _sort_unique(work_df[DIM_MACHINE].tolist(), _machine_sort_key)
    cards: list[str] = []
    for machine in machine_values:
        machine_df = work_df[work_df[DIM_MACHINE].astype(str).eq(machine)].copy()
        cards.append(_machine_card(machine, machine_df, course_id, group_rows))
    return "".join(cards)


def _course_stage_people(course_df: pd.DataFrame) -> pd.DataFrame:
    """People that should flow into group/stage/member lanes.

    Course leaders (課長/主任) and top managers are rendered in the course/root
    layer. They must not be duplicated as a fake 未設定工段/組員 group; otherwise
    the visual hierarchy becomes 製造部 → 課別 → 未設定工段 → 課長, which is the
    incorrect behavior shown in the user's screenshots.
    """
    if course_df is None or course_df.empty or DIM_TITLE not in course_df.columns:
        return pd.DataFrame(columns=course_df.columns if course_df is not None else [])
    leader_words = tuple(dict.fromkeys(tuple(TOP_LEADER_KEYWORDS) + tuple(COURSE_LEADER_KEYWORDS)))
    return course_df[~course_df[DIM_TITLE].map(lambda v: _contains_any(v, leader_words))].copy()


def _course_column_by_structure(course: str, course_df: pd.DataFrame, course_row: pd.Series | None, group_rows: pd.DataFrame, project: bool = False) -> str:
    course_id = _slug(course, "course")
    manager_words = _split_structure_values(course_row.get("管理職稱清單", "")) if isinstance(course_row, pd.Series) else []
    leaders = _leader_cards(course_df, manager_words or COURSE_LEADER_KEYWORDS, f"course_leader_{course_id}", course_id) or '<div class="xl-empty">未設定課長 / 主任</div>'
    work_df = _course_stage_people(course_df)
    machine_html = _machine_lane_html(work_df, course_id, group_rows)
    return f"""
    <section class="xl-course {'project-course' if project else ''}" draggable="true" data-drag-type="course" data-card-id="{course_id}" data-parent-card-id="department_root">
      <div class="xl-course-count">{len(course_df)}員</div>
      <div class="xl-course-title"><span class="xl-grip">☰</span><b>{html.escape(course)}</b><em>{len(course_df)}員｜直接 {_direct_count(course_df)}</em><button class="xl-add" type="button" data-add-type="machine">＋機型</button>{_delete_button()}</div>
      <div class="xl-leader-zone xl-drop-zone" data-accept="person">{leaders}</div>
      <div class="xl-machine-lane xl-drop-zone" data-accept="machine">{machine_html}</div>
    </section>
    """


def _course_column(course: str, course_df: pd.DataFrame, project: bool = False) -> str:
    course_id = _slug(course, "course")
    leaders = _leader_cards(course_df, COURSE_LEADER_KEYWORDS, f"course_leader_{course_id}", course_id) or '<div class="xl-empty">未設定課長 / 主任</div>'
    work_df = _course_stage_people(course_df)
    machine_html = _machine_lane_html(work_df, course_id)
    return f"""
    <section class="xl-course {'project-course' if project else ''}" draggable="true" data-drag-type="course" data-card-id="{course_id}" data-parent-card-id="department_root">
      <div class="xl-course-count">{len(course_df)}員</div>
      <div class="xl-course-title"><span class="xl-grip">☰</span><b>{html.escape(course)}</b><em>{len(course_df)}員｜直接 {_direct_count(course_df)}</em><button class="xl-add" type="button" data-add-type="machine">＋機型</button>{_delete_button()}</div>
      <div class="xl-leader-zone xl-drop-zone" data-accept="person">{leaders}</div>
      <div class="xl-machine-lane xl-drop-zone" data-accept="machine">{machine_html}</div>
    </section>
    """


def _image_style_tree_inner(people: pd.DataFrame, year: object | None = None, structure_settings: pd.DataFrame | None = None) -> str:
    if people is None or people.empty:
        return '<div class="xl-empty">目前沒有組織資料。</div>'
    people = _normalize_people(people)
    structure = _normalize_structure_frame(structure_settings) if structure_settings is not None else pd.DataFrame()
    structure = structure[structure["啟用"]].copy() if not structure.empty else structure
    course_rows = pd.DataFrame()
    group_rows = pd.DataFrame()
    if not structure.empty:
        course_rows = structure[structure["架構類型"].astype(str).eq("課別")].copy()
        group_rows = structure[structure["架構類型"].astype(str).eq("組別")].copy()
    course_source = people[~people[DIM_TITLE].map(lambda v: _contains_any(v, TOP_LEADER_KEYWORDS))].copy()
    if not course_rows.empty:
        course_values = [c for c in _sort_unique(course_rows["顯示名稱"].tolist(), _dept_sort_key) if c not in {FALLBACK_DEPARTMENT, "製造部"}]
    else:
        course_values = [c for c in _sort_unique(course_source[DIM_COURSE].tolist(), _dept_sort_key) if c not in {FALLBACK_DEPARTMENT, "製造部"}]
    if not course_values:
        course_values = [FALLBACK_COURSE]
    course_html_parts: list[str] = []
    assigned_courses: set[str] = set()
    for course in course_values:
        course_row_df = course_rows[course_rows["顯示名稱"].astype(str).eq(course)] if not course_rows.empty else pd.DataFrame()
        course_row = course_row_df.iloc[0] if not course_row_df.empty else None
        course_key = _safe_text(course_row.get("課別"), course) if isinstance(course_row, pd.Series) else course
        course_df = people[people[DIM_COURSE].astype(str).eq(course_key)].copy()
        assigned_courses.add(course_key)
        project = "二" in course or "2" in course
        if not group_rows.empty:
            gdf = group_rows[(group_rows["上層"].astype(str).eq(course)) | (group_rows["課別"].astype(str).eq(course_key))].copy()
            course_html_parts.append(_course_column_by_structure(course, course_df, course_row, gdf, project=project))
        else:
            course_html_parts.append(_course_column(course, course_df, project=project))
    unassigned_courses = [c for c in _sort_unique(course_source[DIM_COURSE].tolist(), _dept_sort_key) if c not in assigned_courses and c not in {FALLBACK_DEPARTMENT, "製造部"}]
    for course in unassigned_courses:
        course_df = people[people[DIM_COURSE].astype(str).eq(course)]
        project = "二" in course or "2" in course
        course_html_parts.append(_course_column(course, course_df, project=project))
    manager_label = html.escape(_department_manager_label(people))
    return f"""
    <div class="xl-org-canvas" data-layout-version="v25-server-xlsx-export">
      <svg class="xl-auto-connectors" aria-hidden="true"></svg>
      {_summary_table(people, year)}
      <div class="xl-exec-zone xl-drop-zone" data-accept="person">{_top_management(people)}</div>
      <div class="xl-root-card" draggable="true" data-drag-type="department" data-card-id="department_root"><span class="xl-grip">☰</span><b>製造部</b><strong>{manager_label}</strong><em>{len(people)}員｜直接 {_direct_count(people)}</em><button class="xl-add" type="button" data-add-type="course">＋課別</button>{_delete_button()}</div>
      <div class="xl-course-lane xl-drop-zone" data-accept="course">{''.join(course_html_parts)}</div>
    </div>
    """


def render_org_html(people: pd.DataFrame) -> str:
    """卷軸式清單備用模式：部門 → 課別 → 工段 → 職稱 → 人力來源 → 人員。"""
    if people is None or people.empty:
        return "<div class='tech-card'>目前沒有組織資料。</div>"
    people = _normalize_people(people)
    parts: list[str] = ["""
<style>
.org-scroll-list { display:flex; flex-direction:column; gap:16px; }
.org-scroll-card { border:1px solid rgba(0,212,255,.28); border-radius:18px; padding:14px; background:rgba(10,25,45,.78); }
.org-scroll-title { color:#fff; font-weight:950; margin-bottom:10px; }
.org-scroll-chip { display:inline-flex; padding:5px 9px; border-radius:999px; background:rgba(0,212,255,.08); color:#BDF4FF; margin:2px 4px 6px 0; font-size:.8rem; }
.org-scroll-person { border-left:3px solid #00D4FF; background:rgba(255,255,255,.055); border-radius:10px; padding:7px 9px; margin:5px 0; color:#fff; }
.org-scroll-person small { color:#9FB6C8; margin-left:8px; }

</style>
<div class='org-scroll-list'>"""]
    for course in _sort_unique(people[DIM_COURSE].tolist(), _dept_sort_key):
        course_df = people[people[DIM_COURSE].astype(str).eq(course)]
        parts.append(f"<section class='org-scroll-card'><div class='org-scroll-title'>{html.escape(course)}｜{len(course_df)} 人</div>")
        for stage in _sort_unique(course_df[DIM_STAGE].tolist(), _stage_sort_key):
            stage_df = course_df[course_df[DIM_STAGE].astype(str).eq(stage)]
            parts.append(f"<div class='org-scroll-chip'>工段：{html.escape(stage)}｜{len(stage_df)} 人</div>")
            for _, row in stage_df.sort_values(PERSON_NAME, kind="stable").iterrows():
                parts.append(f"<div class='org-scroll-person'>{html.escape(_safe_text(row.get(PERSON_NAME), '未命名'))}<small>{html.escape(_safe_text(row.get(DIM_TITLE), '未設定職稱'))}｜{html.escape(_safe_text(row.get(DIM_SOURCE), '未設定來源'))}</small></div>")
        parts.append("</section>")
    parts.append("</div>")
    return "".join(parts)




def _extract_element_by_card_id(markup: str, card_id: str) -> str | None:
    if not markup:
        return None
    pattern = re.compile(r'<(?P<tag>\w+)(?P<attrs>[^>]*data-card-id=["\']' + re.escape(card_id) + r'["\'][^>]*)>', re.IGNORECASE)
    match = pattern.search(markup)
    if not match:
        return None
    tag = match.group('tag')
    start = match.start()
    pos = match.end()
    depth = 1
    token = re.compile(r'</?' + re.escape(tag) + r'\b[^>]*>', re.IGNORECASE)
    while depth and pos < len(markup):
        next_match = token.search(markup, pos)
        if not next_match:
            return None
        text = next_match.group(0)
        if text.startswith('</'):
            depth -= 1
        elif not text.endswith('/>'):
            depth += 1
        pos = next_match.end()
    return markup[start:pos] if depth == 0 else None


def _replace_element_by_card_id(markup: str, card_id: str, replacement: str) -> str:
    current = _extract_element_by_card_id(markup, card_id)
    if not current:
        return markup
    return markup.replace(current, replacement, 1)


def _merge_saved_layout_with_live(saved_layout_html: str | None, live_layout_html: str) -> str:
    """Keep the saved display layout but refresh count-critical blocks from current 01/02 data.

    This prevents an old permanent layout from freezing the upper manpower table or
    department totals after users add people in 01/02. New/changed people are still
    safest in the live layout; therefore the 03 page defaults to live 01/02 sync.
    """
    saved = _safe_text(saved_layout_html, "")
    if not saved:
        return live_layout_html
    # V16 refines the machine layout: 02 派遣名單 is merged into the same hierarchy,
    # 組長 is rendered directly under 機型, and 工段/組別 are laid out horizontally.
    # Older saved layouts do not contain these structure nodes, so they must fall back
    # to the live tree.
    if "xl-auto-connectors" not in saved or "data-parent-card-id" not in saved or "v25-server-xlsx-export" not in saved:
        return live_layout_html
    for card_id in ["summary_top", "department_root"]:
        live_block = _extract_element_by_card_id(live_layout_html, card_id)
        if live_block:
            saved = _replace_element_by_card_id(saved, card_id, live_block)
    return saved




def export_org_chart_layout_xlsx_bytes(people: pd.DataFrame, year: object | None = None, structure_settings: pd.DataFrame | None = None) -> bytes:
    """Create a real Excel workbook for the organization chart layout.

    This export is generated server-side with openpyxl instead of browser-side XML/ZIP
    assembly. It prevents Excel repair warnings and blank worksheets on Windows Excel.
    The layout is deterministic: department -> course -> machine -> group/stage -> title -> people.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    normalized = _normalize_people(people)
    wb = Workbook()
    ws = wb.active
    ws.title = "版面圖"
    detail = wb.create_sheet("卡片明細")

    # Palette uses light fills so printing and Excel preview remain readable.
    fills = {
        "title": PatternFill("solid", fgColor="0B1F33"),
        "department": PatternFill("solid", fgColor="22D3EE"),
        "course": PatternFill("solid", fgColor="FDBA74"),
        "machine": PatternFill("solid", fgColor="A7F3D0"),
        "leader": PatternFill("solid", fgColor="FDE68A"),
        "stage": PatternFill("solid", fgColor="C4B5FD"),
        "role": PatternFill("solid", fgColor="BAE6FD"),
        "person": PatternFill("solid", fgColor="F1F5F9"),
        "dispatch": PatternFill("solid", fgColor="FED7AA"),
        "header": PatternFill("solid", fgColor="1E293B"),
    }
    thin = Side(style="thin", color="94A3B8")
    medium = Side(style="medium", color="0891B2")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col in range(1, 42):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(1, 12):
        detail.column_dimensions[get_column_letter(col)].width = 18
    detail.column_dimensions[get_column_letter(4)].width = 34

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    c = ws.cell(1, 1, "製造部階層式組織圖版面匯出")
    c.fill = fills["title"]
    c.font = Font(color="FFFFFF", bold=True, size=16)
    c.alignment = center
    c.border = border_medium
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)
    c = ws.cell(2, 1, f"年份：{_safe_text(year, '全部')}｜匯出時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}｜來源：01. 超慧員工名單 + 02. 派遣名單")
    c.alignment = left
    c.border = border_thin
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=16)
    c = ws.cell(3, 1, f"匯出狀態：共 {len(normalized)} 人。此檔由伺服端 openpyxl 產生，避免 Excel 修復後空白。")
    c.alignment = left
    c.border = border_thin

    def direct_count(df: pd.DataFrame) -> int:
        return _direct_count(df) if df is not None and not df.empty else 0

    def set_box(sheet, row: int, col: int, row_span: int, col_span: int, value: str, fill_key: str, font_color: str = "0F172A") -> None:
        row_span = max(1, int(row_span))
        col_span = max(1, int(col_span))
        if row_span > 1 or col_span > 1:
            sheet.merge_cells(start_row=row, start_column=col, end_row=row + row_span - 1, end_column=col + col_span - 1)
        cell = sheet.cell(row, col, value)
        cell.fill = fills.get(fill_key, fills["person"])
        cell.font = Font(color=font_color, bold=True, size=10)
        cell.alignment = center
        cell.border = border_medium if fill_key in {"department", "course", "machine"} else border_thin
        for rr in range(row, row + row_span):
            sheet.row_dimensions[rr].height = 28

    # Detail sheet first so it always has a complete non-empty tab even if the visual layout is compact.
    detail_headers = ["序號", "部門", "課別", "機型", "工段/組別", "職稱", "姓名", "人力來源", "名單來源", "是否直接人力", "可用比例"]
    detail.append(["製造部組織圖卡片明細"])
    detail.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(detail_headers))
    detail.cell(1, 1).fill = fills["title"]
    detail.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=14)
    detail.cell(1, 1).alignment = center
    detail.append([])
    detail.append(detail_headers)
    for cell in detail[3]:
        cell.fill = fills["header"]
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = center
        cell.border = border_thin
    for idx, (_, person) in enumerate(normalized.iterrows(), start=1):
        detail.append([
            idx,
            _safe_text(person.get(DIM_DEPARTMENT), FALLBACK_DEPARTMENT),
            _safe_text(person.get(DIM_COURSE), FALLBACK_COURSE),
            _safe_text(person.get(DIM_MACHINE), FALLBACK_MACHINE),
            _safe_text(person.get(DIM_STAGE), FALLBACK_STAGE),
            _safe_text(person.get(DIM_TITLE), ""),
            _safe_text(person.get(PERSON_NAME), ""),
            _safe_text(person.get(DIM_SOURCE), ""),
            _safe_text(person.get(ROSTER_SOURCE), ""),
            _safe_text(person.get("是否直接人力"), ""),
            float(person.get("可用比例", 0) or 0),
        ])
    for row in detail.iter_rows(min_row=4):
        for cell in row:
            cell.border = border_thin
            cell.alignment = left if cell.column in {4, 5, 6, 7, 8, 9} else center
    detail.freeze_panes = "A4"
    detail.auto_filter.ref = f"A3:{get_column_letter(len(detail_headers))}{max(3, detail.max_row)}"

    if normalized.empty:
        set_box(ws, 6, 6, 2, 6, "目前沒有組織資料", "department")
    else:
        manager_label = _department_manager_label(normalized)
        set_box(ws, 5, 7, 2, 4, f"製造部\n{manager_label}\n{len(normalized)}員｜直接 {direct_count(normalized)}", "department")

        course_source = normalized[~normalized[DIM_TITLE].map(lambda v: _contains_any(v, TOP_LEADER_KEYWORDS))].copy()
        course_values = [c for c in _sort_unique(course_source[DIM_COURSE].tolist(), _dept_sort_key) if c not in {FALLBACK_DEPARTMENT, "製造部"}]
        if not course_values:
            course_values = [FALLBACK_COURSE]

        course_start_cols = [1, 11, 21, 31]
        for course_index, course in enumerate(course_values):
            course_col = course_start_cols[course_index % len(course_start_cols)]
            course_block_row = 9 + (course_index // len(course_start_cols)) * 48
            course_df = normalized[normalized[DIM_COURSE].astype(str).eq(str(course))].copy()
            set_box(ws, course_block_row, course_col + 2, 2, 5, f"{course}\n{len(course_df)}員｜直接 {direct_count(course_df)}", "course")

            leader_df = course_df[course_df[DIM_TITLE].map(lambda v: _contains_any(v, COURSE_LEADER_KEYWORDS))].copy()
            leader_text = "未設定課長/主任" if leader_df.empty else "\n".join(
                f"{_safe_text(r.get(DIM_TITLE), '')}｜{_safe_text(r.get(PERSON_NAME), '')}" for _, r in leader_df.iterrows()
            )
            set_box(ws, course_block_row + 3, course_col + 2, 2, 5, leader_text, "leader")

            work_df = _course_stage_people(course_df)
            machines = _sort_unique(work_df[DIM_MACHINE].tolist(), _machine_sort_key) if not work_df.empty else []
            if not machines:
                set_box(ws, course_block_row + 7, course_col + 1, 2, 7, "未設定機型 / 組別", "machine")
                continue
            machine_row = course_block_row + 7
            for machine_index, machine in enumerate(machines):
                machine_df = work_df[work_df[DIM_MACHINE].astype(str).eq(str(machine))].copy()
                if machine_index and machine_index % 2 == 0:
                    machine_row += 18
                machine_col = course_col if machine_index % 2 == 0 else course_col + 5
                set_box(ws, machine_row, machine_col, 2, 4, f"{machine}\n{len(machine_df)}員｜直接 {direct_count(machine_df)}", "machine")
                m_leaders = machine_df[machine_df[DIM_TITLE].map(lambda v: _contains_any(v, STAGE_LEADER_KEYWORDS))].copy()
                if not m_leaders.empty:
                    leader_names = "\n".join(f"組長｜{_safe_text(r.get(PERSON_NAME), '')}" for _, r in m_leaders.iterrows())
                    set_box(ws, machine_row + 3, machine_col, 2, 4, leader_names, "leader")
                    stage_base_row = machine_row + 6
                else:
                    stage_base_row = machine_row + 3
                member_df = _exclude_title_keywords(machine_df, STAGE_LEADER_KEYWORDS)
                stages = _sort_unique(member_df[DIM_STAGE].tolist(), _stage_sort_key) if not member_df.empty else []
                if not stages:
                    set_box(ws, stage_base_row, machine_col, 2, 4, "未設定工段", "stage")
                    continue
                for stage_index, stage in enumerate(stages[:4]):
                    stage_df = member_df[member_df[DIM_STAGE].astype(str).eq(str(stage))].copy()
                    stage_col = machine_col + (stage_index % 2) * 2
                    stage_row = stage_base_row + (stage_index // 2) * 6
                    set_box(ws, stage_row, stage_col, 2, 2, f"{stage}\n{len(stage_df)}員", "stage")
                    titles = _sort_unique(stage_df[DIM_TITLE].tolist(), _title_sort_key)
                    role_lines: list[str] = []
                    for title in titles:
                        title_df = stage_df[stage_df[DIM_TITLE].astype(str).eq(str(title))]
                        names = "、".join(_safe_text(x, "") for x in title_df[PERSON_NAME].head(4).tolist())
                        more = f" +{len(title_df)-4}" if len(title_df) > 4 else ""
                        role_lines.append(f"{title}({len(title_df)}) {names}{more}")
                    fill_key = "dispatch" if _dispatch_count(stage_df) else "person"
                    set_box(ws, stage_row + 2, stage_col, 3, 2, "\n".join(role_lines[:4]) or "未配置組員", fill_key)
                if len(stages) > 4:
                    set_box(ws, stage_base_row + 12, machine_col, 2, 4, f"其他工段：{len(stages)-4} 組，請見卡片明細", "stage")

    ws.freeze_panes = "A4"
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 18, 22)
    for sheet in [ws, detail]:
        sheet.sheet_view.showGridLines = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def render_org_component_html(people: pd.DataFrame, saved_layout_html: str | None = None, storage_key_suffix: str | None = None, year: object | None = None, structure_settings: pd.DataFrame | None = None, viewport_height: int = 1580, canvas_width: int = 3600, canvas_height: int = 2600, default_zoom: int = 75, auto_expand_canvas: bool = True) -> str:
    """Return the draggable hierarchy tree in the requested image-style layout."""
    inner = _image_style_tree_inner(people, year=year, structure_settings=structure_settings)
    mount_inner = _merge_saved_layout_with_live(saved_layout_html, inner) if saved_layout_html else inner
    storage_key = "spt_capacity_org_tree_layout_v25_" + _slug(storage_key_suffix or "default", "layout")
    has_server_layout = "true" if saved_layout_html else "false"
    try:
        viewport_height_i = int(viewport_height)
    except Exception:
        viewport_height_i = 1580
    try:
        canvas_width_i = int(canvas_width)
    except Exception:
        canvas_width_i = 3600
    try:
        canvas_height_i = int(canvas_height)
    except Exception:
        canvas_height_i = 2600
    try:
        default_zoom_i = int(default_zoom)
    except Exception:
        default_zoom_i = 75
    viewport_height_i = max(760, min(2600, viewport_height_i))
    canvas_width_i = max(3000, min(7000, canvas_width_i))
    canvas_height_i = max(1900, min(5200, canvas_height_i))
    default_zoom_i = max(20, min(180, default_zoom_i))
    auto_expand_canvas_js = "true" if auto_expand_canvas else "false"
    template = '''<!doctype html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8" />
<style>
:root { --bg:#06111F; --panel:#071527; --grid:#16415c; --line:#8AA4B8; --cream:#0A1628; --yellow:#FACC15; --peach:#FB923C; --green:#22C55E; --lightgreen:#34D399; --pink:#F472B6; --orange:#FB923C; --blue:#38BDF8; --cyan:#22D3EE; --text:#EAF6FF; --muted:#9FB6C8; --glow:rgba(34,211,238,.28); }
* { box-sizing:border-box; }
body { margin:0; background:linear-gradient(135deg,#06111f,#02040a); color:var(--text); font-family:"Microsoft JhengHei","Noto Sans TC",Arial,sans-serif; overflow:auto; }
.xl-toolbar { position:sticky; top:0; z-index:20; display:flex; flex-wrap:wrap; align-items:center; gap:9px; padding:10px 12px; color:#E8F6FF; background:rgba(2,8,18,.95); border:1px solid rgba(0,212,255,.20); border-radius:14px; margin-bottom:12px; box-shadow:0 0 20px rgba(0,212,255,.10); }
.xl-toolbar b { font-size:16px; color:#fff; }
.xl-toolbar span { font-size:12px; color:#BDF4FF; }
.xl-toolbar button { border:1px solid rgba(0,212,255,.35); background:linear-gradient(135deg,rgba(0,212,255,.18),rgba(59,130,246,.10)); color:#fff; border-radius:10px; padding:7px 10px; font-weight:900; cursor:pointer; box-shadow:0 0 12px rgba(0,212,255,.08); }
.xl-toolbar button:hover { border-color:rgba(34,211,238,.85); box-shadow:0 0 18px rgba(34,211,238,.18); }
.xl-toolbar input[type=range] { width:110px; accent-color:#22D3EE; }
.xl-shell { overflow:auto; height:__VIEWPORT_HEIGHT__px; border:1px solid rgba(0,212,255,.30); border-radius:18px; background:radial-gradient(circle at 20% 0%, rgba(34,211,238,.08), transparent 28%), linear-gradient(135deg,#071527,#020817); padding:16px; box-shadow:inset 0 0 24px rgba(0,212,255,.10), 0 0 24px rgba(0,0,0,.22); }
.xl-shell.hide-scrollbars { scrollbar-width:none; }
.xl-shell.hide-scrollbars::-webkit-scrollbar { width:0; height:0; }
.xl-shell.hide-scrollbars { overflow:hidden; }
.xl-selection-box { position:fixed; display:none; border:2px dashed #00a8ff; background:rgba(0,168,255,.12); z-index:999999; pointer-events:none; }
.xl-selected { outline:3px solid #00a8ff !important; outline-offset:2px; box-shadow:0 0 0 4px rgba(0,168,255,.18), 0 0 18px rgba(0,168,255,.45) !important; }
.xl-resize-hint { display:none; }
.xl-card-resize-handle { display:none; position:absolute; right:-7px; bottom:-7px; width:22px; height:22px; z-index:999; cursor:nwse-resize; border-right:4px solid rgba(56,189,248,.92); border-bottom:4px solid rgba(56,189,248,.92); border-radius:0 0 6px 0; background:linear-gradient(135deg, transparent 0 38%, rgba(56,189,248,.16) 38% 62%, transparent 62% 100%); box-shadow:0 0 10px rgba(56,189,248,.45); pointer-events:auto; touch-action:none; }
body.org-editing .xl-machine-card > .xl-card-resize-handle,
body.org-editing .xl-course > .xl-card-resize-handle,
body.org-editing .xl-stage-card > .xl-card-resize-handle,
body.org-editing .xl-role-block > .xl-card-resize-handle { display:block; }
body.org-editing .xl-card-resizing { outline:2px solid rgba(56,189,248,.92) !important; outline-offset:3px; box-shadow:0 0 0 4px rgba(56,189,248,.16), 0 0 24px rgba(56,189,248,.42) !important; }
.xl-org-canvas { position:relative; min-width:__CANVAS_WIDTH__px; min-height:__CANVAS_HEIGHT__px; padding:0 20px 40px; color:var(--text); }
.xl-summary { position:absolute; left:0; top:0; width:1250px; }
.xl-summary table { border-collapse:collapse; width:100%; table-layout:fixed; font-size:12px; background:rgba(10,22,40,.92); color:var(--text); box-shadow:0 0 16px rgba(34,211,238,.10); }
.xl-summary th, .xl-summary td { border:1px solid rgba(148,163,184,.65); padding:6px 4px; text-align:center; vertical-align:middle; }
.xl-summary th { font-weight:900; }
.xl-summary .year { background:rgba(14,165,233,.16); color:#fff; }
.xl-summary .red, .xl-summary .total-cell { color:#e60000; font-weight:950; }
.xl-exec-zone { position:absolute; left:1420px; top:0; width:850px; height:145px; display:flex; align-items:flex-start; justify-content:center; gap:150px; }
.xl-top-box { position:relative; min-width:116px; border:1px solid rgba(34,211,238,.55); background:linear-gradient(180deg,rgba(15,23,42,.96),rgba(2,8,23,.96)); color:#EAF6FF; text-align:center; padding:0 24px 6px; box-shadow:0 0 15px rgba(34,211,238,.18); }
.xl-top-box b { display:block; background:linear-gradient(90deg,rgba(251,146,60,.95),rgba(250,204,21,.80)); color:#111827; margin:0 -24px 5px; padding:4px; font-size:12px; }
.xl-top-box em { display:block; font-style:normal; font-size:12px; min-height:18px; }
.xl-manager-stem { position:absolute; left:1575px; top:58px; width:2px; height:120px; background:linear-gradient(180deg,rgba(125,249,255,.95),rgba(255,255,255,.90)); z-index:3; box-shadow:0 0 10px rgba(34,211,238,.45); }
.xl-root-card { position:absolute; left:1510px; top:165px; width:140px; border:1px solid rgba(34,211,238,.60); background:linear-gradient(180deg,rgba(15,23,42,.98),rgba(8,47,73,.88)); color:#EAF6FF; text-align:center; padding:0 22px 5px; box-shadow:0 0 18px rgba(34,211,238,.18); z-index:6; }
.xl-root-card b { display:block; background:linear-gradient(90deg,#22D3EE,#38BDF8); color:#04111f; padding:5px; margin:0 -22px 4px; font-size:13px; }
.xl-root-card em { display:block; font-size:11px; font-style:normal; color:#BDF4FF; }
.xl-root-line { position:absolute; left:845px; top:232px; width:1390px; height:2px; background:linear-gradient(90deg,rgba(34,211,238,.18),rgba(255,255,255,.92),rgba(125,249,255,.96),rgba(34,211,238,.18)); z-index:3; box-shadow:0 0 10px rgba(34,211,238,.42); }
.xl-custom-line { position:absolute; z-index:4; background:rgba(255,255,255,.95); box-shadow:0 0 8px rgba(34,211,238,.25); }
.xl-course-lane { position:absolute; left:35px; top:285px; display:grid; grid-template-columns:minmax(1220px,1320px) minmax(980px,1080px); gap:115px; align-items:start; min-height:760px; }
.xl-course { position:relative; border-top:2px solid rgba(125,249,255,.76); padding:36px 12px 18px; min-height:720px; border-radius:12px; background:linear-gradient(180deg,rgba(8,47,73,.08),rgba(2,8,23,.02)); }
.xl-course.project-course { border:2px dashed rgba(34,211,238,.42); border-top:8px solid #22C55E; padding:42px 14px 18px; }
.xl-course::before { content:""; position:absolute; top:-36px; left:50%; width:2px; height:36px; background:linear-gradient(180deg,rgba(255,255,255,.95),rgba(34,211,238,.86)); box-shadow:0 0 9px rgba(34,211,238,.38); }
.xl-course-count { position:absolute; left:50%; transform:translateX(-50%); top:-64px; min-width:80px; background:linear-gradient(90deg,#FACC15,#FDE68A); color:#111827; text-align:center; font-weight:900; font-size:12px; padding:4px; box-shadow:0 0 14px rgba(250,204,21,.22); z-index:7; }
.xl-course-title { position:absolute; left:50%; transform:translateX(-50%); top:-37px; width:190px; background:linear-gradient(135deg,rgba(251,146,60,.95),rgba(34,211,238,.32)); color:#fff; border:1px solid rgba(34,211,238,.55); text-align:center; padding:5px 24px 6px; font-weight:950; box-shadow:0 0 15px rgba(251,146,60,.18); z-index:7; }
.xl-course-title b { display:block; }
.xl-course-title em { display:block; font-size:11px; font-style:normal; font-weight:500; }
.xl-leader-zone { display:flex; justify-content:center; gap:25px; min-height:70px; margin-bottom:22px; position:relative; }
.xl-leader-zone::before { content:""; position:absolute; left:50%; top:-20px; width:2px; height:18px; background:linear-gradient(180deg,rgba(255,255,255,.85),rgba(34,211,238,.70)); box-shadow:0 0 8px rgba(34,211,238,.28); }
.xl-leader-card { position:relative; width:112px; border:1px solid rgba(250,204,21,.60); background:linear-gradient(180deg,rgba(15,23,42,.96),rgba(69,26,3,.55)); color:#fff; text-align:center; padding:0 22px 4px; box-shadow:0 0 12px rgba(250,204,21,.16); }
.xl-leader-card b { display:block; background:linear-gradient(90deg,#FACC15,#FDE68A); color:#111827; margin:0 -22px 4px; padding:3px; font-size:12px; }
.xl-leader-card em { display:block; font-style:normal; font-size:12px; }
.xl-stage-lane { position:relative; display:grid; grid-template-columns:repeat(auto-fill, minmax(235px, 1fr)); gap:58px 54px; align-items:start; justify-content:flex-start; padding-top:30px; }
.xl-stage-lane::before { content:""; position:absolute; left:8%; right:8%; top:0; height:2px; background:linear-gradient(90deg,transparent,rgba(255,255,255,.84),rgba(125,249,255,.92),transparent); box-shadow:0 0 8px rgba(34,211,238,.34); }
.xl-stage-card { position:relative; width:100%; min-width:235px; max-width:285px; min-height:130px; padding-top:42px; justify-self:start; }
.xl-stage-card::before { content:""; position:absolute; top:-30px; left:50%; width:2px; height:60px; background:linear-gradient(180deg,rgba(255,255,255,.88),rgba(34,211,238,.75)); box-shadow:0 0 8px rgba(34,211,238,.32); }
.xl-stage-head { position:absolute; top:30px; left:0; right:0; border:1px solid rgba(34,211,238,.55); background:linear-gradient(135deg,#FACC15,#22D3EE); color:#07111f; text-align:center; padding:5px 30px 6px; min-height:46px; z-index:2; overflow:hidden; box-shadow:0 0 14px rgba(34,211,238,.14); }
.xl-stage-head b { display:block; font-size:12px; }
.xl-stage-head em { display:block; font-size:10px; font-style:normal; color:#111; line-height:1.15; }
.xl-stage-leaders { margin-top:52px; min-height:20px; }
.xl-role-zone { position:relative; display:flex; flex-direction:column; gap:14px; margin-top:12px; min-height:42px; padding-top:8px; }
.xl-role-zone::before { content:""; position:absolute; left:50%; top:-12px; width:2px; height:16px; background:rgba(125,249,255,.55); box-shadow:0 0 6px rgba(34,211,238,.25); }
.xl-role-block { position:relative; border:1px solid rgba(148,163,184,.62); background:rgba(15,23,42,.94); color:#EAF6FF; min-height:52px; min-width:170px; padding-top:30px; overflow:hidden; box-shadow:0 0 10px rgba(34,211,238,.08); }
.xl-role-block::before { content:""; position:absolute; left:50%; top:-14px; width:2px; height:14px; background:rgba(125,249,255,.48); box-shadow:0 0 6px rgba(34,211,238,.22); }
.xl-role-title { position:absolute; left:-1px; right:-1px; top:-1px; min-height:28px; border-bottom:1px solid rgba(148,163,184,.58); text-align:center; padding:4px 34px 4px; font-size:11px; overflow:hidden; }
.xl-role-title b { display:block; line-height:1.05; }
.xl-role-title em { display:block; font-size:9px; font-style:normal; color:#333; }
.xl-person-zone { min-height:30px; }
.xl-person-row { position:relative; display:grid; grid-template-columns:18px minmax(70px,1fr) 62px; align-items:center; min-height:25px; border-top:1px solid rgba(148,163,184,.55); background:rgba(15,23,42,.96); color:#EAF6FF; font-size:11px; padding:3px 28px 3px 2px; overflow:hidden; }
.xl-person-row b { overflow:hidden; white-space:nowrap; text-overflow:ellipsis; }
.xl-person-row small { color:#D7ECFF; font-size:9px; overflow:hidden; white-space:nowrap; text-overflow:ellipsis; }
.xl-empty { border:1px dashed rgba(148,163,184,.55); background:rgba(15,23,42,.72); font-size:11px; color:#BDF4FF; padding:5px; text-align:center; }
.role-leader, .role-leader .xl-role-title { background:linear-gradient(90deg,rgba(250,204,21,.95),rgba(251,191,36,.68)); color:#111827; box-shadow:0 0 10px rgba(250,204,21,.18); }
.role-senior .xl-role-title, .role-senior.xl-person-row { background:linear-gradient(90deg,rgba(34,197,94,.95),rgba(20,184,166,.78)); color:#04111f; box-shadow:0 0 10px rgba(34,197,94,.16); }
.role-engineer .xl-role-title, .role-engineer.xl-person-row { background:linear-gradient(90deg,rgba(52,211,153,.86),rgba(186,230,253,.70)); color:#04111f; }
.role-assistant .xl-role-title, .role-assistant.xl-person-row { background:linear-gradient(90deg,rgba(244,114,182,.92),rgba(192,132,252,.78)); color:#fff; box-shadow:0 0 10px rgba(244,114,182,.16); }
.role-dispatch .xl-role-title, .role-dispatch.xl-person-row { background:linear-gradient(90deg,rgba(251,146,60,.92),rgba(253,186,116,.74)); color:#111827; }
.role-tech .xl-role-title, .role-tech.xl-person-row { background:linear-gradient(90deg,rgba(56,189,248,.95),rgba(34,211,238,.74)); color:#03111d; box-shadow:0 0 10px rgba(56,189,248,.18); }
.role-general .xl-role-title, .role-general.xl-person-row { background:linear-gradient(90deg,rgba(30,41,59,.96),rgba(51,65,85,.88)); color:#EAF6FF; }
.xl-actions, .xl-add, .dim-edit, .dim-delete { position:absolute; }
.xl-add { right:50px; top:2px; border:1px solid #111; background:#e8f7ff; color:#111; border-radius:3px; padding:2px 4px; font-size:10px; font-weight:900; cursor:pointer; }
.dim-edit { right:27px; top:2px; border:1px solid #0369a1; background:#dff6ff; color:#0369a1; border-radius:3px; min-width:18px; height:18px; padding:0 4px; font-weight:950; cursor:pointer; }
.dim-delete { right:3px; top:2px; border:1px solid #9f1239; background:#ffe4e6; color:#be123c; border-radius:3px; min-width:18px; height:18px; padding:0 4px; font-weight:950; cursor:pointer; }
.xl-grip { color:#0070c0; cursor:grab; user-select:none; font-weight:950; }
.dragging { opacity:.70; filter:drop-shadow(0 12px 16px rgba(0,0,0,.28)); }
.manually-positioned { position:relative; z-index:9; }
.xl-drop-zone.drop-active { outline:2px dashed #00a8ff; outline-offset:4px; background:rgba(0,168,255,.10); }
body.org-editing .xl-course,
body.org-editing .xl-machine-card,
body.org-editing .xl-stage-card,
body.org-editing .xl-role-block,
body.org-editing .xl-person-row,
body.org-editing .xl-leader-card,
body.org-editing .xl-top-box,
body.org-editing .xl-root-card {
  resize:both !important;
  overflow:auto !important;
  min-width:80px;
  min-height:24px;
  max-width:none !important;
  max-height:none !important;
  scrollbar-width:thin;
}
body.org-editing .xl-person-zone,
body.org-editing .xl-role-zone,
body.org-editing .xl-machine-lane,
body.org-editing .xl-machine-group-lane,
body.org-editing .xl-stage-lane,
body.org-editing .xl-leader-zone { overflow:visible !important; }
body.org-editing .xl-role-block { min-width:170px; }
body.org-editing .xl-person-row { min-width:150px; width:100%; resize:both !important; }
body.org-editing .xl-course::after,
body.org-editing .xl-machine-card::after,
body.org-editing .xl-stage-card::after,
body.org-editing .xl-role-block::after,
body.org-editing .xl-person-row::after,
body.org-editing .xl-leader-card::after,
body.org-editing .xl-top-box::after,
body.org-editing .xl-root-card::after {
  content:'';
  position:absolute;
  right:2px;
  bottom:2px;
  width:16px;
  height:16px;
  z-index:8;
  pointer-events:none;
  border-right:3px solid rgba(0,112,192,.75);
  border-bottom:3px solid rgba(0,112,192,.75);
  background:linear-gradient(135deg, transparent 0 42%, rgba(0,112,192,.32) 42% 54%, transparent 54% 100%);
}
body.org-editing .xl-course:hover,
body.org-editing .xl-machine-card:hover,
body.org-editing .xl-stage-card:hover,
body.org-editing .xl-role-block:hover,
body.org-editing .xl-person-row:hover,
body.org-editing .xl-leader-card:hover,
body.org-editing .xl-top-box:hover,
body.org-editing .xl-root-card:hover { outline:1px dashed rgba(0,112,192,.55); outline-offset:2px; }
body.org-editing .xl-line-node { overflow:visible !important; cursor:move; min-width:1px; min-height:1px; outline:1px dashed rgba(255,255,255,.44); outline-offset:4px; background:#FFFFFF !important; box-shadow:0 0 8px rgba(255,255,255,.62), 0 0 14px rgba(34,211,238,.20); }
.xl-line-node.xl-selected { outline:2px solid #FFFFFF !important; box-shadow:0 0 16px rgba(255,255,255,.72), 0 0 22px rgba(34,211,238,.45) !important; }
body.org-editing .xl-line-node::after { content:'線'; position:absolute; right:-22px; top:-12px; font-size:10px; color:#FFFFFF; font-weight:900; text-shadow:0 0 8px rgba(255,255,255,.75); }
.xl-line-handle { display:none; position:absolute; width:10px; height:10px; border-radius:50%; background:#FFFFFF; border:1px solid rgba(34,211,238,.95); box-shadow:0 0 10px rgba(255,255,255,.80), 0 0 12px rgba(34,211,238,.65); z-index:10; }
body.org-editing .xl-line-node .xl-line-handle { display:block; }
.xl-line-handle.start { left:-5px; top:50%; transform:translateY(-50%); cursor:crosshair; }
.xl-line-handle.end { right:-5px; top:50%; transform:translateY(-50%); cursor:crosshair; }
.xl-line-node[data-line-axis="vertical"] .xl-line-handle.start { left:50%; top:-5px; transform:translateX(-50%); }
.xl-line-node[data-line-axis="vertical"] .xl-line-handle.end { left:50%; right:auto; bottom:-5px; top:auto; transform:translateX(-50%); }
body.org-editing .xl-org-canvas { background-image:linear-gradient(rgba(0,0,0,.045) 1px, transparent 1px), linear-gradient(90deg, rgba(0,0,0,.045) 1px, transparent 1px); background-size:20px 20px; }
body:not(.org-editing) .xl-add,
body:not(.org-editing) .dim-edit,
body:not(.org-editing) .dim-delete,
body:not(.org-editing) #autoLayout,
body:not(.org-editing) #antiOverlapToggle,
body:not(.org-editing) #addHLine,
body:not(.org-editing) #addVLine,
body:not(.org-editing) #deleteSelectedLine,
body:not(.org-editing) #clearSelection,
body:not(.org-editing) #bringToFront,
body:not(.org-editing) #sendToBack,
body:not(.org-editing) #saveLayout,
body:not(.org-editing) #exportPermanentLayout,
body:not(.org-editing) #resetLayout { display:none !important; }
body.org-editing #enableEdit { display:none !important; }
body:not(.org-editing) #disableEdit { display:none !important; }
body:not(.org-editing) [draggable="true"] { cursor:default; }
body:not(.org-editing) .xl-grip { opacity:.28; cursor:default; }
.edit-state { padding:6px 10px; border-radius:999px; border:1px solid rgba(0,212,255,.28); background:rgba(255,255,255,.045); color:#BDF4FF !important; font-weight:900; }
body.org-editing .edit-state { color:#D8FFE9 !important; border-color:rgba(57,255,136,.42); background:rgba(57,255,136,.10); }

/* V15 professional hierarchy layout: fixed 5-level lanes + clean automatic connectors */
.xl-auto-connectors { position:absolute; left:0; top:0; width:100%; height:100%; overflow:visible; pointer-events:none; z-index:1; }
.xl-auto-connector-path { fill:none; stroke:rgba(189,244,255,.92); stroke-width:1.6; stroke-linecap:round; stroke-linejoin:round; filter:drop-shadow(0 0 6px rgba(34,211,238,.48)); }
.xl-auto-connector-path.org-connector-soft { stroke:rgba(125,249,255,.72); stroke-width:1.2; filter:drop-shadow(0 0 4px rgba(34,211,238,.30)); }
.xl-root-line, .xl-manager-stem { display:none !important; }
.xl-summary, .xl-exec-zone, .xl-root-card, .xl-course, .xl-course-title, .xl-course-count, .xl-leader-card, .xl-stage-card, .xl-role-block, .xl-person-row { z-index:4; }
.xl-org-canvas { min-width:2650px; min-height:1450px; }
.xl-exec-zone { left:1320px; top:16px; width:940px; gap:120px; }
.xl-root-card { left:1515px; top:168px; width:180px; }
.xl-course-lane { left:40px; top:318px; width:2520px; display:grid; grid-template-columns:1180px 1180px; gap:120px; align-items:start; }
.xl-course { min-height:930px; padding:58px 22px 26px; border-radius:18px; overflow:visible; }
.xl-course-title { top:-48px; width:230px; min-height:64px; }
.xl-course-count { top:-78px; min-width:96px; }
.xl-leader-zone { min-height:78px; margin-bottom:34px; gap:30px; }
.xl-stage-lane { grid-template-columns:repeat(auto-fit, minmax(255px, 255px)); gap:64px 32px; justify-content:center; padding-top:36px; }
.xl-stage-card { width:255px; min-width:255px; max-width:255px; justify-self:center; }
.xl-role-block { min-width:230px; }
.xl-person-row { min-height:27px; }
.xl-course::before, .xl-leader-zone::before, .xl-stage-lane::before, .xl-stage-card::before, .xl-role-zone::before { opacity:.18; }
body.org-editing .xl-auto-connectors { opacity:.92; }

/* V15: professional fixed hierarchy, no decorative pseudo-lines that cross cards. */
.xl-shell { height:1360px; }
.xl-org-canvas { min-width:2820px; min-height:1660px; padding:0 60px 80px; }
.xl-summary { left:0; top:0; width:1230px; z-index:7; }
.xl-exec-zone { left:1320px; top:26px; width:620px; height:76px; gap:28px; justify-content:center; }
.xl-exec-zone:empty { display:none; }
.xl-top-box { min-width:128px; }
.xl-manager-stem, .xl-root-line, .xl-course::before, .xl-leader-zone::before, .xl-stage-lane::before, .xl-stage-card::before, .xl-role-zone::before { display:none !important; }
.xl-root-card { left:1290px; top:214px; width:245px; min-height:82px; border-radius:0; padding:0 24px 7px; }
.xl-root-card b { font-size:16px; padding:7px; margin:0 -24px 5px; }
.xl-root-card strong { display:block; color:#fff; font-size:13px; line-height:1.35; font-weight:950; }
.xl-root-card em { font-size:12px; color:#D9F8FF; }
.xl-course-lane { left:120px; top:430px; width:2580px; display:grid; grid-template-columns:1220px 1220px; gap:140px; align-items:start; }
.xl-course { min-height:1060px; padding:128px 28px 34px; border:1px solid rgba(125,249,255,.28); border-top:2px solid rgba(189,244,255,.84); background:linear-gradient(180deg,rgba(15,23,42,.48),rgba(2,8,23,.08)); box-shadow:inset 0 0 28px rgba(34,211,238,.06); border-radius:16px; overflow:visible; }
.xl-course.project-course { border:1px dashed rgba(34,211,238,.55); border-top:8px solid #22C55E; }
.xl-course-title { top:-78px; width:260px; min-height:72px; padding:8px 28px 8px; }
.xl-course-title b { font-size:17px; }
.xl-course-title em { font-size:12px; color:#F1FEFF; }
.xl-course-count { top:-112px; min-width:106px; font-size:13px; padding:5px; }
.xl-leader-zone { min-height:90px; margin:0 auto 64px; display:flex; justify-content:center; gap:34px; }
.xl-leader-card { width:128px; min-height:66px; }
.xl-stage-lane { display:grid; grid-template-columns:repeat(4, 260px); gap:72px 36px; justify-content:center; align-items:start; padding-top:0; }
.xl-stage-card { width:260px; min-width:260px; max-width:260px; padding-top:0; min-height:150px; justify-self:center; }
.xl-stage-head { position:relative; top:auto; left:auto; right:auto; min-height:58px; padding:8px 30px 8px; }
.xl-stage-head b { font-size:13px; }
.xl-stage-head em { font-size:11px; }
.xl-stage-leaders { margin-top:12px; min-height:16px; display:flex; flex-direction:column; align-items:center; gap:8px; }
.xl-role-zone { margin-top:10px; display:flex; flex-direction:column; gap:10px; }
.xl-role-block { width:100%; min-width:236px; }
.xl-person-zone { display:flex; flex-direction:column; gap:0; }
.xl-person-row { min-height:28px; }
.xl-auto-connector-path { stroke:rgba(189,244,255,.88); stroke-width:1.35; filter:drop-shadow(0 0 5px rgba(34,211,238,.35)); }
.xl-auto-connector-path.org-connector-soft { stroke:rgba(125,249,255,.44); stroke-width:.95; filter:none; }
.xl-empty { min-height:36px; display:flex; align-items:center; justify-content:center; }


/* V16: dispatch-inclusive machine hierarchy with leader row + horizontal stage lane */
.xl-org-canvas { min-width:__CANVAS_WIDTH__px; min-height:__CANVAS_HEIGHT__px; }
.xl-shell { height:__VIEWPORT_HEIGHT__px; min-height:__VIEWPORT_HEIGHT__px; max-height:__VIEWPORT_HEIGHT__px; }
.xl-root-card { left:1395px; top:205px; }
.xl-course-lane { left:100px; top:430px; width:2840px; grid-template-columns:1340px 1340px; gap:160px; }
.xl-course { min-height:1240px; padding:128px 34px 38px; }
.xl-leader-zone { margin-bottom:54px; }
.xl-machine-lane { position:relative; display:grid; grid-template-columns:repeat(auto-fit, minmax(410px, 1fr)); gap:84px 52px; justify-content:center; align-items:start; overflow:visible; }
.xl-machine-card { position:relative; min-width:390px; max-width:560px; min-height:280px; justify-self:center; padding:82px 18px 20px; border:1px solid rgba(56,189,248,.42); border-top:5px solid rgba(34,211,238,.88); border-radius:14px; background:linear-gradient(180deg,rgba(8,47,73,.34),rgba(2,8,23,.12)); box-shadow:inset 0 0 22px rgba(34,211,238,.07), 0 0 16px rgba(34,211,238,.08); z-index:4; overflow:visible; }
.xl-machine-head { position:absolute; left:50%; top:-38px; transform:translateX(-50%); width:300px; min-height:72px; border:1px solid rgba(34,211,238,.62); background:linear-gradient(135deg,#D9F99D,#22D3EE); color:#06111F; text-align:center; padding:8px 30px 8px; font-weight:950; box-shadow:0 0 16px rgba(34,211,238,.16); z-index:6; }
.xl-machine-head b { display:block; font-size:15px; }
.xl-machine-head em { display:block; font-style:normal; font-size:11px; color:#102033; line-height:1.2; }
.xl-machine-leader-lane { display:flex; justify-content:center; align-items:flex-start; flex-wrap:wrap; gap:18px; min-height:76px; margin:4px auto 28px; }
.xl-machine-leader-lane .xl-leader-card { width:128px; }
.xl-machine-group-lane { display:grid; grid-template-columns:repeat(var(--machine-stage-cols, 2), minmax(220px, 1fr)); gap:32px 24px; justify-content:center; align-items:start; overflow:visible; }
.xl-machine-card .xl-stage-card { width:100%; min-width:220px; max-width:none; padding-top:0; min-height:140px; justify-self:stretch; }
.xl-machine-card .xl-stage-head { position:relative; top:auto; left:auto; right:auto; min-height:56px; padding:8px 30px 8px; }
.xl-machine-card .xl-stage-head b { font-size:13px; }
.xl-machine-card .xl-role-zone { margin-top:10px; gap:8px; }
.xl-machine-card .xl-role-block { min-width:0; width:100%; }
.xl-machine-card .xl-person-row { min-height:28px; }
.xl-course::before, .xl-leader-zone::before, .xl-stage-lane::before, .xl-stage-card::before, .xl-role-zone::before, .xl-role-block::before, .xl-machine-lane::before, .xl-machine-card::before, .xl-machine-group-lane::before, .xl-machine-leader-lane::before { display:none !important; }
.xl-auto-connectors { z-index:2; }
.xl-auto-connector-path { stroke:rgba(189,244,255,.82); stroke-width:1.25; stroke-linecap:square; stroke-linejoin:miter; filter:drop-shadow(0 0 4px rgba(34,211,238,.30)); }
.xl-auto-connector-path.org-connector-soft { stroke:rgba(125,249,255,.50); stroke-width:1.0; filter:none; }
.xl-summary, .xl-exec-zone, .xl-root-card, .xl-course, .xl-course-title, .xl-course-count, .xl-leader-card, .xl-machine-card, .xl-machine-head, .xl-stage-card, .xl-role-block, .xl-person-row { z-index:5; }
.xl-empty-wide { grid-column:1 / -1; min-height:54px; }
body.org-editing .xl-machine-card { resize:none !important; overflow:visible !important; min-width:220px; min-height:120px; max-width:none !important; }
body.org-editing .xl-machine-lane, body.org-editing .xl-machine-group-lane { overflow:visible !important; }
body.org-editing .xl-machine-card:hover { outline:1px dashed rgba(0,112,192,.55); outline-offset:2px; }
body.org-editing .xl-machine-card::after { content:''; position:absolute; right:2px; bottom:2px; width:16px; height:16px; z-index:8; pointer-events:none; border-right:3px solid rgba(0,112,192,.75); border-bottom:3px solid rgba(0,112,192,.75); background:linear-gradient(135deg, transparent 0 42%, rgba(0,112,192,.32) 42% 54%, transparent 54% 100%); }

/* Dynamic large-canvas viewport: keep scroll range wider than the visible org chart. */
.xl-shell { height:__VIEWPORT_HEIGHT__px !important; min-height:__VIEWPORT_HEIGHT__px !important; max-height:__VIEWPORT_HEIGHT__px !important; overscroll-behavior:contain; }
.xl-org-canvas { min-width:__CANVAS_WIDTH__px !important; min-height:__CANVAS_HEIGHT__px !important; }
.xl-shell::-webkit-scrollbar { width:14px; height:14px; }
.xl-shell::-webkit-scrollbar-thumb { background:linear-gradient(135deg,rgba(34,211,238,.70),rgba(59,130,246,.48)); border-radius:999px; border:3px solid rgba(2,8,23,.95); }
.xl-shell::-webkit-scrollbar-track { background:rgba(15,23,42,.82); border-radius:999px; }
.canvas-chip { color:#EAF6FF !important; font-weight:950; border:1px solid rgba(34,211,238,.25); border-radius:999px; padding:4px 8px; background:rgba(255,255,255,.055); }
.zoom-label { min-width:46px; text-align:center; color:#EAF6FF !important; font-weight:950; border:1px solid rgba(34,211,238,.25); border-radius:999px; padding:4px 8px; background:rgba(255,255,255,.055); }
body.xl-fullscreen-active { background:#020817; overflow:hidden; }
body.xl-fullscreen-active .xl-toolbar { position:fixed; left:10px; right:10px; top:10px; z-index:1000000; margin:0; }
body.xl-fullscreen-active .xl-shell { position:fixed; left:0; top:0; width:100vw; height:100vh !important; min-height:100vh !important; max-height:100vh !important; padding-top:74px; border-radius:0; z-index:999999; overflow:auto !important; }
body.xl-fullscreen-active .xl-org-canvas { min-height:2100px; }
body.xl-fullscreen-active #fullscreenOrg { border-color:rgba(250,204,21,.75); box-shadow:0 0 18px rgba(250,204,21,.24); }

</style>
</head>
<body>
  <div class="xl-toolbar">
    <b>製造部階層式組織圖</b>
    <span>新版依專業階層排列：製造部 → 製一課/製二課 → 機型群 → 組長 → 組別/工段（橫向） → 組員，並自動併入 02. 派遣名單。</span>
    <span>啟動編輯後可拖曳任一卡片微調位置，也可按 ✎ 或雙擊修改卡片文字。</span>
    <span>正式資料來源仍為 01/02；編輯模式僅調整展示版面。</span>
    <span id="editState" class="edit-state">瀏覽模式</span>
    <button id="enableEdit" type="button">啟動編輯</button>
    <button id="disableEdit" type="button">關閉編輯</button>
    <button id="antiOverlapToggle" type="button">防重疊：關</button>
    <button id="autoLayout" type="button">重新依階層排列＋連接線</button>
    <button id="addHLine" type="button">＋橫線</button>
    <button id="addVLine" type="button">＋直線</button>
    <button id="deleteSelectedLine" type="button">刪除選取線</button>
    <button id="bringToFront" type="button">移到最上層</button>
    <button id="sendToBack" type="button">移到最下層</button>
    <button id="clearSelection" type="button">取消選取</button>
    <button id="toggleScrollbars" type="button">隱藏卷軸</button>
    <button id="zoomOut" type="button">－</button>
    <input id="zoomRange" type="range" min="20" max="180" value="__DEFAULT_ZOOM__" />
    <span id="zoomLabel" class="zoom-label">__DEFAULT_ZOOM__%</span>
    <button id="zoomIn" type="button">＋</button>
    <button id="resetZoom" type="button">縮放100%</button>
    <button id="fitScreen" type="button">適合視窗</button>
    <button id="fitAll" type="button">完整顯示</button>
    <button id="expandCanvas" type="button">畫布＋20%</button>
    <button id="shrinkCanvas" type="button">畫布－20%</button>
    <button id="topLeft" type="button">回到左上</button>
    <span id="canvasInfo" class="canvas-chip">畫布 __CANVAS_WIDTH__×__CANVAS_HEIGHT__</span>
    <button id="fullscreenOrg" type="button">全螢幕播放</button>
    <button id="saveLayout" type="button">保存目前版面於瀏覽器</button>
    <button id="exportPermanentLayout" type="button">複製永久記錄碼</button>
    <button id="resetLayout" type="button">重置版面</button>
  </div>
  <div class="xl-shell"><div id="layoutSource" style="display:none">__DEFAULT_INNER__</div><div id="layoutMount">__MOUNT_INNER__</div></div>
<script>
(function() {
  const KEY = '__STORAGE_KEY__';
  const HAS_SERVER_LAYOUT = __HAS_SERVER_LAYOUT__;
  const EXPORT_VERSION = 'spt_org_layout_v1';
  const mount = document.getElementById('layoutMount');
  const source = document.getElementById('layoutSource');
  const shell = document.querySelector('.xl-shell');
  let antiOverlapEnabled = false;
  let zoomValue = __DEFAULT_ZOOM__;
  const BASE_CANVAS_WIDTH = __CANVAS_WIDTH__;
  const BASE_CANVAS_HEIGHT = __CANVAS_HEIGHT__;
  const AUTO_EXPAND_CANVAS = __AUTO_EXPAND_CANVAS__;
  const saved = localStorage.getItem(KEY);
  if (!HAS_SERVER_LAYOUT && saved) mount.innerHTML = saved;
  ensureLineHandles();
  ensureResizeHandles();
  requestAnimationFrame(drawAutoConnectors);
  setTimeout(drawAutoConnectors, 250);
  let dragged = null;
  function isEditing() { return document.body.classList.contains('org-editing'); }
  function setEditing(enabled) {
    document.body.classList.toggle('org-editing', !!enabled);
    const state = document.getElementById('editState');
    if (state) state.textContent = enabled ? '編輯模式' : '瀏覽模式';
    if (enabled) ensureResizeHandles();
  }
  setEditing(false);
  requestAnimationFrame(function() { applyZoom(zoomValue); });
  function applyZoom(value) {
    zoomValue = Math.max(20, Math.min(180, Number(value) || __DEFAULT_ZOOM__));
    const range = document.getElementById('zoomRange');
    const label = document.getElementById('zoomLabel');
    if (range) range.value = String(zoomValue);
    if (label) label.textContent = String(Math.round(zoomValue)) + '%';
    if (mount) mount.style.zoom = String(zoomValue / 100);
    ensureScrollableCanvas();
    requestAnimationFrame(drawAutoConnectors);
  }

  function zoomScale() {
    return Math.max(0.35, Math.min(1.8, zoomValue / 100 || 1));
  }
  function canvasRelativeRect(el) {
    const canvas = mount.querySelector('.xl-org-canvas');
    if (!canvas || !el) return null;
    const c = canvas.getBoundingClientRect();
    const r = el.getBoundingClientRect();
    const z = zoomScale();
    return {
      left: (r.left - c.left) / z,
      top: (r.top - c.top) / z,
      right: (r.right - c.left) / z,
      bottom: (r.bottom - c.top) / z,
      width: r.width / z,
      height: r.height / z,
      cx: (r.left - c.left + r.width / 2) / z,
      cy: (r.top - c.top + r.height / 2) / z
    };
  }
  function updateCanvasInfo(width, height) {
    const info = document.getElementById('canvasInfo');
    if (info) info.textContent = '畫布 ' + Math.round(width) + '×' + Math.round(height);
  }
  function currentCanvasSize(canvas) {
    const cs = window.getComputedStyle(canvas);
    const w = parseFloat(canvas.style.minWidth || cs.minWidth || canvas.scrollWidth || BASE_CANVAS_WIDTH) || BASE_CANVAS_WIDTH;
    const h = parseFloat(canvas.style.minHeight || cs.minHeight || canvas.scrollHeight || BASE_CANVAS_HEIGHT) || BASE_CANVAS_HEIGHT;
    return { width: Math.max(BASE_CANVAS_WIDTH, w), height: Math.max(BASE_CANVAS_HEIGHT, h) };
  }
  function contentBounds() {
    const canvas = mount.querySelector('.xl-org-canvas');
    if (!canvas) return { width: BASE_CANVAS_WIDTH, height: BASE_CANVAS_HEIGHT };
    const crect = canvas.getBoundingClientRect();
    const z = zoomScale();
    let right = BASE_CANVAS_WIDTH;
    let bottom = BASE_CANVAS_HEIGHT;
    const selector = '.xl-summary,.xl-exec-zone,.xl-root-card,.xl-root-line,.xl-course-lane,.xl-course,.xl-course-title,.xl-course-count,.xl-leader-card,.xl-machine-card,.xl-machine-head,.xl-stage-card,.xl-role-block,.xl-person-row,.xl-line-node,[data-drag-type]';
    canvas.querySelectorAll(selector).forEach(el => {
      if (!el || el.offsetParent === null) return;
      const r = el.getBoundingClientRect();
      if (!r || !Number.isFinite(r.right) || !Number.isFinite(r.bottom)) return;
      right = Math.max(right, (r.right - crect.left) / z);
      bottom = Math.max(bottom, (r.bottom - crect.top) / z);
    });
    return { width: Math.ceil(right + 720), height: Math.ceil(bottom + 720) };
  }
  function ensureScrollableCanvas() {
    const canvas = mount.querySelector('.xl-org-canvas');
    if (!canvas) return;
    let size = currentCanvasSize(canvas);
    if (AUTO_EXPAND_CANVAS) {
      const bounds = contentBounds();
      size.width = Math.max(size.width, bounds.width, BASE_CANVAS_WIDTH);
      size.height = Math.max(size.height, bounds.height, BASE_CANVAS_HEIGHT);
    }
    canvas.style.minWidth = Math.round(size.width) + 'px';
    canvas.style.minHeight = Math.round(size.height) + 'px';
    updateCanvasInfo(size.width, size.height);
  }
  function adjustCanvas(factor) {
    const canvas = mount.querySelector('.xl-org-canvas');
    if (!canvas) return;
    const size = currentCanvasSize(canvas);
    const nextW = Math.max(BASE_CANVAS_WIDTH, Math.min(7000, Math.round(size.width * factor)));
    const nextH = Math.max(BASE_CANVAS_HEIGHT, Math.min(5200, Math.round(size.height * factor)));
    canvas.style.minWidth = nextW + 'px';
    canvas.style.minHeight = nextH + 'px';
    updateCanvasInfo(nextW, nextH);
    requestAnimationFrame(drawAutoConnectors);
  }
  function placeSystemConnectors() {
    // V15 uses only SVG hierarchy connectors. Legacy manual root lines are hidden to avoid crossing cards.
  }

  function anchorElement(el) {
    if (!el) return null;
    if (el.classList.contains('xl-course')) return el.querySelector(':scope > .xl-course-title') || el;
    if (el.classList.contains('xl-stage-card')) return el.querySelector(':scope > .xl-stage-head') || el;
    if (el.classList.contains('xl-role-block')) return el.querySelector(':scope > .xl-role-title') || el;
    return el;
  }
  function cardById(canvas, id) {
    if (!canvas || !id) return null;
    return Array.from(canvas.querySelectorAll('[data-card-id]')).find(el => el.dataset.cardId === id) || null;
  }
  function drawAutoConnectors() {
    const canvas = mount.querySelector('.xl-org-canvas');
    if (!canvas) return;
    ensureScrollableCanvas();
    const svg = canvas.querySelector('.xl-auto-connectors');
    if (!svg) return;
    const w = Math.max(canvas.scrollWidth || 0, canvas.offsetWidth || 0, 3000);
    const h = Math.max(canvas.scrollHeight || 0, canvas.offsetHeight || 0, 1800);
    svg.setAttribute('viewBox', '0 0 ' + w + ' ' + h);
    svg.setAttribute('width', String(w));
    svg.setAttribute('height', String(h));
    svg.innerHTML = '';
    const children = Array.from(canvas.querySelectorAll('[data-parent-card-id]'))
      .filter(el => el.dataset.parentCardId && isConnectorTarget(el) && el.offsetParent !== null);
    const groups = new Map();
    children.forEach(child => {
      const parent = cardById(canvas, child.dataset.parentCardId);
      if (!parent || parent === child || parent.offsetParent === null) return;
      const list = groups.get(child.dataset.parentCardId) || [];
      list.push(child);
      groups.set(child.dataset.parentCardId, list);
    });
    function drawPath(d, cls) {
      const path = document.createElementNS('http://www.w3.org/2000/svg', 'path');
      path.setAttribute('class', cls || 'xl-auto-connector-path');
      path.setAttribute('d', d);
      svg.appendChild(path);
    }
    groups.forEach((items, parentId) => {
      const parent = cardById(canvas, parentId);
      const pr = canvasRelativeRect(anchorElement(parent));
      if (!pr) return;
      const childRects = items.map(child => ({ child, rect: canvasRelativeRect(anchorElement(child)) })).filter(x => x.rect);
      if (!childRects.length) return;
      const below = childRects.filter(x => x.rect.cy >= pr.cy);
      const above = childRects.filter(x => x.rect.cy < pr.cy);
      [below, above].forEach((bucket, bucketIndex) => {
        if (!bucket.length) return;
        const isAbove = bucketIndex === 1;
        const centers = bucket.map(x => x.rect.cx);
        const startX = Math.round(pr.cx);
        const startY = Math.round(isAbove ? pr.top : pr.bottom);
        const endYs = bucket.map(x => isAbove ? x.rect.bottom : x.rect.top);
        const nearestEndY = isAbove ? Math.max.apply(null, endYs) : Math.min.apply(null, endYs);
        let busY = Math.round(isAbove ? Math.min(startY - 28, nearestEndY + 28) : Math.max(startY + 28, nearestEndY - 28));
        if (Math.abs(busY - startY) < 16) busY = startY + (isAbove ? -26 : 26);
        const left = Math.round(Math.min.apply(null, centers));
        const right = Math.round(Math.max.apply(null, centers));
        const cls = parent.classList.contains('xl-stage-card') || parent.classList.contains('xl-machine-card') ? 'xl-auto-connector-path org-connector-soft' : 'xl-auto-connector-path';
        drawPath('M ' + startX + ' ' + startY + ' V ' + busY, cls);
        if (right > left) drawPath('M ' + left + ' ' + busY + ' H ' + right, cls);
        bucket.forEach(item => {
          const endX = Math.round(item.rect.cx);
          const endY = Math.round(isAbove ? item.rect.bottom : item.rect.top);
          drawPath('M ' + endX + ' ' + busY + ' V ' + endY, cls);
        });
      });
    });
  }
  function isConnectorTarget(el) {
    if (!el) return false;
    // Professional org charts use trunk connectors only between structure nodes.
    // Member rows stay inside their role boxes; drawing one line per person makes the chart noisy.
    if (el.classList.contains('xl-person-row')) return false;
    if (el.classList.contains('xl-top-box')) return false;
    if (el.classList.contains('xl-empty')) return false;
    if (el.classList.contains('xl-role-block')) return false;
    return el.classList.contains('xl-course') || el.classList.contains('xl-machine-card') || el.classList.contains('xl-stage-card') || el.classList.contains('xl-leader-card');
  }
  function resetHierarchyFlowLayout() {
    const cleanSelectors = '.xl-summary,.xl-exec-zone,.xl-course-lane,.xl-course,.xl-machine-card,.xl-stage-card,.xl-role-block,.xl-person-row,.xl-leader-card,.xl-top-box,.xl-root-card';
    mount.querySelectorAll(cleanSelectors).forEach(el => {
      el.dataset.offsetX = '0';
      el.dataset.offsetY = '0';
      el.style.transform = '';
      el.classList.remove('manually-positioned');
      if (!el.classList.contains('xl-summary') && !el.classList.contains('xl-exec-zone') && !el.classList.contains('xl-course-lane') && !el.classList.contains('xl-root-card')) {
        el.style.width = '';
        el.style.height = '';
      }
      el.style.zIndex = '';
    });
    mount.querySelectorAll('.xl-line-node').forEach(line => {
      if (!line.classList.contains('xl-custom-line')) {
        line.dataset.offsetX = '0';
        line.dataset.offsetY = '0';
        line.style.transform = '';
        line.classList.remove('manually-positioned');
      }
    });
    requestAnimationFrame(drawAutoConnectors);
  }
  function lineHtml(axis) {
    const horizontal = axis !== 'vertical';
    const style = horizontal ? 'left:120px;top:220px;width:360px;height:1px;' : 'left:220px;top:160px;width:1px;height:220px;';
    const cls = 'xl-custom-line xl-line-node';
    return '<div class="' + cls + '" draggable="true" data-drag-type="line" data-card-id="' + cardId(horizontal ? 'hline' : 'vline') + '" data-line-axis="' + (horizontal ? 'horizontal' : 'vertical') + '" style="position:absolute;' + style + '" title="編輯模式可拖拉、拉長、拉短或刪除連接線；拖曳端點可與其他線端對接">' + lineHandles() + '</div>';
  }
  function lineHandles() {
    return '<span class="xl-line-handle start" data-line-handle="start"></span><span class="xl-line-handle end" data-line-handle="end"></span>';
  }
  function ensureLineHandles() {
    mount.querySelectorAll('.xl-line-node').forEach(line => {
      if (!line.querySelector('.xl-line-handle')) line.insertAdjacentHTML('beforeend', lineHandles());
    });
  }
  function ensureResizeHandles() {
    const selector = '.xl-machine-card,.xl-course,.xl-stage-card,.xl-role-block';
    mount.querySelectorAll(selector).forEach(card => {
      if (!card.querySelector(':scope > .xl-card-resize-handle')) {
        const handle = document.createElement('span');
        handle.className = 'xl-card-resize-handle';
        handle.title = '拖曳調整外框大小；調整後請按「保存目前版面於瀏覽器」或「複製永久記錄碼」。';
        handle.setAttribute('data-resize-handle', 'card');
        card.appendChild(handle);
      }
    });
  }
  const selectionBox = document.createElement('div');
  selectionBox.className = 'xl-selection-box';
  document.body.appendChild(selectionBox);
  function cardId(prefix) { return prefix + '_' + Date.now().toString(36) + '_' + Math.random().toString(36).slice(2, 7); }
  function esc(s) { return String(s || '').replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])); }
  function delBtn() { return '<button class="dim-edit" type="button" title="修改卡片內容">✎</button><button class="dim-delete" type="button" title="刪除此卡片">×</button>'; }
  function readOffset(card) {
    return { x: parseFloat(card.dataset.offsetX || '0') || 0, y: parseFloat(card.dataset.offsetY || '0') || 0 };
  }
  function isResizeGesture(e, el) {
    if (!el || !isEditing()) return false;
    return !!(e && e.target && e.target.closest && e.target.closest('.xl-card-resize-handle'));
  }
  function applyOffset(card, x, y) {
    card.dataset.offsetX = String(Math.round(x));
    card.dataset.offsetY = String(Math.round(y));
    card.style.transform = 'translate(' + Math.round(x) + 'px,' + Math.round(y) + 'px)';
    card.classList.add('manually-positioned');
    requestAnimationFrame(drawAutoConnectors);
  }
  function cardMainNode(card) {
    if (!card) return null;
    if (card.classList.contains('xl-course')) return card.querySelector(':scope > .xl-course-title b');
    if (card.classList.contains('xl-machine-card')) return card.querySelector(':scope > .xl-machine-head b');
    if (card.classList.contains('xl-stage-card')) return card.querySelector(':scope > .xl-stage-head b');
    if (card.classList.contains('xl-role-block')) return card.querySelector(':scope > .xl-role-title b');
    return card.querySelector('b');
  }
  function cardSubNode(card) {
    if (!card) return null;
    if (card.classList.contains('xl-course')) return card.querySelector(':scope > .xl-course-title em');
    if (card.classList.contains('xl-machine-card')) return card.querySelector(':scope > .xl-machine-head em');
    if (card.classList.contains('xl-stage-card')) return card.querySelector(':scope > .xl-stage-head em');
    if (card.classList.contains('xl-role-block')) return card.querySelector(':scope > .xl-role-title em');
    return card.querySelector('em, small');
  }
  function editCard(card) {
    if (!card || !isEditing()) return;
    const main = cardMainNode(card);
    const sub = cardSubNode(card);
    const currentMain = main ? main.textContent.trim() : '';
    const nextMain = prompt('請修改卡片主內容（例如姓名、課別、工段或職稱）', currentMain);
    if (nextMain === null) return;
    if (main) main.textContent = nextMain.trim() || currentMain;
    if (sub) {
      const currentSub = sub.textContent.trim();
      const nextSub = prompt('請修改卡片副內容（例如職稱、機型、來源或備註）', currentSub);
      if (nextSub !== null) sub.textContent = nextSub.trim();
    }
  }
  function courseHtml(value) { const id = cardId('course'); return '<section class="xl-course" draggable="true" data-drag-type="course" data-card-id="' + id + '" data-parent-card-id="department_root"><div class="xl-course-count">0員</div><div class="xl-course-title"><span class="xl-grip">☰</span><b>' + esc(value) + '</b><em>自訂課別</em><button class="xl-add" type="button" data-add-type="machine">＋機型</button>' + delBtn() + '</div><div class="xl-leader-zone xl-drop-zone" data-accept="person"></div><div class="xl-machine-lane xl-drop-zone" data-accept="machine"></div></section>'; }
  function machineHtml(value, parentId) { const id = cardId('machine'); return '<section class="xl-machine-card" draggable="true" data-drag-type="machine" data-card-id="' + id + '" data-parent-card-id="' + esc(parentId || '') + '"><div class="xl-machine-head"><span class="xl-grip">☰</span><b>' + esc(value) + '</b><em>自訂機型群</em><button class="xl-add" type="button" data-add-type="stage">＋組別</button>' + delBtn() + '</div><div class="xl-machine-group-lane xl-drop-zone" data-accept="stage"></div></section>'; }
  function stageHtml(value, parentId) { return '<section class="xl-stage-card" draggable="true" data-drag-type="stage" data-card-id="' + cardId('stage') + '" data-parent-card-id="' + esc(parentId || '') + '"><div class="xl-stage-head"><span class="xl-grip">☰</span><b>' + esc(value) + '</b><em>自訂組別 / 工段</em><button class="xl-add" type="button" data-add-type="role">＋職稱</button>' + delBtn() + '</div><div class="xl-stage-leaders xl-drop-zone" data-accept="person"></div><div class="xl-role-zone xl-drop-zone" data-accept="role"></div></section>'; }
  function roleHtml(title, source) { return '<article class="xl-role-block role-general" draggable="true" data-drag-type="role" data-card-id="' + cardId('role') + '"><div class="xl-role-title"><span class="xl-grip">☰</span><b>' + esc(title) + '</b><em>' + esc(source || '自訂來源') + '</em><button class="xl-add" type="button" data-add-type="person">＋人員</button>' + delBtn() + '</div><div class="xl-person-zone xl-drop-zone" data-accept="person"></div></article>'; }
  function personHtml(name, title) { return '<div class="xl-person-row role-general" draggable="true" data-drag-type="person" data-card-id="' + cardId('person') + '"><span class="xl-grip">⋮</span><b>' + esc(name) + '</b><small>' + esc(title || '人員') + '</small>' + delBtn() + '</div>'; }
  function canDrop(card, zone) { return card && zone && card.dataset.dragType === zone.dataset.accept; }
  function zIndexValue(el) {
    const value = parseInt(window.getComputedStyle(el).zIndex, 10);
    return Number.isFinite(value) ? value : 0;
  }
  function setLayerPosition(mode) {
    const targets = selectedCards().filter(el => el && el.dataset && el.dataset.dragType);
    if (!targets.length) { alert('請先點選一個或多個物件，再調整最上層 / 最下層。'); return; }
    const targetSet = new Set(targets);
    const all = movableCards().filter(el => el.offsetParent !== null);
    const values = all.map(zIndexValue);
    if (mode === 'front') {
      let z = Math.max(50, ...values) + 10;
      targets.forEach(el => { el.style.zIndex = String(z++); el.dataset.layerZ = el.style.zIndex; });
      alert('已將選取物件移到最上層。請確認後按「保存目前版面於瀏覽器」或「複製永久記錄碼」。');
    } else {
      // Do not use negative z-index. Negative values can push cards behind the canvas
      // background and make them look lost. Put selected cards on a low positive layer
      // and normalize other cards above them, so the objects stay visible and selectable.
      let bottomZ = 5;
      targets.forEach(el => { el.style.zIndex = String(bottomZ++); el.dataset.layerZ = el.style.zIndex; });
      let otherZ = 30;
      all.forEach(el => {
        if (targetSet.has(el)) return;
        const current = zIndexValue(el);
        if (current < 30) {
          el.style.zIndex = String(otherZ++);
          el.dataset.layerZ = el.style.zIndex;
        }
      });
      alert('已將選取物件移到最下層，但仍保留在可見圖層，不會再消失。請確認後按「保存目前版面於瀏覽器」或「複製永久記錄碼」。');
    }
  }
  function fitScreen() {
    const canvas = mount ? mount.querySelector('.xl-org-canvas') : null;
    if (!canvas || !shell) { applyZoom(__DEFAULT_ZOOM__); return; }
    ensureScrollableCanvas();
    const shellWidth = Math.max(520, shell.clientWidth - 54);
    const shellHeight = Math.max(520, shell.clientHeight - 94);
    const bounds = contentBounds();
    const pct = Math.floor(Math.min(shellWidth / Math.max(bounds.width, 1), shellHeight / Math.max(bounds.height, 1)) * 100);
    applyZoom(Math.max(20, Math.min(120, pct || __DEFAULT_ZOOM__)));
    shell.scrollTo({ left: 0, top: 0, behavior: 'smooth' });
  }
  function fitAll() {
    fitScreen();
  }
  function toggleFullscreen() {
    const active = document.body.classList.toggle('xl-fullscreen-active');
    const btn = document.getElementById('fullscreenOrg');
    if (btn) btn.textContent = active ? '離開全螢幕' : '全螢幕播放';
    if (active && document.documentElement.requestFullscreen) {
      document.documentElement.requestFullscreen().catch(function() {});
    } else if (!active && document.fullscreenElement && document.exitFullscreen) {
      document.exitFullscreen().catch(function() {});
    }
    setTimeout(function() { requestAnimationFrame(drawAutoConnectors); }, 120);
  }
  document.addEventListener('fullscreenchange', function() {
    if (!document.fullscreenElement && document.body.classList.contains('xl-fullscreen-active')) {
      document.body.classList.remove('xl-fullscreen-active');
      const btn = document.getElementById('fullscreenOrg');
      if (btn) btn.textContent = '全螢幕播放';
      requestAnimationFrame(drawAutoConnectors);
    }
  });
  function cardText(el) {
    if (!el) return '';
    const main = cardMainNode(el);
    const sub = cardSubNode(el);
    const mainText = main ? main.textContent.trim() : '';
    const subText = sub ? sub.textContent.trim() : '';
    return (mainText + (subText ? '｜' + subText : '')).replace(/[\\t\\r\\n ]+/g, ' ');
  }
  // Excel export is handled by Streamlit server-side download button to guarantee a valid .xlsx file.
    function exportPermanentLayout() {
    if (!isEditing()) return;
    persistSizes();
    const payload = JSON.stringify({ version: EXPORT_VERSION, exportedAt: new Date().toISOString(), html: mount.innerHTML });
    function fallbackCopy() {
      const box = document.createElement('textarea');
      box.value = payload;
      box.style.position = 'fixed';
      box.style.left = '12px';
      box.style.top = '12px';
      box.style.width = '90vw';
      box.style.height = '160px';
      box.style.zIndex = '999999';
      document.body.appendChild(box);
      box.focus();
      box.select();
      alert('已產生永久記錄碼。若瀏覽器沒有自動複製，請手動複製畫面左上方文字框內容，再貼到 Streamlit 下方「永久記錄」區塊儲存。');
    }
    if (navigator.clipboard && navigator.clipboard.writeText) {
      navigator.clipboard.writeText(payload)
        .then(() => alert('已複製永久記錄碼。請貼到 Streamlit 下方「永久記錄」區塊，按「儲存永久記錄」。'))
        .catch(fallbackCopy);
    } else {
      fallbackCopy();
    }
  }
  document.addEventListener('click', function(e) {
    const edit = e.target.closest('.dim-edit');
    if (edit) {
      e.preventDefault();
      if (!isEditing()) return;
      editCard(edit.closest('[data-drag-type]'));
      return;
    }
    const del = e.target.closest('.dim-delete');
    if (del) {
      e.preventDefault();
      if (!isEditing()) return;
      const card = del.closest('[data-drag-type]');
      if (card && confirm('確定刪除此卡片？這只會影響組織圖展示版面，不會刪除 01/02 權威資料。')) { card.remove(); requestAnimationFrame(drawAutoConnectors); }
      return;
    }
    const add = e.target.closest('[data-add-type]');
    if (add) {
      e.preventDefault();
      if (!isEditing()) return;
      const type = add.dataset.addType;
      if (type === 'course') {
        const value = prompt('請輸入課別名稱，例如：製一課');
        if (!value) return;
        const zone = mount.querySelector('.xl-course-lane');
        if (zone) { zone.insertAdjacentHTML('beforeend', courseHtml(value)); requestAnimationFrame(drawAutoConnectors); }
        return;
      }
      if (type === 'machine') {
        const value = prompt('請輸入機型群，例如：Sorter / EFEM、BWBS、NTB、GPTC');
        if (!value) return;
        const parent = add.closest('.xl-course');
        const zone = parent ? parent.querySelector(':scope > .xl-machine-lane') : null;
        if (zone) { zone.insertAdjacentHTML('beforeend', machineHtml(value, parent.dataset.cardId)); ensureResizeHandles(); requestAnimationFrame(drawAutoConnectors); }
        return;
      }
      if (type === 'stage') {
        const value = prompt('請輸入組別 / 工段，例如：模組、S.T.、前置鈑金、機電');
        if (!value) return;
        const parent = add.closest('.xl-machine-card') || add.closest('.xl-course');
        const zone = parent ? (parent.querySelector(':scope > .xl-machine-group-lane') || parent.querySelector(':scope > .xl-stage-lane')) : null;
        if (zone) { zone.insertAdjacentHTML('beforeend', stageHtml(value, parent.dataset.cardId)); ensureResizeHandles(); requestAnimationFrame(drawAutoConnectors); }
        return;
      }
      if (type === 'role') {
        const title = prompt('請輸入職稱，例如：工程師、助理工程師、派遣');
        if (!title) return;
        const source = prompt('請輸入人力來源，例如：超慧正職、派遣/外包', '超慧正職') || '超慧正職';
        const parent = add.closest('.xl-stage-card');
        const zone = parent ? parent.querySelector(':scope > .xl-role-zone') : null;
        if (zone) { zone.insertAdjacentHTML('beforeend', roleHtml(title, source)); requestAnimationFrame(drawAutoConnectors); }
        return;
      }
      if (type === 'person') {
        const name = prompt('請輸入姓名');
        if (!name) return;
        const title = prompt('請輸入職稱', '人員') || '人員';
        const parent = add.closest('.xl-role-block');
        const zone = parent ? parent.querySelector(':scope > .xl-person-zone') : null;
        if (zone) { zone.insertAdjacentHTML('beforeend', personHtml(name, title)); requestAnimationFrame(drawAutoConnectors); }
        return;
      }
    }
    if (e.target.id === 'antiOverlapToggle') {
      if (!isEditing()) return;
      antiOverlapEnabled = !antiOverlapEnabled;
      e.target.textContent = antiOverlapEnabled ? '防重疊：開' : '防重疊：關';
      return;
    }
    if (e.target.id === 'addHLine') {
      if (!isEditing()) return;
      const canvas = mount.querySelector('.xl-org-canvas');
      if (canvas) canvas.insertAdjacentHTML('beforeend', lineHtml('horizontal'));
      return;
    }
    if (e.target.id === 'addVLine') {
      if (!isEditing()) return;
      const canvas = mount.querySelector('.xl-org-canvas');
      if (canvas) canvas.insertAdjacentHTML('beforeend', lineHtml('vertical'));
      return;
    }
    if (e.target.id === 'deleteSelectedLine') {
      if (!isEditing()) return;
      const selectedLines = Array.from(mount.querySelectorAll('.xl-line-node.xl-selected'));
      if (!selectedLines.length) { alert('請先點選要刪除的黑線。'); return; }
      if (confirm('確定刪除選取的連接線？')) selectedLines.forEach(line => line.remove());
      return;
    }
    if (e.target.id === 'bringToFront') { if (!isEditing()) return; setLayerPosition('front'); return; }
    if (e.target.id === 'sendToBack') { if (!isEditing()) return; setLayerPosition('back'); return; }
    if (e.target.id === 'exportPermanentLayout') { exportPermanentLayout(); return; }
    if (e.target.id === 'toggleScrollbars') {
      if (!shell) return;
      shell.classList.toggle('hide-scrollbars');
      e.target.textContent = shell.classList.contains('hide-scrollbars') ? '顯示卷軸' : '隱藏卷軸';
      return;
    }
    if (e.target.id === 'zoomOut') { applyZoom(zoomValue - 10); return; }
    if (e.target.id === 'zoomIn') { applyZoom(zoomValue + 10); return; }
    if (e.target.id === 'resetZoom') { applyZoom(100); return; }
    if (e.target.id === 'fitScreen') { fitScreen(); return; }
    if (e.target.id === 'fitAll') { fitAll(); return; }
    if (e.target.id === 'expandCanvas') { adjustCanvas(1.2); return; }
    if (e.target.id === 'shrinkCanvas') { adjustCanvas(0.84); return; }
    if (e.target.id === 'topLeft') { if (shell) shell.scrollTo({ left:0, top:0, behavior:'smooth' }); return; }
    if (e.target.id === 'fullscreenOrg') { toggleFullscreen(); return; }
    if (e.target.id === 'saveLayout') {
      if (!isEditing()) return;
      persistSizes();
      drawAutoConnectors();
      localStorage.setItem(KEY, mount.innerHTML);
      alert('已保存目前組織圖展示版面到此瀏覽器。');
    }
    if (e.target.id === 'resetLayout') {
      if (!isEditing()) return;
      if (confirm('確定重置為系統依 01/02 權威資料產生的版面？')) {
        localStorage.removeItem(KEY);
        mount.innerHTML = source.innerHTML;
        ensureLineHandles();
        ensureResizeHandles();
        resetHierarchyFlowLayout();
        drawAutoConnectors();
        clearSelected();
      }
    }
    if (e.target.id === 'autoLayout') {
      if (!isEditing()) return;
      clearSelected();
      antiOverlapEnabled = false;
      const antiBtn = document.getElementById('antiOverlapToggle');
      if (antiBtn) antiBtn.textContent = '防重疊：關';
      localStorage.removeItem(KEY);
      mount.innerHTML = source.innerHTML;
      ensureLineHandles();
      ensureResizeHandles();
      resetHierarchyFlowLayout();
      drawAutoConnectors();
      alert('已重新依專業階層排列並重繪連接線。若下方卡片看不到，可用縮放、適合視窗或全螢幕播放。');
    }
    if (e.target.id === 'clearSelection') clearSelected();
    if (e.target.id === 'enableEdit') setEditing(true);
    if (e.target.id === 'disableEdit') { clearSelected(); setEditing(false); }
  });
  document.addEventListener('input', function(e) {
    if (e.target && e.target.id === 'zoomRange') applyZoom(e.target.value);
  });
  window.addEventListener('resize', function() { requestAnimationFrame(drawAutoConnectors); });
  document.addEventListener('dblclick', function(e) {
    if (!isEditing()) return;
    if (e.target.closest('button')) return;
    const card = e.target.closest('[data-drag-type]');
    if (!card) return;
    e.preventDefault();
    editCard(card);
  });
  let moveState = null;
  let resizeState = null;
  let selectState = null;
  function movableCards() {
    return Array.from(mount.querySelectorAll('[data-drag-type]'));
  }
  function layoutCards() {
    return movableCards().filter(el => el.dataset.dragType !== 'line' && !el.classList.contains('xl-summary'));
  }
  function isNestedRelated(a, b) {
    return !!(a && b && (a.contains(b) || b.contains(a)));
  }
  function dragDistance(state, e) {
    if (!state || !e) return 0;
    const dx = e.clientX - state.startX;
    const dy = e.clientY - state.startY;
    return Math.sqrt(dx * dx + dy * dy);
  }
  function selectedCards() {
    return Array.from(mount.querySelectorAll('.xl-selected'));
  }
  function clearSelected() {
    mount.querySelectorAll('.xl-selected').forEach(el => el.classList.remove('xl-selected'));
  }
  function addSelected(el, append) {
    if (!append) clearSelected();
    if (el) el.classList.add('xl-selected');
  }
  function rectsOverlap(a, b, gap) {
    gap = gap || 8;
    return !(a.right + gap < b.left || a.left > b.right + gap || a.bottom + gap < b.top || a.top > b.bottom + gap);
  }
  function persistSizes() {
    mount.querySelectorAll('.xl-course,.xl-machine-card,.xl-stage-card,.xl-role-block,.xl-person-row,.xl-leader-card,.xl-top-box,.xl-root-card,.xl-line-node').forEach(el => {
      const w = el.offsetWidth || 0;
      const h = el.offsetHeight || 0;
      if (w > 0) el.style.width = Math.round(w) + 'px';
      if (h > 0) el.style.height = Math.round(h) + 'px';
    });
  }
  function nudgeOutOfOverlap(card) {
    if (!card || card.dataset.dragType === 'line') return;
    let tries = 0;
    let moved = false;
    while (tries < 24) {
      tries += 1;
      const cr = card.getBoundingClientRect();
      const hit = layoutCards().find(other => other !== card && other.offsetParent !== null && !isNestedRelated(card, other) && rectsOverlap(cr, other.getBoundingClientRect(), 6));
      if (!hit) break;
      const current = readOffset(card);
      const hr = hit.getBoundingClientRect();
      const shiftY = Math.max(16, hr.bottom - cr.top + 10);
      const shiftX = tries % 3 === 0 ? 24 : 0;
      applyOffset(card, current.x + shiftX, current.y + shiftY);
      moved = true;
    }
    return moved;
  }
  function autoArrangeOverlaps(resetSmallOffsets) {
    const cards = layoutCards().filter(el => el.offsetParent !== null);
    if (resetSmallOffsets) {
      cards.forEach(el => {
        const off = readOffset(el);
        if (Math.abs(off.x) < 3 && Math.abs(off.y) < 3) applyOffset(el, 0, 0);
      });
    }
    cards.sort((a,b) => {
      const ar = a.getBoundingClientRect();
      const br = b.getBoundingClientRect();
      return (ar.top - br.top) || (ar.left - br.left);
    });
    for (let pass = 0; pass < 4; pass++) {
      cards.forEach(card => nudgeOutOfOverlap(card));
    }
  }
  function finishSelection() {
    if (!selectState) return;
    const boxRect = selectionBox.getBoundingClientRect();
    clearSelected();
    layoutCards().forEach(card => {
      if (rectsOverlap(boxRect, card.getBoundingClientRect(), -2)) card.classList.add('xl-selected');
    });
    selectionBox.style.display = 'none';
    selectState = null;
  }
  function updateSelectionBox(e) {
    if (!selectState) return;
    const x1 = Math.min(selectState.startX, e.clientX);
    const y1 = Math.min(selectState.startY, e.clientY);
    const x2 = Math.max(selectState.startX, e.clientX);
    const y2 = Math.max(selectState.startY, e.clientY);
    selectionBox.style.left = x1 + 'px';
    selectionBox.style.top = y1 + 'px';
    selectionBox.style.width = Math.max(2, x2 - x1) + 'px';
    selectionBox.style.height = Math.max(2, y2 - y1) + 'px';
    selectionBox.style.display = 'block';
  }

  let lineHandleState = null;
  function canvasRect() {
    const canvas = mount.querySelector('.xl-org-canvas') || mount;
    return canvas.getBoundingClientRect();
  }
  function px(v, fallback) {
    const n = parseFloat(String(v || '').replace('px',''));
    return Number.isFinite(n) ? n : (fallback || 0);
  }
  function getLineGeom(line) {
    return {
      left: px(line.style.left, line.offsetLeft),
      top: px(line.style.top, line.offsetTop),
      width: Math.max(1, px(line.style.width, line.offsetWidth)),
      height: Math.max(1, px(line.style.height, line.offsetHeight)),
      axis: line.dataset.lineAxis || 'horizontal'
    };
  }
  function setLineGeom(line, geom) {
    line.style.left = Math.round(geom.left) + 'px';
    line.style.top = Math.round(geom.top) + 'px';
    line.style.width = Math.max(1, Math.round(geom.width)) + 'px';
    line.style.height = Math.max(1, Math.round(geom.height)) + 'px';
    line.style.transform = '';
    line.dataset.offsetX = '0';
    line.dataset.offsetY = '0';
  }
  function lineEndpoint(line, which) {
    const g = getLineGeom(line);
    if (g.axis === 'vertical') {
      return which === 'start' ? {x:g.left, y:g.top} : {x:g.left, y:g.top + g.height};
    }
    return which === 'start' ? {x:g.left, y:g.top} : {x:g.left + g.width, y:g.top};
  }
  function allLineEndpoints(exceptLine) {
    const out = [];
    mount.querySelectorAll('.xl-line-node').forEach(line => {
      if (line === exceptLine) return;
      out.push({ line, which:'start', point:lineEndpoint(line, 'start') });
      out.push({ line, which:'end', point:lineEndpoint(line, 'end') });
    });
    return out;
  }
  function snapLineHandle(line, handleName, maxDist) {
    maxDist = maxDist || 18;
    const current = lineEndpoint(line, handleName);
    let best = null;
    allLineEndpoints(line).forEach(item => {
      const dx = item.point.x - current.x;
      const dy = item.point.y - current.y;
      const d = Math.sqrt(dx*dx + dy*dy);
      if (d <= maxDist && (!best || d < best.d)) best = { d, point:item.point };
    });
    if (!best) return;
    const g = getLineGeom(line);
    if (g.axis === 'vertical') {
      if (handleName === 'start') {
        const oldEnd = g.top + g.height;
        g.left = best.point.x;
        g.top = best.point.y;
        g.height = Math.max(1, oldEnd - g.top);
      } else {
        g.left = best.point.x;
        g.height = Math.max(1, best.point.y - g.top);
      }
    } else {
      if (handleName === 'start') {
        const oldEnd = g.left + g.width;
        g.left = best.point.x;
        g.top = best.point.y;
        g.width = Math.max(1, oldEnd - g.left);
      } else {
        g.top = best.point.y;
        g.width = Math.max(1, best.point.x - g.left);
      }
    }
    setLineGeom(line, g);
  }
  function cardMinSize(card) {
    if (!card) return { width: 120, height: 80 };
    if (card.classList.contains('xl-course')) return { width: 520, height: 260 };
    if (card.classList.contains('xl-machine-card')) return { width: 240, height: 130 };
    if (card.classList.contains('xl-stage-card')) return { width: 180, height: 100 };
    if (card.classList.contains('xl-role-block')) return { width: 150, height: 70 };
    return { width: 120, height: 60 };
  }
  function startCardResize(e, handle) {
    const card = handle ? handle.closest('.xl-machine-card,.xl-course,.xl-stage-card,.xl-role-block') : null;
    if (!card) return false;
    e.preventDefault();
    e.stopPropagation();
    addSelected(card, false);
    const min = cardMinSize(card);
    resizeState = {
      card: card,
      startX: e.clientX,
      startY: e.clientY,
      baseWidth: Math.max(min.width, card.offsetWidth || card.getBoundingClientRect().width),
      baseHeight: Math.max(min.height, card.offsetHeight || card.getBoundingClientRect().height),
      minWidth: min.width,
      minHeight: min.height
    };
    card.classList.add('xl-card-resizing');
    return true;
  }
  document.addEventListener('pointerdown', function(e) {
    if (!isEditing()) return;
    if (e.button !== 0) return;
    if (e.target.closest('button,input,textarea,select')) return;
    const lineHandle = e.target.closest('.xl-line-handle');
    if (lineHandle) {
      const line = lineHandle.closest('.xl-line-node');
      if (!line) return;
      e.preventDefault();
      addSelected(line, false);
      lineHandleState = { line, handle: lineHandle.dataset.lineHandle || 'end', startX: e.clientX, startY: e.clientY, geom: getLineGeom(line), canvas: canvasRect() };
      return;
    }
    const resizeHandle = e.target.closest('.xl-card-resize-handle');
    if (resizeHandle && startCardResize(e, resizeHandle)) return;
    const card = e.target.closest('[data-drag-type]');
    if (!card) {
      if (e.target.closest('.xl-org-canvas,.xl-shell')) {
        selectState = { startX: e.clientX, startY: e.clientY };
        updateSelectionBox(e);
      }
      return;
    }
    if (isResizeGesture(e, card)) {
      // 讓瀏覽器原生 resize handle 接手，不啟動拖曳，否則卡片外框無法被拉大。
      return;
    }
    e.preventDefault();
    const selected = selectedCards();
    let targets = [];
    if (card.classList.contains('xl-selected') && selected.length > 1) {
      targets = selected;
    } else {
      addSelected(card, e.shiftKey || e.metaKey || e.ctrlKey);
      targets = selectedCards();
    }
    moveState = {
      cards: targets,
      startX: e.clientX,
      startY: e.clientY,
      base: targets.map(c => ({ card: c, offset: readOffset(c) })),
      didMove: false
    };
    targets.forEach(c => c.classList.add('dragging'));
  });
  document.addEventListener('pointermove', function(e) {
    if (resizeState) {
      e.preventDefault();
      const zoom = zoomValue / 100 || 1;
      const dx = (e.clientX - resizeState.startX) / zoom;
      const dy = (e.clientY - resizeState.startY) / zoom;
      const nextW = Math.max(resizeState.minWidth, Math.round(resizeState.baseWidth + dx));
      const nextH = Math.max(resizeState.minHeight, Math.round(resizeState.baseHeight + dy));
      resizeState.card.style.width = nextW + 'px';
      resizeState.card.style.height = nextH + 'px';
      resizeState.card.dataset.manualWidth = String(nextW);
      resizeState.card.dataset.manualHeight = String(nextH);
      ensureScrollableCanvas();
      requestAnimationFrame(drawAutoConnectors);
      return;
    }
    if (lineHandleState) {
      e.preventDefault();
      const st = lineHandleState;
      const zoom = zoomValue / 100 || 1;
      const dx = (e.clientX - st.startX) / zoom;
      const dy = (e.clientY - st.startY) / zoom;
      const g = Object.assign({}, st.geom);
      if (g.axis === 'vertical') {
        if (st.handle === 'start') {
          g.left = st.geom.left + dx;
          g.top = st.geom.top + dy;
          g.height = Math.max(1, st.geom.height - dy);
        } else {
          g.left = st.geom.left + dx;
          g.height = Math.max(1, st.geom.height + dy);
        }
      } else {
        if (st.handle === 'start') {
          g.left = st.geom.left + dx;
          g.top = st.geom.top + dy;
          g.width = Math.max(1, st.geom.width - dx);
        } else {
          g.top = st.geom.top + dy;
          g.width = Math.max(1, st.geom.width + dx);
        }
      }
      setLineGeom(st.line, g);
      return;
    }
    if (selectState) {
      e.preventDefault();
      updateSelectionBox(e);
      return;
    }
    if (!moveState) return;
    const distance = dragDistance(moveState, e);
    if (distance < 6) return;
    moveState.didMove = true;
    e.preventDefault();
    const dx = e.clientX - moveState.startX;
    const dy = e.clientY - moveState.startY;
    moveState.base.forEach(item => {
      const snap = 5;
      const nx = Math.round((item.offset.x + dx) / snap) * snap;
      const ny = Math.round((item.offset.y + dy) / snap) * snap;
      applyOffset(item.card, nx, ny);
    });
  });
  document.addEventListener('pointerup', function() {
    if (resizeState) {
      resizeState.card.classList.remove('xl-card-resizing');
      resizeState = null;
      requestAnimationFrame(drawAutoConnectors);
      return;
    }
    if (lineHandleState) {
      snapLineHandle(lineHandleState.line, lineHandleState.handle, 22);
      lineHandleState = null;
      return;
    }
    if (selectState) {
      finishSelection();
      return;
    }
    if (!moveState) return;
    const didMove = !!moveState.didMove;
    moveState.cards.forEach(c => c.classList.remove('dragging'));
    if (didMove && antiOverlapEnabled) moveState.cards.forEach(c => nudgeOutOfOverlap(c));
    moveState = null;
    requestAnimationFrame(drawAutoConnectors);
  });
  document.addEventListener('dragstart', function(e) {
    // 使用 pointer 直接微調卡片位置；停用 HTML5 原生拖放，避免卡片被搬到錯誤區塊造成重疊。
    e.preventDefault();
    return false;
  });
  document.addEventListener('dragend', function() {
    if (dragged) dragged.classList.remove('dragging');
    dragged = null;
    document.querySelectorAll('.drop-active').forEach(z => z.classList.remove('drop-active'));
  });
  document.addEventListener('dragover', function(e) {
    if (!isEditing()) return;
    const zone = e.target.closest('.xl-drop-zone');
    if (!zone || !canDrop(dragged, zone)) return;
    e.preventDefault();
    zone.classList.add('drop-active');
  });
  document.addEventListener('dragleave', function(e) {
    const zone = e.target.closest('.xl-drop-zone');
    if (zone) zone.classList.remove('drop-active');
  });
  document.addEventListener('drop', function(e) {
    if (!isEditing()) return;
    const zone = e.target.closest('.xl-drop-zone');
    if (!zone || !canDrop(dragged, zone)) return;
    e.preventDefault();
    zone.classList.remove('drop-active');
    zone.appendChild(dragged);
  });
})();
</script>
</body>
</html>'''
    return (
        template.replace("__DEFAULT_INNER__", inner)
        .replace("__MOUNT_INNER__", mount_inner)
        .replace("__STORAGE_KEY__", storage_key)
        .replace("__HAS_SERVER_LAYOUT__", has_server_layout)
        .replace("__VIEWPORT_HEIGHT__", str(viewport_height_i))
        .replace("__CANVAS_WIDTH__", str(canvas_width_i))
        .replace("__CANVAS_HEIGHT__", str(canvas_height_i))
        .replace("__DEFAULT_ZOOM__", str(default_zoom_i))
        .replace("__AUTO_EXPAND_CANVAS__", auto_expand_canvas_js)
    )
