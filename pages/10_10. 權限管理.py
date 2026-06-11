# -*- coding: utf-8 -*-
from __future__ import annotations

from io import StringIO, BytesIO
from datetime import datetime
import hashlib
import pandas as pd
import streamlit as st

from services.theme_service import apply_theme, render_header
from services.security_service import require_module_access, check_permission
from services.permission_service import (
    ACTIONS,
    MODULES,
    ROLE_DESCRIPTIONS,
    delete_users,
    get_account_permissions,
    get_security_settings,
    get_users,
    init_permission_tables,
    save_account_permissions,
    save_security_settings,
    save_users,
    restore_default_accounts_once_v57,
)
try:
    from services.permission_service import save_account_master as _v94_save_account_master
except Exception:
    _v94_save_account_master = None

apply_theme()
require_module_access("10_permissions", "can_manage")
render_header("10 | 權限管理", "帳號密碼總表、帳號匯入、帳號貼上、帳號級模組權限 / Account & Permission Management")

try:
    from services.performance_profiler_service import start_page_event as _spt_v40_start_page_event, finish_page_event as _spt_v40_finish_page_event
    _SPT_V40_PAGE_TOKEN = _spt_v40_start_page_event("10", "權限管理")
except Exception:
    _SPT_V40_PAGE_TOKEN = None

# V67：權限表初始化/預設帳號補齊只需每個 Streamlit session 執行一次。
# 避免 10 頁每次 widget rerun 都觸發 Neon schema/default account 檢查。
if not st.session_state.get("_spt_v67_permission_bootstrap_ready", False):
    init_permission_tables()
    try:
        # V57：補回原始六個預設帳號（只新增缺少帳號，不覆蓋已存在帳號資料）
        restore_default_accounts_once_v57()
    except Exception:
        pass
    st.session_state["_spt_v67_permission_bootstrap_ready"] = True

st.caption("V300.48 loaded｜權限管理頁：Account Editor 儲存只送出有異動列，避免套用帳號密碼總表時長時間運轉。")

ROLE_OPTIONS = ["admin", "manager", "leader", "operator", "viewer", "auditor"]
ACTION_COLS = [a[0] for a in ACTIONS]


# ===== V104 PERMISSION PAGE LIGHTWEIGHT CACHE =====
# 10 權限管理是設定頁，帳號、權限矩陣、安全設定都需要永久保存，
# 但不能在每個 widget rerun 都重複讀 Neon。這裡只做前台 session 快取：
# - 第一次進入區塊時讀一次
# - 按重新載入 / 儲存後才刷新
# - 正式資料仍以 permission_service / Neon 為權威
def _v104_cache_get(name: str):
    obj = st.session_state.get(f"v104_perm_cache_{name}")
    if not isinstance(obj, dict):
        return None
    try:
        age = (datetime.now() - obj.get("ts")).total_seconds()
    except Exception:
        return None
    if age > 300:
        return None
    return obj.get("data")


def _v104_cache_set(name: str, data):
    st.session_state[f"v104_perm_cache_{name}"] = {"ts": datetime.now(), "data": data}


def _v30022_clear_permission_drafts() -> None:
    """Clear editable permission matrix drafts when authority data changes.

    Permission edits are kept as front-end drafts until the explicit Apply button.
    When accounts/permissions are saved or reloaded, stale drafts must be removed
    so the next edit starts from the latest Neon authority data.
    """
    for key in list(st.session_state.keys()):
        if str(key).startswith("v30022_permission_draft_"):
            st.session_state.pop(key, None)


def _v104_cache_clear(*names: str) -> None:
    if not names:
        names = ("users", "permissions", "security")
    for name in names:
        st.session_state.pop(f"v104_perm_cache_{name}", None)
        if name in {"users", "permissions", "security"}:
            # V300.35：權限資料已變更時，手動產生的 Excel 匯出也必須失效，
            # 避免下載到儲存前的舊權限檔；只清 session bytes，不碰正式資料。
            st.session_state.pop("v95_permission_management_excel_bytes", None)
            st.session_state.pop("v95_permission_management_excel_ts", None)
        if name in {"users", "permissions"}:
            _v30022_clear_permission_drafts()


def _v104_get_users_cached(force: bool = False) -> list[dict]:
    if not force:
        cached = _v104_cache_get("users")
        if cached is not None:
            return [dict(r) for r in cached]
    rows = get_users()
    _v104_cache_set("users", [dict(r) for r in rows])
    return rows


def _v104_get_permissions_cached(force: bool = False) -> list[dict]:
    if not force:
        cached = _v104_cache_get("permissions")
        if cached is not None:
            return [dict(r) for r in cached]
    rows = get_account_permissions()
    _v104_cache_set("permissions", [dict(r) for r in rows])
    return rows


def _v104_get_security_cached(force: bool = False) -> dict:
    if not force:
        cached = _v104_cache_get("security")
        if isinstance(cached, dict):
            return dict(cached)
    data = get_security_settings() or {}
    _v104_cache_set("security", dict(data))
    return data


# ===== V300.35 ACCOUNT IMPORT PARSE CACHE =====
def _v30035_digest_bytes(data: bytes) -> str:
    try:
        return hashlib.sha1(data or b"").hexdigest()
    except Exception:
        return str(len(data or b""))


def _v30035_parse_account_excel_cached(uploaded_file, has_header: bool) -> pd.DataFrame:
    """Parse uploaded account Excel once per identical file/header setting.

    Streamlit tabs execute all tab bodies on rerun.  Without this cache, a loaded
    Excel file is parsed again whenever the Account Editor reruns, even when the
    user is only editing cells.  The cache is session-local and keyed by file
    digest + header flag, so it never changes formal Neon authority data.
    """
    data = uploaded_file.getvalue() if uploaded_file is not None else b""
    sig = ("excel", _v30035_digest_bytes(data), bool(has_header))
    cached = st.session_state.get("v30035_account_excel_parse_cache")
    if isinstance(cached, dict) and cached.get("sig") == sig and isinstance(cached.get("df"), pd.DataFrame):
        return pd.DataFrame(cached["df"]).copy()
    raw_excel = pd.read_excel(BytesIO(data), header=None, dtype=str).fillna("")
    import_df = _normalize_account_import_df(raw_excel, has_header=has_header)
    st.session_state["v30035_account_excel_parse_cache"] = {"sig": sig, "df": import_df.copy()}
    return import_df


def _v30035_parse_account_paste_cached(paste_text: str, has_header: bool) -> pd.DataFrame:
    data = str(paste_text or "").encode("utf-8", errors="ignore")
    sig = ("paste", _v30035_digest_bytes(data), bool(has_header))
    cached = st.session_state.get("v30035_account_paste_parse_cache")
    if isinstance(cached, dict) and cached.get("sig") == sig and isinstance(cached.get("df"), pd.DataFrame):
        return pd.DataFrame(cached["df"]).copy()
    raw_paste = pd.read_csv(StringIO(str(paste_text or "")), sep="\t", header=None, dtype=str, engine="python").fillna("")
    if raw_paste.shape[1] <= 1:
        raw_paste = pd.read_csv(StringIO(str(paste_text or "")), sep=r"\s{2,}|,", header=None, dtype=str, engine="python").fillna("")
    import_df = _normalize_account_import_df(raw_paste, has_header=has_header)
    st.session_state["v30035_account_paste_parse_cache"] = {"sig": sig, "df": import_df.copy()}
    return import_df
# ===== V300.35 ACCOUNT IMPORT PARSE CACHE END =====


# ===== V300.46 ACCOUNT PASTE/IMPORT ACTION FASTPATH =====
def _v30046_import_df_signature(df: pd.DataFrame, action: str = "") -> str:
    """Stable signature for one pasted/imported account batch.

    Used only for UI action guards and widget cleanup.  It does not replace Neon
    transactions or authority writes.
    """
    try:
        work = pd.DataFrame(df).fillna("").astype(str)
        payload = work.to_csv(index=False).encode("utf-8", errors="ignore")
    except Exception:
        payload = str(df).encode("utf-8", errors="ignore")
    return hashlib.sha1(str(action).encode("utf-8") + b"\0" + payload).hexdigest()


def _v30046_get_editor_base_df() -> pd.DataFrame:
    """Return current Account Editor draft without accidentally reloading Neon.

    Avoid using st.session_state.get("v133_users_df", _users_for_editor()) because
    Python evaluates the default argument eagerly, causing an unnecessary get_users()
    even when the draft already exists.  That was one cause of Paste Data actions
    continuing to spin after merging rows.
    """
    existing = st.session_state.get("v133_users_df")
    if isinstance(existing, pd.DataFrame):
        return _v30047_normalize_account_editor_df(existing)
    return _users_for_editor()


def _v30046_bump_widget_nonce(kind: str) -> None:
    key = f"v30046_{kind}_nonce"
    try:
        st.session_state[key] = int(st.session_state.get(key, 0) or 0) + 1
    except Exception:
        st.session_state[key] = 1
    if kind == "paste":
        st.session_state.pop("v30035_account_paste_parse_cache", None)
    if kind == "excel":
        st.session_state.pop("v30035_account_excel_parse_cache", None)


def _v30046_import_action_recent(sig: str, window_seconds: float = 8.0) -> bool:
    """Prevent accidental duplicate direct-save from double clicks/reruns."""
    import time
    rec = st.session_state.get("v30046_last_account_import_action")
    if not isinstance(rec, dict) or rec.get("sig") != sig:
        return False
    try:
        return (time.time() - float(rec.get("ts") or 0)) <= window_seconds
    except Exception:
        return False


def _v30046_mark_import_action(sig: str, result: dict | None = None) -> None:
    import time
    st.session_state["v30046_last_account_import_action"] = {"sig": sig, "ts": time.time(), "result": result or {}}


def _v30046_set_notice(level: str, message: str) -> None:
    st.session_state["v30046_account_notice"] = {"level": str(level or "info"), "message": str(message or "")}


def _v30046_show_notice() -> None:
    notice = st.session_state.pop("v30046_account_notice", None)
    if not isinstance(notice, dict) or not notice.get("message"):
        return
    level = str(notice.get("level") or "info").lower()
    msg = str(notice.get("message") or "")
    if level == "success":
        st.success(msg)
    elif level == "warning":
        st.warning(msg)
    elif level == "error":
        st.error(msg)
    else:
        st.info(msg)
# ===== V300.46 ACCOUNT PASTE/IMPORT ACTION FASTPATH END =====
_v30046_show_notice()
# ===== V104 PERMISSION PAGE LIGHTWEIGHT CACHE END =====


# ===== V300.21 PERMISSION MATRIX COLUMN COMPATIBILITY =====
def _v30021_ensure_permission_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure permission editor has all columns expected by the page.

    Some cleaned authority files may use an older/narrower permission matrix schema.
    The page must not crash when a column is missing; it should fill safe defaults
    and let the save button write back the complete schema.
    """
    base_cols = ["username", "display_name", "role_code", "module_code", "module_name_zh", "module_name_en"]
    if df is None:
        df = pd.DataFrame()
    df = pd.DataFrame(df).copy()
    for col in base_cols:
        if col not in df.columns:
            df[col] = ""
    for col in ACTION_COLS:
        if col not in df.columns:
            df[col] = False
        else:
            df[col] = df[col].map(lambda v: _as_bool_value(v, default=False))
    ordered = base_cols + ACTION_COLS
    other_cols = [c for c in df.columns if c not in ordered]
    return df[ordered + other_cols]
# ===== V300.21 PERMISSION MATRIX COLUMN COMPATIBILITY END =====


# ===== V300.22 PERMISSION MATRIX FRONTEND DRAFT =====
def _v30022_permission_draft_key(selected_user: str, selected_module: str, selected_role: str) -> str:
    raw = f"{selected_user}|{selected_module}|{selected_role}"
    safe = "".join(ch if ch.isalnum() else "_" for ch in raw)[:160]
    return f"v30022_permission_draft_{safe}"


def _v30022_permission_get_draft(key: str, base_df: pd.DataFrame, *, force_reset: bool = False) -> pd.DataFrame:
    if force_reset or key not in st.session_state or not isinstance(st.session_state.get(key), pd.DataFrame):
        st.session_state[key] = pd.DataFrame(base_df).copy().reset_index(drop=True)
    return pd.DataFrame(st.session_state[key]).copy().reset_index(drop=True)


def _v30022_permission_set_draft(key: str, df: pd.DataFrame) -> pd.DataFrame:
    work = _v30021_ensure_permission_columns(pd.DataFrame(df).copy())
    st.session_state[key] = work.copy().reset_index(drop=True)
    return work.copy().reset_index(drop=True)
# ===== V300.22 PERMISSION MATRIX FRONTEND DRAFT END =====

ACCOUNT_HEADER_ALIASES = {
    "username": ["帳號", "登入帳號", "使用者", "使用者帳號", "username", "user", "account", "login id", "login_id"],
    "new_password": ["密碼", "新密碼", "預設密碼", "password", "new password", "new_password", "pwd"],
    "employee_id": ["工號", "員工編號", "人員編號", "employee id", "employee_id", "emp id", "emp_id"],
    "display_name": ["姓名", "顯示名稱", "員工姓名", "人員姓名", "display name", "display_name", "name"],
    "email": ["email", "e-mail", "信箱", "電子信箱", "mail"],
    "role_code": ["角色", "角色代碼", "權限角色", "role", "role_code", "role code"],
    "is_active": ["啟用", "啟用狀態", "active", "is_active", "在職", "有效"],
    "force_password_change": ["強制改密碼", "首次登入改密碼", "force change", "force_password_change", "force password change"],
    "note": ["備註", "說明", "note", "remark", "remarks", "memo"],
}

ACCOUNT_DISPLAY_COLUMNS = {
    "username": "帳號 / Username",
    "new_password": "密碼 / Password",
    "employee_id": "工號 / Employee ID",
    "display_name": "姓名 / Display Name",
    "email": "Email",
    "role_code": "角色 / Role",
    "is_active": "啟用 / Active",
    "force_password_change": "強制改密碼 / Force Change",
    "note": "備註 / Note",
}

# V300.47：Account Editor 欄位順序與標準欄名。
# Streamlit data_editor 在部分版本會把「資料欄名 + column_config label」同時顯示；
# 若欄名本身已經是中英雙語，再給同一個中英 label，就會變成
# 「帳號 / Username / 帳號 / Username」。因此 Account Editor 以資料欄名作為唯一標題，
# column_config 只負責型別/選項，不再重複指定相同 label。
ACCOUNT_EDITOR_COLUMNS = [
    "刪除 / Delete",
    "帳號 / Username",
    "密碼狀態 / Password Status",
    "強制改密碼 / Force Change",
    "新密碼 / New Password",
    "工號 / Employee ID",
    "姓名 / Display Name",
    "Email",
    "角色 / Role",
    "啟用 / Active",
    "備註 / Note",
    "最後登入 / Last Login",
    "更新時間 / Updated At",
]


# ===== V95 FAST RAW EDITOR HELPER =====
def _v95_raw_data_editor(data=None, *args, **kwargs):
    """Use Streamlit native data_editor for 10. 權限管理 heavy editors.

    The global column-settings wrapper renders an extra settings data_editor for every
    table. In this module it caused 2~3 minute reruns and duplicate-key failures.
    This helper bypasses that wrapper only for the two permission-management editors.
    """
    try:
        import services.column_settings_service as _css
        _orig = getattr(_css, "_ORIGINAL_DATA_EDITOR", None)
        if callable(_orig):
            return _orig(data, *args, **kwargs)
    except Exception:
        pass
    return st.data_editor(data, *args, **kwargs)
# ===== V95 FAST RAW EDITOR HELPER END =====

with st.expander("⧠ 權限設定使用說明 / User Guide", expanded=False):
    st.markdown("""
### 密碼欄位說明 / Password Field Design
- **密碼 / Password** 欄可直接輸入新密碼；既有帳號顯示 `********` 代表維持原密碼。
- 系統不會顯示既有密碼明碼，只顯示 `********` 或狀態提示。
- 要新增或修改密碼，可填 **密碼 / Password** 或 **新密碼 / New Password**；匯入資料中的 **密碼 / Password** 也會套用。
- 空白代表維持原密碼，不會清掉密碼。

### 匯入帳號格式 / Account Import Format
建議 Excel 或貼上資料包含標題列，系統會依標題自動對應欄位：

`帳號、密碼、工號、姓名、Email、角色、啟用、強制改密碼、備註`

角色可填：`admin / manager / leader / operator / viewer / auditor`，也可填中文：系統管理員、製造主管、現場幹部、作業人員、查詢者、稽核。
""")

def _role_text(info, key: str, fallback_key: str = "", default: str = "") -> str:
    """V300 clean compatibility: role metadata may come from old/new schema."""
    if not isinstance(info, dict):
        return default
    value = info.get(key)
    if value not in (None, ""):
        return str(value)
    if fallback_key:
        value = info.get(fallback_key)
        if value not in (None, ""):
            return str(value)
    return default

with st.expander("⌖ 角色權限說明 / Role Permission Description", expanded=False):
    st.dataframe(pd.DataFrame([
        {
            "角色代碼 / Role Code": role_code,
            "中文角色 / Chinese Role": _role_text(info, "zh", "label", role_code),
            "英文角色 / English Role": _role_text(info, "en", "role_en", role_code),
            "建議用途 / Recommendation": _role_text(info, "desc", "description", ""),
        }
        for role_code, info in ROLE_DESCRIPTIONS.items()
    ]), use_container_width=True, hide_index=True)


def _as_bool_value(value, default: bool = False) -> bool:
    """Robust checkbox value parser for Streamlit data_editor.

    Streamlit versions may return bool, numpy bool, integers, or text-like values
    after column conversions.  Deletion must not silently fail because the UI
    value came back as a string.
    """
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except Exception:
        pass
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y", "on", "checked", "☑", "✅", "是", "勾選", "刪除", "delete"}:
        return True
    if text in {"false", "0", "no", "n", "off", "unchecked", "☐", "□", "否", "不刪除", ""}:
        return False
    return default


def _to_bool_series(df: pd.DataFrame, col: str, default: bool = False) -> pd.Series:
    if col not in df.columns:
        return pd.Series([default] * len(df), index=df.index)
    return df[col].map(lambda v: _as_bool_value(v, default=default)).fillna(default).astype(bool)


def _selected_delete_usernames(df: pd.DataFrame, editor_key: str) -> list[str]:
    """Return selected usernames from data_editor + widget delta state.

    In some Streamlit builds, checkbox edits inside a form can be present in the
    widget state even when the returned dataframe is not yet fully refreshed.
    This fallback prevents the 'Delete' checkbox from looking checked but saving
    zero deletions.
    """
    selected: set[str] = set()
    if df is None or df.empty or "帳號 / Username" not in df.columns:
        return []

    mask = _to_bool_series(df, "刪除 / Delete")
    for _, row in df.loc[mask].iterrows():
        username = str(row.get("帳號 / Username", "")).strip()
        if username:
            selected.add(username)

    state = st.session_state.get(editor_key, {})
    if isinstance(state, dict):
        edited_rows = state.get("edited_rows", {}) or {}
        for row_idx, changes in edited_rows.items():
            if not isinstance(changes, dict) or "刪除 / Delete" not in changes:
                continue
            if not _as_bool_value(changes.get("刪除 / Delete"), default=False):
                continue
            try:
                idx = int(row_idx)
            except Exception:
                continue
            if 0 <= idx < len(df):
                username = str(df.iloc[idx].get("帳號 / Username", "")).strip()
                if username:
                    selected.add(username)
    return sorted(selected)



def _v30023_apply_account_editor_delta(base_df: pd.DataFrame, returned_df: pd.DataFrame, editor_key: str) -> pd.DataFrame:
    """Merge Streamlit data_editor widget deltas into the submitted dataframe.

    Some Streamlit builds can show edited cells in the browser while the dataframe
    returned from a data_editor inside a form is still the pre-edit value on the
    submit rerun.  V300.22 only compensated for the Delete checkbox.  Account
    Editor must also preserve text/selectbox/active/force-change edits before
    calling save_users(), otherwise the page appears to save, reruns, and then
    reloads the old authority values.
    """
    if isinstance(returned_df, pd.DataFrame):
        work = _v30047_normalize_account_editor_df(returned_df).reset_index(drop=True)
    elif isinstance(base_df, pd.DataFrame):
        work = _v30047_normalize_account_editor_df(base_df).reset_index(drop=True)
    else:
        work = pd.DataFrame()

    # Ensure the submitted dataframe has the same visible columns as the current draft.
    if isinstance(base_df, pd.DataFrame) and not base_df.empty:
        base_norm = _v30047_normalize_account_editor_df(base_df)
        for col in base_norm.columns:
            if col not in work.columns:
                work[col] = base_norm[col].values[: len(work)] if len(base_norm) >= len(work) else ""

    state = st.session_state.get(editor_key, {})
    if isinstance(state, dict):
        edited_rows = state.get("edited_rows", {}) or {}
        for row_idx, changes in edited_rows.items():
            if not isinstance(changes, dict):
                continue
            try:
                idx = int(row_idx)
            except Exception:
                continue
            if idx < 0:
                continue
            # data_editor num_rows is fixed here, but be defensive for future compatibility.
            while idx >= len(work):
                work = pd.concat([work, pd.DataFrame([_blank_user_row()])], ignore_index=True)
            for col, value in changes.items():
                canonical_col = _v30047_canonical_account_editor_col(col)
                if canonical_col not in work.columns:
                    work[canonical_col] = ""
                work.at[idx, canonical_col] = value

        deleted_rows = state.get("deleted_rows", []) or []
        if deleted_rows:
            try:
                drop_idx = {int(x) for x in deleted_rows}
                work = work.loc[[i for i in range(len(work)) if i not in drop_idx]].reset_index(drop=True)
            except Exception:
                pass

        added_rows = state.get("added_rows", []) or []
        for row in added_rows:
            if isinstance(row, dict):
                new_row = _blank_user_row()
                new_row.update(row)
                work = pd.concat([work, pd.DataFrame([new_row])], ignore_index=True)

    return _v30047_normalize_account_editor_df(work)

def _blank_user_row() -> dict:
    return {
        "刪除 / Delete": False,
        "帳號 / Username": "",
        "密碼狀態 / Password Status": "新帳號請輸入新密碼",
        "新密碼 / New Password": "",
        "工號 / Employee ID": "",
        "姓名 / Display Name": "",
        "Email": "",
        "角色 / Role": "operator",
        "啟用 / Active": True,
        "強制改密碼 / Force Change": False,
        "備註 / Note": "",
        "最後登入 / Last Login": "",
        "更新時間 / Updated At": "",
    }


def _v30047_canonical_account_editor_col(col) -> str:
    """Return the official Account Editor column name.

    This fixes stale session drafts created by previous versions where a bilingual
    column title could be stored twice, for example:
    「帳號 / Username / 帳號 / Username」.  It also protects save_users() from
    reading the wrong column after a rerun.
    """
    s0 = str(col or "").strip()
    if not s0:
        return s0
    alias_map = {
        "密碼 / Password": "新密碼 / New Password",
        "姓名 / Name": "姓名 / Display Name",
        "登入帳號 / Username": "帳號 / Username",
        "使用者 / Username": "帳號 / Username",
        "Force Change": "強制改密碼 / Force Change",
        "force_password_change": "強制改密碼 / Force Change",
        "password_status": "密碼狀態 / Password Status",
        "new_password": "新密碼 / New Password",
        "employee_id": "工號 / Employee ID",
        "display_name": "姓名 / Display Name",
        "role_code": "角色 / Role",
        "is_active": "啟用 / Active",
    }
    if s0 in alias_map:
        return alias_map[s0]
    for canonical in sorted(ACCOUNT_EDITOR_COLUMNS, key=len, reverse=True):
        if s0 == canonical:
            return canonical
        # Streamlit / stored column settings duplication pattern:
        #   canonical + " / " + canonical
        # or canonical + " / " + another label.  The data column itself is the
        # canonical source of truth, so map back to the first canonical segment.
        if s0.startswith(canonical + " / "):
            return canonical
        if s0.count(canonical) >= 2:
            return canonical
    return s0


def _v30047_normalize_account_editor_df(df: pd.DataFrame | None) -> pd.DataFrame:
    """Normalize Account Editor dataframe columns and keep Force Change present.

    The page must accept old session_state data, imported rows, and Streamlit
    data_editor returned rows.  This helper only changes column names/order and
    safe boolean types; it does not write Neon or change any saved authority data.
    """
    src = pd.DataFrame(df).copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    merged: dict[str, pd.Series] = {}
    for idx, col in enumerate(list(src.columns)):
        canonical = _v30047_canonical_account_editor_col(col)
        series = src.iloc[:, idx] if len(src.columns) else pd.Series(dtype=object)
        if canonical in merged:
            prev = merged[canonical]
            try:
                prev_blank = prev.isna() | prev.astype(str).str.strip().isin(["", "nan", "None", "NONE"])
                merged[canonical] = prev.where(~prev_blank, series)
            except Exception:
                merged[canonical] = prev
        else:
            merged[canonical] = series.copy()
    out = pd.DataFrame(merged) if merged else pd.DataFrame(index=src.index)
    defaults = _blank_user_row()
    for col, default in defaults.items():
        if col not in out.columns:
            out[col] = default
    for col, default in [
        ("刪除 / Delete", False),
        ("啟用 / Active", True),
        ("強制改密碼 / Force Change", False),
    ]:
        out[col] = _to_bool_series(out, col, default=bool(default)).fillna(bool(default)).astype(bool)
    ordered = [c for c in ACCOUNT_EDITOR_COLUMNS if c in out.columns]
    other_cols = [c for c in out.columns if c not in ordered]
    return out[ordered + other_cols].reset_index(drop=True)


def _users_for_editor() -> pd.DataFrame:
    raw = pd.DataFrame(_v104_get_users_cached())
    if raw.empty:
        return pd.DataFrame([_blank_user_row()]).iloc[0:0]
    out = pd.DataFrame()
    out["刪除 / Delete"] = False
    out["帳號 / Username"] = raw.get("username", "")
    out["密碼狀態 / Password Status"] = raw.get("password_display", "********")
    out["新密碼 / New Password"] = raw.get("new_password", "")
    out["工號 / Employee ID"] = raw.get("employee_id", "")
    out["姓名 / Display Name"] = raw.get("display_name", "")
    out["Email"] = raw.get("email", "")
    out["角色 / Role"] = raw.get("role_code", "operator")
    out["啟用 / Active"] = raw.get("is_active", 1).fillna(1).astype(bool) if "is_active" in raw else True
    out["強制改密碼 / Force Change"] = raw.get("force_password_change", 0).fillna(0).astype(bool) if "force_password_change" in raw else False
    out["備註 / Note"] = raw.get("note", "")
    out["最後登入 / Last Login"] = raw.get("last_login_at", "")
    out["更新時間 / Updated At"] = raw.get("updated_at", "")
    return _v30047_normalize_account_editor_df(out)


def _password_from_editor_row(r: pd.Series) -> str:
    """Return a new password only when the user really typed one.

    V1.76：使用者常誤以為「密碼狀態」不能輸入是錯誤。
    現在允許直接在該欄輸入新密碼；若仍為 ******** 或提示文字，代表維持原密碼。
    """
    explicit = str(r.get("新密碼 / New Password", "") or "").strip()
    if explicit:
        return explicit
    status_value = str(r.get("密碼狀態 / Password Status", "") or "").strip()
    blocked_values = {
        "", "none", "nan", "********", "*******", "******",
        "新帳號請輸入新密碼", "匯入後將更新密碼", "未提供密碼，維持原密碼",
        "已設定", "維持原密碼", "password set",
    }
    if status_value.lower() in blocked_values or set(status_value) == {"*"}:
        return ""
    return status_value


def _users_to_service_rows(df: pd.DataFrame) -> list[dict]:
    df = _v30047_normalize_account_editor_df(df)
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "username": str(r.get("帳號 / Username", "")).strip(),
            "new_password": _password_from_editor_row(r),
            "employee_id": str(r.get("工號 / Employee ID", "")).strip(),
            "display_name": str(r.get("姓名 / Display Name", "")).strip(),
            "email": str(r.get("Email", "")).strip(),
            "role_code": str(r.get("角色 / Role", "operator")).strip() or "operator",
            "is_active": bool(r.get("啟用 / Active", True)),
            "force_password_change": bool(r.get("強制改密碼 / Force Change", False)),
            "note": str(r.get("備註 / Note", "")).strip(),
        })
    return rows




# ===== V300.48 ACCOUNT EDITOR SAVE DELTA FASTPATH =====
def _v30048_set_account_authority_baseline(df: pd.DataFrame | None) -> None:
    """Store the last Neon authority snapshot used by Account Editor.

    Account Editor can contain many rows.  Before V300.48, pressing Apply sent
    every row to save_users(), so the service still had to compare every account
    and fetch all submitted usernames from Neon.  This session-local baseline lets
    the page send only rows that actually changed.  It is not an authority source;
    Neon remains the authority and the baseline is refreshed on reload/save.
    """
    try:
        st.session_state["v30048_account_authority_df"] = _v30047_normalize_account_editor_df(df).copy()
    except Exception:
        st.session_state.pop("v30048_account_authority_df", None)


def _v30048_get_account_authority_baseline(default_df: pd.DataFrame | None = None) -> pd.DataFrame:
    base = st.session_state.get("v30048_account_authority_df")
    if isinstance(base, pd.DataFrame):
        return _v30047_normalize_account_editor_df(base)
    if isinstance(default_df, pd.DataFrame):
        _v30048_set_account_authority_baseline(default_df)
        return _v30047_normalize_account_editor_df(default_df)
    return pd.DataFrame()


def _v30048_account_compare_value(v) -> str:
    if pd.isna(v):
        return ""
    if isinstance(v, bool):
        return "1" if v else "0"
    return str(v or "").strip()


def _v30048_account_changed_rows(current_df: pd.DataFrame, baseline_df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Return only Account Editor rows that changed versus the last authority snapshot."""
    current = _v30047_normalize_account_editor_df(current_df)
    baseline = _v30047_normalize_account_editor_df(baseline_df)
    base_rows: dict[str, dict] = {}
    if isinstance(baseline, pd.DataFrame) and not baseline.empty:
        for _, br in baseline.iterrows():
            uname = str(br.get("帳號 / Username", "") or "").strip().lower()
            if uname:
                base_rows[uname] = dict(br)
    compare_cols = [
        "工號 / Employee ID",
        "姓名 / Display Name",
        "Email",
        "角色 / Role",
        "啟用 / Active",
        "強制改密碼 / Force Change",
        "備註 / Note",
    ]
    changed_indexes: list[int] = []
    new_count = 0
    edited_count = 0
    password_count = 0
    for idx, row in current.iterrows():
        username = str(row.get("帳號 / Username", "") or "").strip()
        if not username:
            continue
        old = base_rows.get(username.lower())
        row_changed = old is None
        if old is None:
            new_count += 1
        password = _password_from_editor_row(row)
        if password:
            row_changed = True
            password_count += 1
        if old is not None:
            for col in compare_cols:
                if col in {"啟用 / Active", "強制改密碼 / Force Change"}:
                    old_val = bool(_to_bool_series(pd.DataFrame([old]), col, default=False).iloc[0])
                    new_val = bool(_to_bool_series(pd.DataFrame([row]), col, default=False).iloc[0])
                    if old_val != new_val:
                        row_changed = True
                        break
                else:
                    if _v30048_account_compare_value(old.get(col, "")) != _v30048_account_compare_value(row.get(col, "")):
                        row_changed = True
                        break
        if row_changed:
            if old is not None:
                edited_count += 1
            changed_indexes.append(idx)
    changed = current.loc[changed_indexes].copy() if changed_indexes else current.iloc[0:0].copy()
    return changed.reset_index(drop=True), {
        "submitted": int(len(current)),
        "changed": int(len(changed)),
        "new": int(new_count),
        "edited": int(edited_count),
        "password": int(password_count),
    }


def _v30048_prepare_account_post_save_df(df: pd.DataFrame, deleted_usernames: list[str] | None = None) -> pd.DataFrame:
    """Keep a lightweight local post-save preview instead of immediately reloading Neon."""
    out = _v30047_normalize_account_editor_df(df)
    deleted = {str(u or "").strip().lower() for u in (deleted_usernames or []) if str(u or "").strip()}
    if deleted and "帳號 / Username" in out.columns:
        out = out.loc[~out["帳號 / Username"].astype(str).str.strip().str.lower().isin(deleted)].copy()
    if "新密碼 / New Password" in out.columns:
        out["新密碼 / New Password"] = ""
    if "密碼狀態 / Password Status" in out.columns:
        out["密碼狀態 / Password Status"] = out["密碼狀態 / Password Status"].map(
            lambda v: "********" if str(v or "").strip() not in {"", "新帳號請輸入新密碼"} else str(v or "")
        )
        out.loc[out["帳號 / Username"].astype(str).str.strip() != "", "密碼狀態 / Password Status"] = "********"
    return _v30047_normalize_account_editor_df(out)
# ===== V300.48 ACCOUNT EDITOR SAVE DELTA FASTPATH END =====


def _norm_header(v) -> str:
    return str(v).strip().lower().replace("　", " ").replace("/", " ").replace("\\", " ").replace("-", " ").replace("_", " ")


def _parse_bool_value(v, default: bool = True) -> bool:
    if pd.isna(v):
        return default
    t = str(v).strip().lower()
    if t in ("", "nan", "none"):
        return default
    if t in ("1", "true", "yes", "y", "啟用", "是", "在職", "有效", "active", "勾選", "已啟用"):
        return True
    if t in ("0", "false", "no", "n", "停用", "否", "離職", "無效", "inactive", "取消", "未啟用"):
        return False
    return default


def _normalize_role(v) -> str:
    t = str(v).strip().lower()
    mapping = {
        "系統管理員": "admin", "管理員": "admin", "administrator": "admin", "admin": "admin",
        "製造主管": "manager", "主管": "manager", "manager": "manager",
        "現場幹部": "leader", "幹部": "leader", "leader": "leader", "line leader": "leader",
        "作業人員": "operator", "人員": "operator", "operator": "operator",
        "查詢者": "viewer", "viewer": "viewer", "檢視者": "viewer",
        "稽核": "auditor", "auditor": "auditor", "稽核人員": "auditor",
    }
    return mapping.get(t, t if t in ROLE_OPTIONS else "operator")


def _map_account_headers(headers: list) -> dict:
    mapped = {}
    normalized = [_norm_header(h) for h in headers]
    for internal, aliases in ACCOUNT_HEADER_ALIASES.items():
        alias_norms = {_norm_header(a) for a in aliases}
        for idx, h in enumerate(normalized):
            if h in alias_norms and internal not in mapped:
                mapped[internal] = headers[idx]
                break
    return mapped


def _normalize_account_import_df(df_raw: pd.DataFrame, has_header: bool = True) -> pd.DataFrame:
    if df_raw is None or df_raw.empty:
        return pd.DataFrame(columns=list(ACCOUNT_DISPLAY_COLUMNS.values()))
    df = df_raw.copy().dropna(how="all")
    if df.empty:
        return pd.DataFrame(columns=list(ACCOUNT_DISPLAY_COLUMNS.values()))
    if has_header:
        headers = [str(x).strip() for x in df.iloc[0].tolist()]
        data = df.iloc[1:].copy()
        data.columns = headers
        mapped = _map_account_headers(headers)
        out = pd.DataFrame()
        for internal in ACCOUNT_DISPLAY_COLUMNS:
            src = mapped.get(internal)
            out[ACCOUNT_DISPLAY_COLUMNS[internal]] = data[src] if src in data.columns else ""
    else:
        data = df.copy().reset_index(drop=True)
        out = pd.DataFrame()
        order = list(ACCOUNT_DISPLAY_COLUMNS.keys())
        for idx, internal in enumerate(order):
            out[ACCOUNT_DISPLAY_COLUMNS[internal]] = data.iloc[:, idx] if idx < data.shape[1] else ""
    out = out.dropna(how="all")
    if out.empty:
        return out
    out["帳號 / Username"] = out["帳號 / Username"].fillna("").astype(str).str.strip()
    out["密碼 / Password"] = out["密碼 / Password"].fillna("").astype(str).str.strip()
    out["工號 / Employee ID"] = out["工號 / Employee ID"].fillna("").astype(str).str.strip()
    out["姓名 / Display Name"] = out["姓名 / Display Name"].fillna("").astype(str).str.strip()
    out["Email"] = out["Email"].fillna("").astype(str).str.strip()
    out["角色 / Role"] = out["角色 / Role"].fillna("operator").apply(_normalize_role)
    out["啟用 / Active"] = out["啟用 / Active"].apply(lambda x: _parse_bool_value(x, True))
    out["強制改密碼 / Force Change"] = out["強制改密碼 / Force Change"].apply(lambda x: _parse_bool_value(x, False))
    out["備註 / Note"] = out["備註 / Note"].fillna("").astype(str).str.strip()
    out = out[out["帳號 / Username"] != ""].copy()
    return out.reset_index(drop=True)


def _account_import_to_editor_rows(import_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, r in import_df.iterrows():
        username = str(r.get("帳號 / Username", "")).strip()
        password = str(r.get("密碼 / Password", "")).strip()
        rows.append({
            "刪除 / Delete": False,
            "帳號 / Username": username,
            "密碼狀態 / Password Status": "匯入後將更新密碼" if password else "未提供密碼，維持原密碼",
            "新密碼 / New Password": password,
            "工號 / Employee ID": str(r.get("工號 / Employee ID", "")).strip(),
            "姓名 / Display Name": str(r.get("姓名 / Display Name", "")).strip() or username,
            "Email": str(r.get("Email", "")).strip(),
            "角色 / Role": _normalize_role(r.get("角色 / Role", "operator")),
            "啟用 / Active": bool(r.get("啟用 / Active", True)),
            "強制改密碼 / Force Change": bool(r.get("強制改密碼 / Force Change", False)),
            "備註 / Note": str(r.get("備註 / Note", "")).strip(),
            "最後登入 / Last Login": "",
            "更新時間 / Updated At": "",
        })
    return pd.DataFrame(rows)


def _merge_users_editor(base_df: pd.DataFrame, new_rows: pd.DataFrame) -> pd.DataFrame:
    base_df = _v30047_normalize_account_editor_df(base_df)
    new_rows = _v30047_normalize_account_editor_df(new_rows)
    if base_df is None or base_df.empty:
        return new_rows.copy()
    if new_rows is None or new_rows.empty:
        return base_df.copy()
    base = base_df.copy()
    for _, r in new_rows.iterrows():
        username = str(r.get("帳號 / Username", "")).strip()
        if not username:
            continue
        mask = base["帳號 / Username"].fillna("").astype(str).str.strip().str.lower() == username.lower()
        if mask.any():
            idx = base.index[mask][0]
            for c in new_rows.columns:
                val = r.get(c, "")
                if c == "新密碼 / New Password" and not str(val).strip():
                    continue
                base.at[idx, c] = val
        else:
            base = pd.concat([base, pd.DataFrame([r])], ignore_index=True)
    return base.reset_index(drop=True)


def _save_imported_accounts(import_df: pd.DataFrame) -> dict:
    editor_rows = _account_import_to_editor_rows(import_df)
    rows = _users_to_service_rows(editor_rows)
    return save_users(rows)


def _permission_summary(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    calc = df.copy()
    for c in ACTION_COLS:
        calc[c] = calc[c].fillna(False).astype(bool).astype(int) if c in calc.columns else 0
    return calc.groupby("username", as_index=False).agg(
        可進入模組數_View=("can_view", "sum"),
        可新增_Create=("can_create", "sum"),
        可編輯_Edit=("can_edit", "sum"),
        可刪除_Delete=("can_delete", "sum"),
        可匯入_Import=("can_import", "sum"),
        可匯出_Export=("can_export", "sum"),
        可備份_Backup=("can_backup", "sum"),
        可還原_Restore=("can_restore", "sum"),
        可管理_Manage=("can_manage", "sum"),
    )


def _v93_safe_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value)


def _v93_bool_text(value) -> str:
    return "是 / Yes" if _as_bool_value(value, default=False) else "否 / No"


def _v93_account_export_df() -> pd.DataFrame:
    raw = pd.DataFrame(_v104_get_users_cached())
    if raw.empty:
        return pd.DataFrame(columns=[
            "帳號 / Username", "密碼狀態 / Password Status", "新密碼 / New Password",
            "工號 / Employee ID", "姓名 / Display Name", "Email", "角色 / Role",
            "啟用 / Active", "強制改密碼 / Force Change", "備註 / Note",
            "最後登入 / Last Login", "更新時間 / Updated At",
        ])
    out = pd.DataFrame()
    out["帳號 / Username"] = raw.get("username", "").map(_v93_safe_text)
    # 系統不輸出既有密碼明碼，也不輸出 password_hash；只輸出畫面使用的遮罩狀態。
    out["密碼狀態 / Password Status"] = raw.get("password_display", "********").map(_v93_safe_text) if "password_display" in raw else "********"
    out["新密碼 / New Password"] = ""
    out["工號 / Employee ID"] = raw.get("employee_id", "").map(_v93_safe_text) if "employee_id" in raw else ""
    out["姓名 / Display Name"] = raw.get("display_name", "").map(_v93_safe_text) if "display_name" in raw else ""
    out["Email"] = raw.get("email", "").map(_v93_safe_text) if "email" in raw else ""
    out["角色 / Role"] = raw.get("role_code", "").map(_v93_safe_text) if "role_code" in raw else ""
    out["啟用 / Active"] = raw.get("is_active", 1).map(_v93_bool_text) if "is_active" in raw else "是 / Yes"
    out["強制改密碼 / Force Change"] = raw.get("force_password_change", 0).map(_v93_bool_text) if "force_password_change" in raw else "否 / No"
    out["備註 / Note"] = raw.get("note", "").map(_v93_safe_text) if "note" in raw else ""
    out["最後登入 / Last Login"] = raw.get("last_login_at", "").map(_v93_safe_text) if "last_login_at" in raw else ""
    out["更新時間 / Updated At"] = raw.get("updated_at", "").map(_v93_safe_text) if "updated_at" in raw else ""
    return out


def _v93_permission_export_df() -> pd.DataFrame:
    raw = pd.DataFrame(_v104_get_permissions_cached())
    base_cols = ["username", "display_name", "role_code", "module_code", "module_name_zh", "module_name_en"]
    if raw.empty:
        raw = pd.DataFrame(columns=base_cols + ACTION_COLS)
    out = pd.DataFrame()
    col_map = {
        "username": "帳號 / Username",
        "display_name": "姓名 / Name",
        "role_code": "角色 / Role",
        "module_code": "模組代碼 / Module Code",
        "module_name_zh": "模組中文 / Module Chinese",
        "module_name_en": "模組英文 / Module English",
    }
    for src, dst in col_map.items():
        out[dst] = raw[src].map(_v93_safe_text) if src in raw.columns else ""
    action_map = {key: f"{zh} / {en}" for key, zh, en in ACTIONS}
    for key in ACTION_COLS:
        dst = action_map.get(key, key)
        out[dst] = raw[key].map(_v93_bool_text) if key in raw.columns else "否 / No"
    return out


def _v93_security_export_df() -> pd.DataFrame:
    settings = _v104_get_security_cached() or {}
    labels = {
        "idle_timeout_minutes": "閒置自動登出分鐘數 / Idle Auto Logout Minutes",
        "ask_continue_after_record": "工時完成後詢問是否繼續記錄 / Ask Continue After Time Record",
    }
    rows = []
    for key in sorted(settings.keys()):
        value = settings.get(key, "")
        display_value = _v93_safe_text(value)
        if key == "ask_continue_after_record":
            display_value = "啟用 / Enabled" if str(value) != "0" else "停用 / Disabled"
        rows.append({
            "設定鍵 / Setting Key": key,
            "設定名稱 / Setting Name": labels.get(key, key),
            "設定值 / Setting Value": display_value,
        })
    if not rows:
        rows.append({"設定鍵 / Setting Key": "", "設定名稱 / Setting Name": "", "設定值 / Setting Value": ""})
    return pd.DataFrame(rows)


def _v93_write_dataframe_sheet(wb, sheet_name: str, df: pd.DataFrame, title: str):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=2, column=1, value=f"匯出時間 / Export Time：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=3, column=1, value="說明 / Note：既有密碼不輸出明碼；密碼欄僅提供狀態或匯入更新用空欄。")
    header_row = 5
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="17365D")
    ws.cell(row=1, column=1).fill = title_fill
    ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=14)
    for col_idx, col_name in enumerate(df.columns.tolist(), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col_name in enumerate(df.columns.tolist(), start=1):
        values = [str(col_name)] + [str(v) for v in df.iloc[:, col_idx - 1].head(200).tolist()] if not df.empty else [str(col_name)]
        width = min(max(max(len(v) for v in values) + 2, 12), 42)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    return ws


def _build_permission_excel_export_v93() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    _v93_write_dataframe_sheet(
        wb,
        "帳號密碼總表",
        _v93_account_export_df(),
        "帳號密碼總表 / Account Password Master",
    )
    _v93_write_dataframe_sheet(
        wb,
        "帳號模組權限",
        _v93_permission_export_df(),
        "帳號模組權限 / Account Module Permissions",
    )
    _v93_write_dataframe_sheet(
        wb,
        "安全設定",
        _v93_security_export_df(),
        "安全設定 / Security",
    )
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


with st.expander("⟰ 權限管理 Excel 下載 / Permission Management Excel Export", expanded=False):
    st.caption("V95：Excel 檔改成按下產生後才建立，不再每次進入 10 頁或勾選 checkbox 都重新產生，避免拖慢頁面。")
    cex1, cex2 = st.columns([1, 1])
    with cex1:
        if st.button("▣ 產生權限管理 Excel / Build Permission Workbook", use_container_width=True, key="v95_build_permission_management_excel"):
            try:
                st.session_state["v95_permission_management_excel_bytes"] = _build_permission_excel_export_v93()
                st.session_state["v95_permission_management_excel_ts"] = datetime.now().strftime('%Y%m%d_%H%M%S')
                st.success("Excel 下載檔已產生 / Workbook generated")
            except Exception as ex:
                st.session_state.pop("v95_permission_management_excel_bytes", None)
                st.warning(f"Excel 下載檔建立失敗 / Export failed：{ex}")
    with cex2:
        excel_bytes = st.session_state.get("v95_permission_management_excel_bytes")
        if excel_bytes:
            st.download_button(
                "⬇ 下載權限管理 Excel / Download Permission Workbook",
                data=excel_bytes,
                file_name=f"SPT_permission_management_{st.session_state.get('v95_permission_management_excel_ts', datetime.now().strftime('%Y%m%d_%H%M%S'))}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="v95_permission_management_excel_download",
            )
        else:
            st.info("請先按左側產生 Excel。")


# V103: Streamlit tabs execute every tab body on each rerun.
# For 權限管理 this meant account list, permission matrix, and security settings
# were all read from Neon even when the user only opened one section.
# Use a lightweight section selector so only the selected section is rendered/read.
_permission_sections = [
    "帳號密碼總表 / Account Password Master",
    "帳號模組權限 / Account Module Permissions",
    "安全設定 / Security",
]
_selected_permission_section = st.radio(
    "管理區塊 / Management Section",
    _permission_sections,
    horizontal=True,
    key="v103_permission_section_selector",
)

if _selected_permission_section == "帳號密碼總表 / Account Password Master":
    st.subheader("帳號密碼總表 / Account & Password Master")
    st.info("V1.76：密碼欄可直接輸入新密碼。若顯示 ******** 代表維持原密碼；系統不會顯示既有密碼明碼。也可使用『新密碼 / New Password』欄。")
    account_tab_edit, account_tab_excel, account_tab_paste = st.tabs([
        "帳號清單編輯 / Account Editor", "Excel 匯入 / Excel Import", "貼上資料 / Paste Data"
    ])

    # V1.69: Edit mode is shown BEFORE tab content so it is always visible.
    if "v166_account_edit_enabled" not in st.session_state:
        st.session_state["v166_account_edit_enabled"] = False
    _edit_on = bool(st.session_state.get("v166_account_edit_enabled", False))
    c_edit1, c_edit2, c_edit3 = st.columns([1.2, 1.2, 3])
    with c_edit1:
        if st.button("◇ 啟動編輯 / Enable Edit", use_container_width=True, disabled=_edit_on, key="v169_enable_account_edit_top"):
            st.session_state["v166_account_edit_enabled"] = True
            st.rerun()
    with c_edit2:
        if st.button("◌ 停止編輯 / Lock Edit", use_container_width=True, disabled=not _edit_on, key="v169_disable_account_edit_top"):
            # V101：停止編輯只切換唯讀保護，不重新載入權威檔。
            # 舊版在這裡呼叫 _users_for_editor()；若 GitHub/權威檔短暫讀到空表，
            # 會把目前畫面草稿覆蓋成空白，造成「啟動編輯 → 停止編輯 → 全部資料消失」。
            # 正式重新載入請使用「重新載入 / Reload」按鈕；未儲存的草稿仍留在 session_state。
            st.session_state["v166_account_edit_enabled"] = False
            st.rerun()
    with c_edit3:
        if _edit_on:
            st.success("目前：已啟動編輯。修改後請按儲存才會正式寫入。")
        else:
            st.info("目前：唯讀保護。請先啟動編輯，再新增、修改、刪除、匯入或貼上帳號。")

    if "v133_users_df" not in st.session_state:
        _v30048_initial_users_df = _users_for_editor()
        st.session_state["v133_users_df"] = _v30048_initial_users_df.copy()
        _v30048_set_account_authority_baseline(_v30048_initial_users_df)
    elif "v30048_account_authority_df" not in st.session_state and isinstance(st.session_state.get("v133_users_df"), pd.DataFrame):
        _v30048_set_account_authority_baseline(st.session_state.get("v133_users_df"))

    with account_tab_edit:
        st.markdown("### 新增帳號專用表單 / Stable Add User Form")
        st.caption("建議新增帳號先使用此表單；送出後會直接寫入資料庫，不會因表格 rerun 造成資料消失。")
        with st.form("v175_create_account_form", clear_on_submit=True):
            f1, f2, f3, f4 = st.columns([1.2, 1.2, 1.2, 1.2])
            with f1:
                new_username = st.text_input("帳號 / Username", key="v175_new_username", disabled=not _edit_on)
                new_emp_id = st.text_input("工號 / Employee ID", key="v175_new_employee_id", disabled=not _edit_on)
            with f2:
                new_password = st.text_input("新密碼 / New Password", type="password", key="v175_new_password", disabled=not _edit_on)
                new_name = st.text_input("姓名 / Display Name", key="v175_new_display_name", disabled=not _edit_on)
            with f3:
                new_email = st.text_input("Email", key="v175_new_email", disabled=not _edit_on)
                new_role = st.selectbox("角色 / Role", ROLE_OPTIONS, index=ROLE_OPTIONS.index("operator"), key="v175_new_role", disabled=not _edit_on)
            with f4:
                new_active = st.checkbox("啟用 / Active", value=True, key="v175_new_active", disabled=not _edit_on)
                new_force_change = st.checkbox("強制改密碼 / Force Change", value=False, key="v175_new_force_change", disabled=not _edit_on)
            new_note = st.text_input("備註 / Note", key="v175_new_note", disabled=not _edit_on)
            submit_new_user = st.form_submit_button("⊕ 建立帳號 / Create User", type="primary", use_container_width=True, disabled=not _edit_on)
        if submit_new_user:
            username = str(new_username or "").strip()
            password = str(new_password or "").strip()
            display_name = str(new_name or "").strip() or username
            if not username:
                st.error("請輸入帳號 / Username is required")
            elif not password:
                st.error("新增帳號必須輸入新密碼 / New password is required for new account")
            else:
                result = save_users([{
                    "username": username,
                    "new_password": password,
                    "employee_id": str(new_emp_id or "").strip(),
                    "display_name": display_name,
                    "email": str(new_email or "").strip(),
                    "role_code": str(new_role or "operator").strip() or "operator",
                    "is_active": bool(new_active),
                    "force_password_change": bool(new_force_change),
                    "note": str(new_note or "").strip(),
                }])
                if result.get("saved", 0) > 0:
                    st.success(f"帳號已建立 / Account created：{username}")
                    _v104_cache_clear("users", "permissions")
                    _v30048_reloaded_users_df = _users_for_editor()
                    st.session_state["v133_users_df"] = _v30048_reloaded_users_df.copy()
                    _v30048_set_account_authority_baseline(_v30048_reloaded_users_df)
                    try:
                        from services.column_settings_service import clear_editor_draft
                        clear_editor_draft("account")
                    except Exception:
                        pass
                    st.rerun()
                if result.get("skipped"):
                    st.warning("；".join(result.get("skipped", [])))


        st.markdown("### 帳號清單編輯 / Editable Account Master")

        # V1.74：啟動/停止編輯只保留頁面上方唯一一組，避免帳號總表區重複顯示。
        if "v166_account_edit_enabled" not in st.session_state:
            st.session_state["v166_account_edit_enabled"] = False
        account_edit_enabled = bool(st.session_state.get("v166_account_edit_enabled", False))
        st.session_state.setdefault("v235_account_editor_rev", 0)

        def _v56_touch_account_editor() -> None:
            """V56：批次按鈕後換新 editor key，避免 st.data_editor 舊前端狀態蓋回 checkbox。"""
            try:
                for _k0 in list(st.session_state.keys()):
                    if str(_k0).startswith(("v171_account_password_editor_", "v56_account_password_editor_")):
                        st.session_state.pop(_k0, None)
            except Exception:
                pass
            try:
                from services.column_settings_service import clear_editor_draft
                clear_editor_draft("v171_account_password_editor")
                clear_editor_draft("v56_account_password_editor")
                clear_editor_draft("account")
            except Exception:
                pass
            st.session_state["v235_account_editor_rev"] = int(st.session_state.get("v235_account_editor_rev", 0)) + 1

        def _v56_prepare_account_draft() -> pd.DataFrame:
            """V56：統一 checkbox 欄位型別，確保 data_editor 以 bool checkbox 顯示。

            V97C：移除 session_state 內可能殘留的重複欄位，避免 data_editor
            顯示兩個「刪除 / Delete」欄，造成使用者誤判與勾選狀態錯亂。
            """
            df0 = st.session_state.get("v133_users_df")
            if not isinstance(df0, pd.DataFrame):
                df0 = _users_for_editor()
            df0 = _v30047_normalize_account_editor_df(df0)
            st.session_state["v133_users_df"] = df0
            return df0

        c1, c2, c3, c4, c5, c6 = st.columns(6)
        with c1:
            if st.button("⊕ 新增帳號 / Add User", use_container_width=True, disabled=not account_edit_enabled, key="v56_add_user_row_once"):
                base_df = _v56_prepare_account_draft()
                st.session_state["v133_users_df"] = pd.concat([base_df, pd.DataFrame([_blank_user_row()])], ignore_index=True)
                _v56_touch_account_editor()
                st.rerun()
        with c2:
            if st.button("☑ 刪除全選 / Select Delete", use_container_width=True, disabled=not account_edit_enabled, key="v56_account_select_delete"):
                df0 = _v56_prepare_account_draft()
                df0["刪除 / Delete"] = True
                st.session_state["v133_users_df"] = df0.copy()
                _v56_touch_account_editor()
                st.rerun()
        with c3:
            if st.button("☐ 刪除取消 / Clear Delete", use_container_width=True, disabled=not account_edit_enabled, key="v56_account_clear_delete"):
                df0 = _v56_prepare_account_draft()
                df0["刪除 / Delete"] = False
                st.session_state["v133_users_df"] = df0.copy()
                _v56_touch_account_editor()
                st.rerun()
        with c4:
            if st.button("☑ 啟用全選 / Active All", use_container_width=True, disabled=not account_edit_enabled, key="v56_account_active_all"):
                df0 = _v56_prepare_account_draft()
                df0["啟用 / Active"] = True
                st.session_state["v133_users_df"] = df0.copy()
                _v56_touch_account_editor()
                st.rerun()
        with c5:
            if st.button("☐ 啟用取消 / Inactive All", use_container_width=True, disabled=not account_edit_enabled, key="v56_account_inactive_all"):
                df0 = _v56_prepare_account_draft()
                df0["啟用 / Active"] = False
                st.session_state["v133_users_df"] = df0.copy()
                _v56_touch_account_editor()
                st.rerun()
        with c6:
            if st.button("⟳ 重新載入 / Reload", use_container_width=True, key="v56_account_reload"):
                _v104_cache_clear("users")
                _v30048_reloaded_users_df = _users_for_editor()
                st.session_state["v133_users_df"] = _v30048_reloaded_users_df.copy()
                _v30048_set_account_authority_baseline(_v30048_reloaded_users_df)
                _v56_touch_account_editor()
                st.rerun()

        # V107：只新增「強制改密碼」批次按鈕；不調整原本六顆按鈕、不改表格欄位、不改儲存流程。
        fc1, fc2, fc_spacer = st.columns([1, 1, 4])
        with fc1:
            if st.button("☑ 強制改密碼全選 / Force Change All", use_container_width=True, disabled=not account_edit_enabled, key="v107_account_force_change_all"):
                df0 = _v56_prepare_account_draft()
                df0["強制改密碼 / Force Change"] = True
                st.session_state["v133_users_df"] = df0.copy()
                _v56_touch_account_editor()
                st.rerun()
        with fc2:
            if st.button("☐ 強制改密碼取消 / Clear Force Change", use_container_width=True, disabled=not account_edit_enabled, key="v107_account_force_change_clear"):
                df0 = _v56_prepare_account_draft()
                df0["強制改密碼 / Force Change"] = False
                st.session_state["v133_users_df"] = df0.copy()
                _v56_touch_account_editor()
                st.rerun()

        st.warning("V107：帳號總表為穩定送出模式。批次按鈕只更新畫面草稿；修改完成後請按表格下方『套用並儲存』才會正式寫入權威檔。")

        account_editor_key = f"v102_account_password_editor_{st.session_state.get('v235_account_editor_rev', 0)}"
        draft_df = _v56_prepare_account_draft()

        # V104：唯讀狀態使用輕量表格預覽，不建立重型 data_editor。
        # 啟動編輯後才建立 data_editor，且放在 form 內，編輯時不寫 Neon、不重算、不重讀。
        if account_edit_enabled:
            with st.form("v102_account_editor_commit_form", clear_on_submit=False):
                edited_users = _v95_raw_data_editor(
                    draft_df,
                    key=account_editor_key,
                    use_container_width=True,
                    num_rows="fixed",
                    hide_index=True,
                    disabled=False,
                    height=360,
                    column_order=[c for c in ACCOUNT_EDITOR_COLUMNS if c in draft_df.columns] + [c for c in draft_df.columns if c not in ACCOUNT_EDITOR_COLUMNS],
                    column_config={
                        # V300.47：不要再把同一個中英標題傳給 column_config，
                        # 部分 Streamlit 版本會顯示成「欄名 / label」而造成標題重複。
                        # 這裡只指定欄位型別/選項，標題直接使用 dataframe 欄名。
                        "刪除 / Delete": st.column_config.CheckboxColumn(),
                        "帳號 / Username": st.column_config.TextColumn(required=True),
                        "密碼狀態 / Password Status": st.column_config.TextColumn(help="可直接輸入新密碼；******** 或提示文字代表維持原密碼"),
                        "強制改密碼 / Force Change": st.column_config.CheckboxColumn(help="勾選後，該帳號下次登入必須變更密碼。"),
                        "新密碼 / New Password": st.column_config.TextColumn(help="要改密碼才填寫；新增帳號必填"),
                        "工號 / Employee ID": st.column_config.TextColumn(),
                        "姓名 / Display Name": st.column_config.TextColumn(required=True),
                        "Email": st.column_config.TextColumn(),
                        "角色 / Role": st.column_config.SelectboxColumn(options=ROLE_OPTIONS, required=True),
                        "啟用 / Active": st.column_config.CheckboxColumn(),
                        "備註 / Note": st.column_config.TextColumn(),
                        "最後登入 / Last Login": st.column_config.TextColumn(disabled=True),
                        "更新時間 / Updated At": st.column_config.TextColumn(disabled=True),
                    },
                )

                submitted_accounts = st.form_submit_button(
                    "▣ 套用並儲存帳號密碼總表 / Apply and Save Account Master",
                    type="primary",
                    use_container_width=True,
                    disabled=False,
                )
        else:
            preview_cols = [c for c in ACCOUNT_EDITOR_COLUMNS if c in draft_df.columns and c not in {"新密碼 / New Password", "刪除 / Delete"}]
            preview_cols += [c for c in draft_df.columns if c not in preview_cols and c not in {"新密碼 / New Password", "刪除 / Delete"}]
            st.dataframe(draft_df[preview_cols].head(200), use_container_width=True, hide_index=True, height=320)
            st.caption("唯讀輕量預覽：啟動編輯後才建立可編輯表格；避免每次進入 10 頁都重建大型 data_editor。")
            edited_users = draft_df.copy()
            submitted_accounts = False

        # 未送出時不把 data_editor 回傳值寫回 session_state，避免 widget 前端草稿
        # 在 rerun 時被舊資料蓋掉。統計顯示 session 草稿值，正式送出才更新。
        metric_df = st.session_state.get("v133_users_df", draft_df)
        if not isinstance(metric_df, pd.DataFrame):
            metric_df = draft_df
        active_count = int(_to_bool_series(metric_df, "啟用 / Active").sum())
        delete_count = int(_to_bool_series(metric_df, "刪除 / Delete").sum())
        new_password_count = int(sum(1 for _, _row in metric_df.iterrows() if _password_from_editor_row(_row)))
        m1, m2, m3 = st.columns(3)
        m1.metric("啟用帳號 / Active", active_count)
        m2.metric("待刪除 / Pending Delete", delete_count)
        m3.metric("密碼異動 / Password Changes", new_password_count)

        if submitted_accounts:
            df = _v30047_normalize_account_editor_df(_v30023_apply_account_editor_delta(draft_df, edited_users, account_editor_key))
            try:
                df = df.loc[:, ~pd.Index(df.columns).duplicated()].copy()
            except Exception:
                pass
            for _col, _default in [
                ("刪除 / Delete", False),
                ("啟用 / Active", True),
                ("強制改密碼 / Force Change", False),
            ]:
                if _col not in df.columns:
                    df[_col] = _default
                df[_col] = _to_bool_series(df, _col).fillna(bool(_default)).astype(bool)
            # 送出瞬間同步 session 草稿，避免 checkbox delta 尚未併入 dataframe 時漏判。
            st.session_state["v133_users_df"] = df.copy()
            to_delete = _selected_delete_usernames(df, account_editor_key)
            if to_delete:
                save_df = df.loc[~df["帳號 / Username"].astype(str).str.strip().isin(to_delete)].copy()
            else:
                save_df = df.copy()
            baseline_df = _v30048_get_account_authority_baseline(draft_df)
            changed_df, delta_info = _v30048_account_changed_rows(save_df, baseline_df)
            service_rows = _users_to_service_rows(changed_df)
            master_result = {}
            if not service_rows and not to_delete:
                st.info("沒有偵測到帳號資料異動，未寫入 Neon。/ No account changes detected; Neon write skipped.")
            else:
                if callable(_v94_save_account_master):
                    master_result = _v94_save_account_master(service_rows, delete_usernames=to_delete)
                    result = {"saved": int(master_result.get("saved", 0)), "skipped": master_result.get("skipped", [])}
                    deleted = int(master_result.get("deleted", 0))
                else:
                    result = save_users(service_rows) if service_rows else {"saved": 0, "skipped": []}
                    deleted = delete_users(to_delete)
                if to_delete and deleted == 0:
                    st.warning("已偵測到刪除勾選，但未刪除任何帳號；admin 系統帳號不可刪除，其他帳號請確認帳號欄位是否有效。")
                st.success(
                    f"帳號已儲存：{result['saved']} 筆；刪除：{deleted} 筆；"
                    f"本次送出 {delta_info.get('submitted', 0)} 筆，只寫入異動 {delta_info.get('changed', 0)} 筆 / Accounts saved with delta fastpath"
                )
                _v125_rec = master_result.get("permission_reconcile") if callable(_v94_save_account_master) and isinstance(master_result, dict) else result.get("permission_reconcile")
                if isinstance(_v125_rec, dict):
                    st.info(
                        "權限矩陣已同步核對 / Permission matrix reconciled："
                        f"帳號 {_v125_rec.get('users', 0)}，權限列 {_v125_rec.get('permissions', 0)}，"
                        f"補齊 {_v125_rec.get('added', 0)}，修正 {_v125_rec.get('upgraded', 0)}。"
                    )
                if result.get("skipped"):
                    st.warning("；".join(result["skipped"]))
                _v104_cache_clear("users", "permissions")
                post_save_df = _v30048_prepare_account_post_save_df(save_df, to_delete)
                st.session_state["v133_users_df"] = post_save_df.copy()
                _v30048_set_account_authority_baseline(post_save_df)
                st.session_state["v166_account_edit_enabled"] = False
                _v56_touch_account_editor()
                st.rerun()

    with account_tab_excel:
        st.markdown("### Excel 匯入帳號密碼設定 / Import Account Password Settings")
        if not st.session_state.get("v166_account_edit_enabled", False):
            st.info("請先到『帳號清單編輯』按『啟動編輯』，再執行匯入或直接儲存。")
        st.caption("若第一列有標題，系統會依標題自動對應欄位。")
        st.code("帳號、密碼、工號、姓名、Email、角色、啟用、強制改密碼、備註", language="text")
        excel_nonce = int(st.session_state.get("v30046_excel_nonce", 0) or 0)
        uploaded = st.file_uploader("上傳帳號設定 Excel / Upload Account Excel", type=["xlsx", "xls"], key=f"v136_account_excel_upload_{excel_nonce}")
        excel_has_header = st.checkbox("Excel 第一列為標題列 / First row is header", value=True, key="v136_account_excel_header")
        if uploaded is not None:
            try:
                import_df = _v30035_parse_account_excel_cached(uploaded, excel_has_header)
                st.success(f"解析完成：{len(import_df)} 筆 / Parsed {len(import_df)} accounts")
                preview_df = import_df.head(300).copy()
                st.dataframe(preview_df, use_container_width=True, hide_index=True, height=320)
                if len(import_df) > len(preview_df):
                    st.caption(f"僅預覽前 {len(preview_df)} 筆，實際會處理全部 {len(import_df)} 筆。")
                e1, e2 = st.columns(2)
                with e1:
                    if st.button("⊕ 加入帳號總表編輯 / Add to Account Editor", use_container_width=True, key="v136_excel_add_editor", disabled=not st.session_state.get("v166_account_edit_enabled", False)):
                        new_rows = _account_import_to_editor_rows(import_df)
                        st.session_state["v133_users_df"] = _merge_users_editor(_v30046_get_editor_base_df(), new_rows)
                        _v30046_bump_widget_nonce("excel")
                        _v30046_set_notice("success", "已加入帳號總表編輯頁，請到『帳號清單編輯』確認後儲存。")
                        st.rerun()
                with e2:
                    if st.button("▣ 直接儲存 Excel 帳號 / Save Imported Accounts", type="primary", use_container_width=True, key="v136_excel_save_direct", disabled=not st.session_state.get("v166_account_edit_enabled", False)):
                        sig = _v30046_import_df_signature(import_df, "excel_save_direct")
                        if _v30046_import_action_recent(sig):
                            _v30046_set_notice("warning", "已偵測到同一批 Excel 帳號剛剛已送出，已略過重複儲存。")
                        else:
                            result = _save_imported_accounts(import_df)
                            _v30046_mark_import_action(sig, result)
                            msg = f"帳號已儲存：{result.get('saved', 0)} 筆；未異動：{result.get('skipped_unchanged', 0)} 筆 / Accounts saved"
                            if result.get("skipped"):
                                msg += "；" + "；".join(result["skipped"])
                            _v30046_set_notice("success", msg)
                        _v104_cache_clear("users", "permissions")
                        st.session_state.pop("v133_users_df", None)
                        _v30046_bump_widget_nonce("excel")
                        st.rerun()
            except Exception as ex:
                st.error(f"Excel 匯入失敗 / Import failed：{ex}")

    with account_tab_paste:
        st.markdown("### 貼上帳號密碼設定 / Paste Account Password Settings")
        if not st.session_state.get("v166_account_edit_enabled", False):
            st.info("請先到『帳號清單編輯』按『啟動編輯』，再執行貼上或直接儲存。")
        st.caption("支援從 Excel 直接複製貼上，建議包含標題列。")
        st.code("帳號\t密碼\t工號\t姓名\tEmail\t角色\t啟用\t強制改密碼\t備註", language="text")
        paste_nonce = int(st.session_state.get("v30046_paste_nonce", 0) or 0)
        paste_text = st.text_area("貼上 Excel 複製資料 / Paste copied Excel data", height=220, key=f"v136_account_paste_text_{paste_nonce}")
        paste_has_header = st.checkbox("貼上資料第一列為標題列 / First row is header", value=True, key="v136_account_paste_header")
        if paste_text.strip():
            try:
                import_df = _v30035_parse_account_paste_cached(paste_text, paste_has_header)
                st.success(f"解析完成：{len(import_df)} 筆 / Parsed {len(import_df)} accounts")
                preview_df = import_df.head(300).copy()
                st.dataframe(preview_df, use_container_width=True, hide_index=True, height=320)
                if len(import_df) > len(preview_df):
                    st.caption(f"僅預覽前 {len(preview_df)} 筆，實際會處理全部 {len(import_df)} 筆。")
                p1, p2 = st.columns(2)
                with p1:
                    if st.button("⊕ 加入帳號總表編輯 / Add to Account Editor", use_container_width=True, key="v136_paste_add_editor", disabled=not st.session_state.get("v166_account_edit_enabled", False)):
                        new_rows = _account_import_to_editor_rows(import_df)
                        st.session_state["v133_users_df"] = _merge_users_editor(_v30046_get_editor_base_df(), new_rows)
                        _v30046_bump_widget_nonce("paste")
                        _v30046_set_notice("success", "已加入帳號總表編輯頁，請到『帳號清單編輯』確認後儲存。貼上輸入框已清空，避免每次 rerun 重複解析。")
                        st.rerun()
                with p2:
                    if st.button("▣ 直接儲存貼上帳號 / Save Pasted Accounts", type="primary", use_container_width=True, key="v136_paste_save_direct", disabled=not st.session_state.get("v166_account_edit_enabled", False)):
                        sig = _v30046_import_df_signature(import_df, "paste_save_direct")
                        if _v30046_import_action_recent(sig):
                            _v30046_set_notice("warning", "已偵測到同一批貼上帳號剛剛已送出，已略過重複儲存。")
                        else:
                            result = _save_imported_accounts(import_df)
                            _v30046_mark_import_action(sig, result)
                            msg = f"帳號已儲存：{result.get('saved', 0)} 筆；未異動：{result.get('skipped_unchanged', 0)} 筆 / Accounts saved"
                            if result.get("skipped"):
                                msg += "；" + "；".join(result["skipped"])
                            _v30046_set_notice("success", msg)
                        _v104_cache_clear("users", "permissions")
                        st.session_state.pop("v133_users_df", None)
                        _v30046_bump_widget_nonce("paste")
                        st.rerun()
            except Exception as ex:
                st.error(f"貼上資料解析失敗 / Paste parse failed：{ex}")

elif _selected_permission_section == "帳號模組權限 / Account Module Permissions":
    st.subheader("帳號模組權限 / Account × Module Permission Matrix")
    st.info("每個帳號可針對每個模組獨立勾選權限。V104：唯讀先用輕量預覽；按『啟動權限編輯』才建立大型權限矩陣 editor。")
    st.session_state.setdefault("v104_permission_edit_enabled", False)
    pc0, pc1, pc2 = st.columns([1.2, 1.2, 3])
    with pc0:
        if st.button("◇ 啟動權限編輯 / Enable Permission Edit", use_container_width=True, disabled=bool(st.session_state.get("v104_permission_edit_enabled")), key="v104_enable_permission_edit"):
            st.session_state["v104_permission_edit_enabled"] = True
            st.rerun()
    with pc1:
        if st.button("◌ 停止權限編輯 / Lock Permission Edit", use_container_width=True, disabled=not bool(st.session_state.get("v104_permission_edit_enabled")), key="v104_disable_permission_edit"):
            st.session_state["v104_permission_edit_enabled"] = False
            st.rerun()
    with pc2:
        if st.button("⟳ 重新載入權限矩陣 / Reload Permissions", use_container_width=True, key="v104_reload_permission_matrix"):
            _v104_cache_clear("permissions")
            st.session_state["v104_permission_edit_enabled"] = False
            st.rerun()
    permission_edit_enabled = bool(st.session_state.get("v104_permission_edit_enabled"))
    perm_df = pd.DataFrame(_v104_get_permissions_cached())
    if perm_df.empty:
        perm_df = pd.DataFrame(columns=["username", "display_name", "role_code", "module_code", "module_name_zh", "module_name_en"] + ACTION_COLS)
    perm_df = _v30021_ensure_permission_columns(perm_df)
    f1, f2, f3 = st.columns(3)
    with f1:
        user_opts = ["全部 / All"] + sorted(perm_df["username"].dropna().unique().tolist())
        selected_user = st.selectbox("帳號篩選 / Filter Username", user_opts)
    with f2:
        module_opts = ["全部 / All"] + [f"{m['module_code']} {m['module_name_zh']} / {m['module_name_en']}" for m in MODULES]
        selected_module = st.selectbox("模組篩選 / Filter Module", module_opts)
    with f3:
        role_opts = ["全部 / All"] + ROLE_OPTIONS
        selected_role = st.selectbox("角色篩選 / Filter Role", role_opts)
    view_df = perm_df.copy()
    st.session_state.setdefault("v235_permission_editor_rev", 0)
    if selected_user != "全部 / All":
        view_df = view_df[view_df["username"] == selected_user]
    if selected_module != "全部 / All":
        view_df = view_df[view_df["module_code"] == selected_module.split(" ", 1)[0]]
    if selected_role != "全部 / All":
        view_df = view_df[view_df["role_code"] == selected_role]
    permission_draft_key = _v30022_permission_draft_key(selected_user, selected_module, selected_role)
    if permission_edit_enabled:
        view_df = _v30022_permission_get_draft(permission_draft_key, view_df)
    else:
        # 唯讀預覽永遠顯示權威快取，不建立/污染編輯草稿。
        view_df = view_df.copy().reset_index(drop=True)

    st.markdown("#### 快速勾選 / Quick Toggle")
    b1, b2, b3, b4, b5 = st.columns(5)
    with b1:
        if st.button("◈ 可進入全選 / Select View", use_container_width=True, disabled=not permission_edit_enabled):
            view_df["can_view"] = True
            view_df = _v30022_permission_set_draft(permission_draft_key, view_df)
            st.session_state["v235_permission_editor_rev"] = int(st.session_state.get("v235_permission_editor_rev", 0)) + 1
    with b2:
        if st.button("◌ 可進入取消 / Clear View", use_container_width=True, disabled=not permission_edit_enabled):
            view_df["can_view"] = False
            view_df = _v30022_permission_set_draft(permission_draft_key, view_df)
            st.session_state["v235_permission_editor_rev"] = int(st.session_state.get("v235_permission_editor_rev", 0)) + 1
    with b3:
        if st.button("◈ 編輯全選 / Select Edit", use_container_width=True, disabled=not permission_edit_enabled):
            view_df["can_edit"] = True
            view_df = _v30022_permission_set_draft(permission_draft_key, view_df)
            st.session_state["v235_permission_editor_rev"] = int(st.session_state.get("v235_permission_editor_rev", 0)) + 1
    with b4:
        if st.button("⟰ 匯出全選 / Select Export", use_container_width=True, disabled=not permission_edit_enabled):
            view_df["can_export"] = True
            view_df = _v30022_permission_set_draft(permission_draft_key, view_df)
            st.session_state["v235_permission_editor_rev"] = int(st.session_state.get("v235_permission_editor_rev", 0)) + 1
    with b5:
        if st.button("⛨ 管理全選 / Select Manage", use_container_width=True, disabled=not permission_edit_enabled):
            view_df["can_manage"] = True
            view_df = _v30022_permission_set_draft(permission_draft_key, view_df)
            st.session_state["v235_permission_editor_rev"] = int(st.session_state.get("v235_permission_editor_rev", 0)) + 1
    base_cols = ["username", "display_name", "role_code", "module_code", "module_name_zh", "module_name_en"]
    col_cfg = {
        "username": st.column_config.TextColumn("帳號 / Username", disabled=True),
        "display_name": st.column_config.TextColumn("姓名 / Name", disabled=True),
        "role_code": st.column_config.TextColumn("角色 / Role", disabled=True),
        "module_code": st.column_config.TextColumn("模組代碼 / Module Code", disabled=True),
        "module_name_zh": st.column_config.TextColumn("模組中文 / Module Chinese", disabled=True),
        "module_name_en": st.column_config.TextColumn("模組英文 / Module English", disabled=True),
    }
    for key, zh, en in ACTIONS:
        col_cfg[key] = st.column_config.CheckboxColumn(f"{zh} / {en}")
    st.info("V104：權限表改為輕量預覽 + 確認後才套用。勾選權限時不寫 Neon、不重讀、不重算。")
    view_df = _v30021_ensure_permission_columns(view_df)
    if permission_edit_enabled:
        with st.form("permission_editor_commit_form", clear_on_submit=False):
            edited_perm = _v95_raw_data_editor(
                view_df[base_cols + ACTION_COLS],
                key=f"v189_permission_editor_{st.session_state.get('v235_permission_editor_rev', 0)}",
                use_container_width=True,
                hide_index=True,
                height=520,
                column_config=col_cfg,
            )
            submitted_perm = st.form_submit_button("▣ 確認套用並儲存權限 / Apply and Save Permissions", type="primary", use_container_width=True)
    else:
        st.dataframe(view_df[base_cols + ACTION_COLS].head(300), use_container_width=True, hide_index=True, height=420)
        st.caption("唯讀輕量預覽：按『啟動權限編輯』後才建立可編輯權限矩陣。")
        edited_perm = view_df[base_cols + ACTION_COLS].copy()
        submitted_perm = False
    st.markdown("#### 權限摘要預覽 / Permission Summary Preview")
    st.dataframe(_permission_summary(edited_perm), use_container_width=True, hide_index=True)
    if submitted_perm:
        edited_perm = _v30022_permission_set_draft(permission_draft_key, edited_perm)
        saved = save_account_permissions(edited_perm.to_dict("records"))
        saved_n = saved.get("saved", saved) if isinstance(saved, dict) else saved
        skipped_n = saved.get("skipped_unchanged", 0) if isinstance(saved, dict) else 0
        _v104_cache_clear("permissions")
        _v30022_clear_permission_drafts()
        st.session_state["v104_permission_edit_enabled"] = False
        st.success(f"權限已套用並儲存：{saved_n} 筆；未異動略過：{skipped_n} 筆 / Permissions saved")
        st.rerun()

elif _selected_permission_section == "安全設定 / Security":
    st.subheader("安全設定 / Security Settings")
    settings = _v104_get_security_cached()
    idle = int(settings.get("idle_timeout_minutes", "15") or 15)
    with st.form("security_settings_commit_form", clear_on_submit=False):
        new_idle = st.number_input("閒置自動登出分鐘數 / Idle Auto Logout Minutes", min_value=1, max_value=240, value=idle, step=1)
        confirm_after_record = st.checkbox("工時完成後詢問是否繼續記錄 / Ask continue after time record", value=settings.get("ask_continue_after_record", "1") != "0")
        submitted_security = st.form_submit_button("⛨ 確認套用安全設定 / Apply Security Settings", type="primary", use_container_width=True)
    if submitted_security:
        save_security_settings({"idle_timeout_minutes": str(int(new_idle)), "ask_continue_after_record": "1" if confirm_after_record else "0"})
        _v104_cache_clear("security")
        try:
            from services.security_service import set_idle_timeout_minutes
            set_idle_timeout_minutes(int(new_idle))
        except Exception:
            pass
        st.success(f"安全設定已永久儲存：閒置自動登出 {int(new_idle)} 分鐘 / Security settings permanently saved")
        st.rerun()

try:
    _spt_v40_finish_page_event(_SPT_V40_PAGE_TOKEN)
except Exception:
    pass

